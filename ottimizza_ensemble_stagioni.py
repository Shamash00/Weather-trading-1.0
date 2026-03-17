import pandas as pd
import numpy as np
import os
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
storico_path = os.path.join(BASE_DIR, "Temperature Storiche",
                            "Temperature Storiche FM-15 Tutte le Citta 2021-2025 Daily Max.xlsx")
CITIES = ['Ankara','Atlanta','BuenosAires','Chicago','Dallas','Londra','Lucknow',
          'Miami','Monaco','New York','Parigi','SaoPaulo','Seattle','Seoul',
          'Shanghai','Singapore','TelAviv','Tokyo','Toronto','Wellington']
CACHE_DIR = os.path.join(BASE_DIR, "_cache_modelli")
out_path = os.path.join(BASE_DIR, "Ottimizzazione Ensemble Stagioni.xlsx")

SOUTHERN = {'BuenosAires', 'SaoPaulo', 'Wellington'}

green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
blue_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
bold = Font(bold=True)
bold_white = Font(bold=True, color="FFFFFF")

SEASONS_NORTH = {
    "Inverno": [12, 1, 2],
    "Primavera": [3, 4, 5],
    "Estate": [6, 7, 8],
    "Autunno": [9, 10, 11],
}
SEASONS_SOUTH = {
    "Estate": [12, 1, 2],
    "Autunno": [3, 4, 5],
    "Inverno": [6, 7, 8],
    "Primavera": [9, 10, 11],
}
SEASON_ORDER = ["Inverno", "Primavera", "Estate", "Autunno"]

SEAMLESS_FAMILIES = {
    "icon_seamless": ["icon_d2", "icon_eu", "icon_global"],
    "gfs_seamless": ["gfs_hrrr", "gfs_global"],
    "gem_seamless": ["gem_hrdps_continental", "gem_hrdps_west", "gem_regional", "gem_global"],
    "jma_seamless": ["jma_msm", "jma_gsm"],
    "kma_seamless": ["kma_ldps", "kma_gdps"],
    "meteoswiss_icon_seamless": ["meteoswiss_icon_ch1", "meteoswiss_icon_ch2"],
    "meteofrance_seamless": ["meteofrance_arome_france_hd", "meteofrance_arome_france",
                              "meteofrance_arpege_europe", "meteofrance_arpege_world"],
    "ukmo_seamless": ["ukmo_global_deterministic_10km", "ukmo_uk_deterministic_2km"],
    "knmi_seamless": ["knmi_harmonie_arome_europe", "knmi_harmonie_arome_netherlands"],
    "dmi_seamless": ["dmi_harmonie_arome_europe"],
    "metno_seamless": ["metno_nordic"],
}

# Minimo 1 anno di dati (almeno 365 giorni totali per modello) per essere considerato
MIN_DAYS_TOTAL = 365


def compute_daily_max(hourly_df):
    hourly_df["Data"] = hourly_df["time"].dt.date
    prev1 = [c for c in hourly_df.columns if c.startswith("temperature_2m_previous_day1_")]
    prev2 = [c for c in hourly_df.columns if c.startswith("temperature_2m_previous_day2_")]
    agg = {c: "max" for c in prev1 + prev2}
    daily = hourly_df.groupby("Data").agg(agg).reset_index()
    rename = {}
    for c in prev1:
        rename[c] = f"1GG_{c.replace('temperature_2m_previous_day1_', '')}"
    for c in prev2:
        rename[c] = f"2GG_{c.replace('temperature_2m_previous_day2_', '')}"
    daily.rename(columns=rename, inplace=True)
    return daily


def get_season(month, city_name):
    seasons = SEASONS_SOUTH if city_name in SOUTHERN else SEASONS_NORTH
    for season, months in seasons.items():
        if month in months:
            return season
    return "?"


def deduplicate_models(merged, model_cols, prefix):
    to_remove = set()
    for seamless, components in SEAMLESS_FAMILIES.items():
        s_col = f"{prefix}_{seamless}"
        if s_col not in model_cols:
            continue
        for comp in components:
            c_col = f"{prefix}_{comp}"
            if c_col not in model_cols:
                continue
            both_valid = merged[s_col].notna() & merged[c_col].notna()
            if both_valid.sum() == 0:
                continue
            if (merged.loc[both_valid, s_col] == merged.loc[both_valid, c_col]).all():
                to_remove.add(s_col)
                break
    return [c for c in model_cols if c not in to_remove]


def filter_min_1year(merged, model_cols):
    """Keep only models with at least MIN_DAYS_TOTAL valid days across all data."""
    return [c for c in model_cols if merged[c].notna().sum() >= MIN_DAYS_TOTAL]


def optimize_subset(merged_subset, prefix, reg_c_subset, eligible_models):
    """Run ensemble optimization on a subset, using only eligible models."""
    model_cols = [c for c in eligible_models if c.startswith(f"{prefix}_")]
    # Within the subset, still need minimum data
    min_valid = max(len(merged_subset) * 0.10, 5)
    model_cols = [c for c in model_cols if merged_subset[c].notna().sum() >= min_valid]

    if not model_cols:
        return None

    model_info = {}
    for c in model_cols:
        name = c.replace(f"{prefix}_", "")
        stima = ((merged_subset[c].round(0) - 32) * 5 / 9).round(0)
        delta = stima - reg_c_subset
        valid = delta.dropna()
        if len(valid) == 0:
            continue
        verde_pct = (valid.abs() <= 1).sum() / len(valid) * 100
        bias_c = float(valid.mean())
        model_info[name] = {
            "stima": stima,
            "delta": delta,
            "verde": verde_pct,
            "bias_c": bias_c,
            "n": len(valid),
            "mae": float(valid.abs().mean()),
        }

    if not model_info:
        return None

    ranked = sorted(model_info.keys(), key=lambda m: model_info[m]["verde"], reverse=True)
    best_single = ranked[0]
    best_single_v = model_info[best_single]["verde"]

    all_configs = []
    all_configs.append({
        "metodo": "Singolo",
        "n_modelli": 1,
        "modelli": [best_single],
        "verde_c": best_single_v,
        "n_dati": model_info[best_single]["n"],
        "mae": model_info[best_single]["mae"],
        "bc": {},
    })

    max_n = len(ranked)
    for top_n in range(2, max_n + 1):
        top_models = ranked[:top_n]
        stime_raw = pd.DataFrame({m: model_info[m]["stima"] for m in top_models})
        stime_bc = pd.DataFrame()
        for m in top_models:
            bias = model_info[m]["bias_c"]
            stime_bc[m] = model_info[m]["stima"] - round(bias)

        for use_bc, bc_label in [(False, ""), (True, " +BC")]:
            stime = stime_bc if use_bc else stime_raw

            media = stime.mean(axis=1).round(0)
            delta_media = media - reg_c_subset
            valid = delta_media.dropna()
            if len(valid) > 0:
                v = (valid.abs() <= 1).sum() / len(valid) * 100
                bc_dict = {m: round(model_info[m]["bias_c"]) for m in top_models} if use_bc else {}
                all_configs.append({
                    "metodo": f"Media{bc_label}",
                    "n_modelli": top_n,
                    "modelli": list(top_models),
                    "verde_c": v,
                    "n_dati": len(valid),
                    "mae": float(valid.abs().mean()),
                    "bc": bc_dict,
                })

            mediana = stime.median(axis=1).round(0)
            delta_med = mediana - reg_c_subset
            valid_m = delta_med.dropna()
            if len(valid_m) > 0:
                v = (valid_m.abs() <= 1).sum() / len(valid_m) * 100
                bc_dict = {m: round(model_info[m]["bias_c"]) for m in top_models} if use_bc else {}
                all_configs.append({
                    "metodo": f"Mediana{bc_label}",
                    "n_modelli": top_n,
                    "modelli": list(top_models),
                    "verde_c": v,
                    "n_dati": len(valid_m),
                    "mae": float(valid_m.abs().mean()),
                    "bc": bc_dict,
                })

    all_configs.sort(key=lambda x: x["verde_c"], reverse=True)
    best = all_configs[0]

    return {
        "best": best,
        "best_single": best_single,
        "best_single_v": best_single_v,
        "top_configs": all_configs[:10],
        "model_info": model_info,
        "ranked": ranked,
        "all_configs_count": len(all_configs),
    }


def format_bc(bc_dict, modelli):
    if not bc_dict:
        return "Nessuna"
    parts = []
    for m in modelli:
        corr = bc_dict.get(m, 0)
        parts.append(f"{m}: {corr:+d}")
    return ", ".join(parts)


xl_st = pd.ExcelFile(storico_path)

print("Ottimizzazione ensemble per stagione (min 1 anno storico)...")
print("=" * 110)

# city -> season -> { "1GG": opt, "2GG": opt }
all_data = {}

for city_name in CITIES:
    if city_name not in xl_st.sheet_names:
        continue
    cache_file = os.path.join(CACHE_DIR, f"{city_name}.pkl")
    if not os.path.exists(cache_file):
        continue

    raw_df = pd.read_pickle(cache_file)
    raw_df["time"] = pd.to_datetime(raw_df["time"])
    daily = compute_daily_max(raw_df)
    df_st = pd.read_excel(xl_st, sheet_name=city_name)
    df_st["Data"] = pd.to_datetime(df_st["Data"]).dt.date
    merged = pd.merge(daily, df_st[["Data", "Max_Temperatura_F"]], on="Data", how="inner")

    dates = pd.to_datetime(merged["Data"])
    merged["Mese"] = dates.dt.month
    merged["Stagione"] = merged["Mese"].apply(lambda m: get_season(m, city_name))

    reg_c = ((merged["Max_Temperatura_F"] - 32) * 5 / 9).round(0)
    merged["_reg_c"] = reg_c

    # Pre-filter: deduplicate + min 1 year on FULL data
    eligible_1gg = sorted([c for c in merged.columns if c.startswith("1GG_")])
    eligible_1gg = deduplicate_models(merged, eligible_1gg, "1GG")
    eligible_1gg = filter_min_1year(merged, eligible_1gg)

    eligible_2gg = sorted([c for c in merged.columns if c.startswith("2GG_")])
    eligible_2gg = deduplicate_models(merged, eligible_2gg, "2GG")
    eligible_2gg = filter_min_1year(merged, eligible_2gg)

    n_1gg = len(eligible_1gg)
    n_2gg = len(eligible_2gg)

    print(f"\n{city_name} {'(Sud)' if city_name in SOUTHERN else ''}"
          f"  [{n_1gg} modelli 1GG, {n_2gg} modelli 2GG con >= 1 anno]")
    print("-" * 70)

    all_data[city_name] = {}

    for season in SEASON_ORDER:
        mask = merged["Stagione"] == season
        subset = merged[mask].copy()
        reg_c_sub = subset["_reg_c"]

        if len(subset) < 10:
            continue

        all_data[city_name][season] = {}

        for prefix, eligible in [("1GG", eligible_1gg), ("2GG", eligible_2gg)]:
            opt = optimize_subset(subset, prefix, reg_c_sub, eligible)
            if opt:
                all_data[city_name][season][prefix] = opt
                b = opt["best"]
                diff = b["verde_c"] - opt["best_single_v"]
                print(f"  {season:<12} {prefix}  Singolo: {opt['best_single']:<22} "
                      f"{opt['best_single_v']:5.1f}%  |  Ottimo: {b['metodo']} top{b['n_modelli']} "
                      f"{b['verde_c']:5.1f}% ({diff:+.1f}%)  [{len(subset)} gg]")


# ── Scrivi Excel ──
print("\n\nScrittura Excel...")

with pd.ExcelWriter(out_path, engine="openpyxl") as writer:

    # ── Foglio 1: Configurazione Operativa 1GG ──
    print("  Configurazione operativa 1GG...")
    op_rows = []
    for city in CITIES:
        if city not in all_data:
            continue
        for season in SEASON_ORDER:
            if season not in all_data[city] or "1GG" not in all_data[city][season]:
                continue
            opt = all_data[city][season]["1GG"]
            b = opt["best"]
            bc = b.get("bc", {})
            op_rows.append({
                "Citta": city,
                "Stagione": season,
                "Metodo": b["metodo"],
                "N Modelli": b["n_modelli"],
                "Verde%": round(b["verde_c"], 1),
                "MAE": round(b["mae"], 2),
                "N Dati": b["n_dati"],
                "Singolo Migl.": opt["best_single"],
                "V% Singolo": round(opt["best_single_v"], 1),
                "Miglioram.": round(b["verde_c"] - opt["best_single_v"], 1),
                "Modelli": ", ".join(b["modelli"]),
                "Correzioni BC (C)": format_bc(bc, b["modelli"]),
            })

    df_op = pd.DataFrame(op_rows)
    df_op.to_excel(writer, sheet_name="Operativa 1GG", index=False)
    ws = writer.sheets["Operativa 1GG"]
    col_names = list(df_op.columns)
    for col_idx in range(1, len(col_names) + 1):
        ws.cell(row=1, column=col_idx).font = bold_white
        ws.cell(row=1, column=col_idx).fill = blue_fill
    for row_idx in range(2, len(df_op) + 2):
        for col_idx, col_name in enumerate(col_names, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if col_name == "Verde%" and cell.value is not None:
                cell.fill = green_fill
                cell.font = bold
            elif col_name == "V% Singolo" and cell.value is not None:
                cell.fill = green_fill
            elif col_name == "Miglioram." and cell.value is not None:
                cell.fill = green_fill if cell.value > 0 else red_fill
                cell.font = bold
    for col_idx in range(1, len(col_names) + 1):
        max_len = max(len(str(ws.cell(row=r, column=col_idx).value or ""))
                      for r in range(1, len(df_op) + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    # ── Foglio 2: Configurazione Operativa 2GG ──
    print("  Configurazione operativa 2GG...")
    op_rows_2 = []
    for city in CITIES:
        if city not in all_data:
            continue
        for season in SEASON_ORDER:
            if season not in all_data[city] or "2GG" not in all_data[city][season]:
                continue
            opt = all_data[city][season]["2GG"]
            b = opt["best"]
            bc = b.get("bc", {})
            op_rows_2.append({
                "Citta": city,
                "Stagione": season,
                "Metodo": b["metodo"],
                "N Modelli": b["n_modelli"],
                "Verde%": round(b["verde_c"], 1),
                "MAE": round(b["mae"], 2),
                "N Dati": b["n_dati"],
                "Singolo Migl.": opt["best_single"],
                "V% Singolo": round(opt["best_single_v"], 1),
                "Miglioram.": round(b["verde_c"] - opt["best_single_v"], 1),
                "Modelli": ", ".join(b["modelli"]),
                "Correzioni BC (C)": format_bc(bc, b["modelli"]),
            })

    df_op2 = pd.DataFrame(op_rows_2)
    df_op2.to_excel(writer, sheet_name="Operativa 2GG", index=False)
    ws2 = writer.sheets["Operativa 2GG"]
    col_names2 = list(df_op2.columns)
    for col_idx in range(1, len(col_names2) + 1):
        ws2.cell(row=1, column=col_idx).font = bold_white
        ws2.cell(row=1, column=col_idx).fill = blue_fill
    for row_idx in range(2, len(df_op2) + 2):
        for col_idx, col_name in enumerate(col_names2, 1):
            cell = ws2.cell(row=row_idx, column=col_idx)
            if col_name == "Verde%" and cell.value is not None:
                cell.fill = green_fill
                cell.font = bold
            elif col_name == "V% Singolo" and cell.value is not None:
                cell.fill = green_fill
            elif col_name == "Miglioram." and cell.value is not None:
                cell.fill = green_fill if cell.value > 0 else red_fill
                cell.font = bold
    for col_idx in range(1, len(col_names2) + 1):
        max_len = max(len(str(ws2.cell(row=r, column=col_idx).value or ""))
                      for r in range(1, len(df_op2) + 2))
        ws2.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    # ── Foglio 3: Riepilogo V% per stagione (1GG) ──
    print("  Riepilogo V%...")
    riepilogo_rows = []
    for city in CITIES:
        if city not in all_data:
            continue
        row = {"Citta": city}
        for season in SEASON_ORDER:
            if season in all_data[city] and "1GG" in all_data[city][season]:
                opt = all_data[city][season]["1GG"]
                row[f"{season} V%"] = round(opt["best"]["verde_c"], 1)
                row[f"{season} Metodo"] = (f"{opt['best']['metodo']} "
                                           f"top{opt['best']['n_modelli']}")
        vals = [row.get(f"{s} V%") for s in SEASON_ORDER if row.get(f"{s} V%") is not None]
        if vals:
            row["Media"] = round(np.mean(vals), 1)
            row["Range"] = round(max(vals) - min(vals), 1)
            best_idx = np.argmax([row.get(f"{s} V%", 0) for s in SEASON_ORDER])
            worst_idx = np.argmin([row.get(f"{s} V%", 999)
                                   if row.get(f"{s} V%") is not None else 999
                                   for s in SEASON_ORDER])
            row["Migliore"] = SEASON_ORDER[best_idx]
            row["Peggiore"] = SEASON_ORDER[worst_idx]
        riepilogo_rows.append(row)

    df_riep = pd.DataFrame(riepilogo_rows)
    df_riep.to_excel(writer, sheet_name="Riepilogo 1GG", index=False)
    ws_r = writer.sheets["Riepilogo 1GG"]
    for col_idx in range(1, len(df_riep.columns) + 1):
        ws_r.cell(row=1, column=col_idx).font = bold_white
        ws_r.cell(row=1, column=col_idx).fill = blue_fill
    for row_idx in range(2, len(df_riep) + 2):
        for col_idx in range(1, len(df_riep.columns) + 1):
            header = ws_r.cell(row=1, column=col_idx).value
            cell = ws_r.cell(row=row_idx, column=col_idx)
            if header and "V%" in header and cell.value is not None:
                cell.fill = green_fill
                cell.font = bold
            elif header == "Media" and cell.value is not None:
                cell.fill = orange_fill
                cell.font = bold
    for col_idx in range(1, len(df_riep.columns) + 1):
        max_len = max(len(str(ws_r.cell(row=r, column=col_idx).value or ""))
                      for r in range(1, len(df_riep) + 2))
        ws_r.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 30)

    # ── Foglio 4: Riepilogo V% per stagione (2GG) ──
    riepilogo_rows_2 = []
    for city in CITIES:
        if city not in all_data:
            continue
        row = {"Citta": city}
        for season in SEASON_ORDER:
            if season in all_data[city] and "2GG" in all_data[city][season]:
                opt = all_data[city][season]["2GG"]
                row[f"{season} V%"] = round(opt["best"]["verde_c"], 1)
                row[f"{season} Metodo"] = (f"{opt['best']['metodo']} "
                                           f"top{opt['best']['n_modelli']}")
        vals = [row.get(f"{s} V%") for s in SEASON_ORDER if row.get(f"{s} V%") is not None]
        if vals:
            row["Media"] = round(np.mean(vals), 1)
            row["Range"] = round(max(vals) - min(vals), 1)
            best_idx = np.argmax([row.get(f"{s} V%", 0) for s in SEASON_ORDER])
            worst_idx = np.argmin([row.get(f"{s} V%", 999)
                                   if row.get(f"{s} V%") is not None else 999
                                   for s in SEASON_ORDER])
            row["Migliore"] = SEASON_ORDER[best_idx]
            row["Peggiore"] = SEASON_ORDER[worst_idx]
        riepilogo_rows_2.append(row)

    df_riep2 = pd.DataFrame(riepilogo_rows_2)
    df_riep2.to_excel(writer, sheet_name="Riepilogo 2GG", index=False)
    ws_r2 = writer.sheets["Riepilogo 2GG"]
    for col_idx in range(1, len(df_riep2.columns) + 1):
        ws_r2.cell(row=1, column=col_idx).font = bold_white
        ws_r2.cell(row=1, column=col_idx).fill = blue_fill
    for row_idx in range(2, len(df_riep2) + 2):
        for col_idx in range(1, len(df_riep2.columns) + 1):
            header = ws_r2.cell(row=1, column=col_idx).value
            cell = ws_r2.cell(row=row_idx, column=col_idx)
            if header and "V%" in header and cell.value is not None:
                cell.fill = green_fill
                cell.font = bold
            elif header == "Media" and cell.value is not None:
                cell.fill = orange_fill
                cell.font = bold
    for col_idx in range(1, len(df_riep2.columns) + 1):
        max_len = max(len(str(ws_r2.cell(row=r, column=col_idx).value or ""))
                      for r in range(1, len(df_riep2) + 2))
        ws_r2.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 30)

    # ── Foglio 5 & 6: Top 10 per Stagione (1GG e 2GG) ──
    for prefix, sheet_name in [("1GG", "Top10 Stagione 1GG"),
                                ("2GG", "Top10 Stagione 2GG")]:
        print(f"  Top 10 {prefix}...")
        ws_t = writer.book.create_sheet(sheet_name)
        row_t = 1

        for city in CITIES:
            if city not in all_data:
                continue

            cell = ws_t.cell(row=row_t, column=1, value=city)
            cell.font = Font(bold=True, size=14)
            row_t += 1

            for season in SEASON_ORDER:
                if season not in all_data[city]:
                    continue
                if prefix not in all_data[city][season]:
                    continue
                opt = all_data[city][season][prefix]

                cell = ws_t.cell(row=row_t, column=1, value=season)
                cell.font = Font(bold=True, size=11)
                ws_t.cell(row=row_t, column=2,
                          value=f"({opt['all_configs_count']} combinazioni)")
                row_t += 1

                headers = ["#", "Metodo", "N Mod.", "Verde%", "MAE",
                           "N Dati", "Modelli", "Correzioni BC"]
                for j, h in enumerate(headers):
                    c = ws_t.cell(row=row_t, column=j + 1, value=h)
                    c.font = bold_white
                    c.fill = blue_fill
                row_t += 1

                for i, cfg in enumerate(opt["top_configs"]):
                    ws_t.cell(row=row_t, column=1, value=i + 1)
                    ws_t.cell(row=row_t, column=2, value=cfg["metodo"])
                    ws_t.cell(row=row_t, column=3, value=cfg["n_modelli"])
                    vc = ws_t.cell(row=row_t, column=4, value=round(cfg["verde_c"], 1))
                    vc.number_format = "0.0"
                    if i == 0:
                        vc.fill = green_fill
                        vc.font = bold
                    ws_t.cell(row=row_t, column=5,
                              value=round(cfg["mae"], 2)).number_format = "0.00"
                    ws_t.cell(row=row_t, column=6, value=cfg["n_dati"])
                    ws_t.cell(row=row_t, column=7, value=", ".join(cfg["modelli"]))
                    bc = cfg.get("bc", {})
                    ws_t.cell(row=row_t, column=8,
                              value=format_bc(bc, cfg["modelli"]))
                    row_t += 1

                row_t += 1
            row_t += 2

print(f"\nFile salvato: {out_path}")

# ── Riepilogo console ──
print()
print("=" * 110)
print("RIEPILOGO STAGIONALE 1GG (Verde% Ottimo, solo modelli con >= 1 anno storico)")
print("=" * 110)
print(f"{'Citta':<15} {'Inverno':>10} {'Primavera':>10} {'Estate':>10} {'Autunno':>10}  | "
      f"{'Media':>8} {'Range':>8}")
print("-" * 90)

for city in CITIES:
    if city not in all_data:
        continue
    vals = {}
    for s in SEASON_ORDER:
        if s in all_data[city] and "1GG" in all_data[city][s]:
            vals[s] = all_data[city][s]["1GG"]["best"]["verde_c"]
    parts = []
    for s in SEASON_ORDER:
        v = vals.get(s)
        parts.append(f"{v:5.1f}%" if v else "   N/A")
    sv = list(vals.values())
    media = np.mean(sv) if sv else 0
    rng = max(sv) - min(sv) if len(sv) >= 2 else 0
    print(f"  {city:<15} {'  '.join(parts)}  | {media:6.1f}% {rng:6.1f}%")
