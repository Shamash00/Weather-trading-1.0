"""
Legge il foglio Riepilogo dal file 'Confronto Tutti Modelli vs Storiche Daily Max.xlsx'
e crea un nuovo Excel con i top 5 modelli deterministici per Verde% per ogni città.
"""
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
src_path = os.path.join(BASE_DIR, "Confronto Tutti Modelli vs Storiche Daily Max.xlsx")
out_path = os.path.join(BASE_DIR, "Top 5 Modelli per Citta.xlsx")

green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
light_green = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
gold_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
silver_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
bronze_fill = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")
bold = Font(bold=True)
bold_white = Font(bold=True, color="FFFFFF")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

print("Caricamento file sorgente...")
wb_src = load_workbook(src_path, data_only=True)
ws = wb_src["Riepilogo"]

# ── Parse the Riepilogo sheet ──
# Structure: City header (bold, size 14), then "Previsione 1 Giorno" section,
# then "Previsione 2 Giorni" section, with column headers and model rows.

data = {}  # {city: {"1GG": [(model, VF_Tot), ...], "2GG": [...]}}

current_city = None
current_section = None
header_row_cols = {}  # column index -> header name

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
    cell0 = row[0]
    val0 = cell0.value

    if val0 is None:
        continue

    # Detect city header (bold, size >= 14)
    if cell0.font and cell0.font.size and cell0.font.size >= 14:
        current_city = str(val0).strip()
        data[current_city] = {"1GG": [], "2GG": []}
        current_section = None
        header_row_cols = {}
        continue

    # Detect section header
    if val0 == "Previsione 1 Giorno":
        current_section = "1GG"
        header_row_cols = {}
        continue
    elif val0 == "Previsione 2 Giorni":
        current_section = "2GG"
        header_row_cols = {}
        continue

    # Detect column header row
    if val0 == "Modello" and current_city and current_section:
        header_row_cols = {}
        for i, c in enumerate(row):
            if c.value:
                header_row_cols[i] = str(c.value)
        continue

    # Data row: model name in col 0
    if current_city and current_section and header_row_cols and val0 and val0 != "Modello":
        model_name = str(val0).strip()
        # Find V%F Tot column
        vf_tot_idx = None
        for idx, hname in header_row_cols.items():
            if hname == "V%F Tot":
                vf_tot_idx = idx
                break

        vf_tot = None
        if vf_tot_idx is not None and vf_tot_idx < len(row):
            vf_tot = row[vf_tot_idx].value

        # Also get V%C Tot as fallback
        vc_tot_idx = None
        for idx, hname in header_row_cols.items():
            if hname == "V%C Tot":
                vc_tot_idx = idx
                break

        vc_tot = None
        if vc_tot_idx is not None and vc_tot_idx < len(row):
            vc_tot = row[vc_tot_idx].value

        if vf_tot is not None and isinstance(vf_tot, (int, float)):
            data[current_city][current_section].append((model_name, round(vf_tot, 1), round(vc_tot, 1) if vc_tot else None))
        elif vc_tot is not None and isinstance(vc_tot, (int, float)):
            data[current_city][current_section].append((model_name, None, round(vc_tot, 1)))

# Sort all models by VF_Tot descending
for city in data:
    for section in ["1GG", "2GG"]:
        models = data[city][section]
        models.sort(key=lambda x: (x[1] if x[1] is not None else 0), reverse=True)

# ── Create output Excel ──
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()

def write_top_sheet(wb, sheet_title, section, top_n, is_first=False):
    if is_first:
        ws_out = wb.active
        ws_out.title = sheet_title
    else:
        ws_out = wb.create_sheet(sheet_title)

    # Headers
    headers = ["Città"]
    for rank in range(1, top_n + 1):
        headers += [f"{rank}° Modello", "V%F", "V%C"]

    for j, h in enumerate(headers):
        c = ws_out.cell(row=1, column=j+1, value=h)
        c.font = bold_white
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border

    row_num = 2
    for city in sorted(data.keys()):
        models = data[city][section][:top_n]
        if not models:
            continue

        c = ws_out.cell(row=row_num, column=1, value=city)
        c.font = bold
        c.border = thin_border

        for i, (model, vf, vc) in enumerate(models):
            base_col = 2 + i * 3

            # Model name
            cm = ws_out.cell(row=row_num, column=base_col, value=model)
            cm.border = thin_border
            if i == 0:
                cm.fill = gold_fill
            elif i == 1:
                cm.fill = silver_fill
            elif i == 2:
                cm.fill = bronze_fill

            # V%F
            cv = ws_out.cell(row=row_num, column=base_col + 1, value=vf)
            cv.number_format = '0.0'
            cv.fill = green_fill
            cv.font = bold
            cv.alignment = Alignment(horizontal="center")
            cv.border = thin_border

            # V%C
            cc = ws_out.cell(row=row_num, column=base_col + 2, value=vc)
            cc.number_format = '0.0'
            cc.fill = light_green
            cc.alignment = Alignment(horizontal="center")
            cc.border = thin_border

        row_num += 1

    # Auto-width columns
    for col in range(1, len(headers) + 1):
        max_len = 0
        for r in range(1, row_num + 1):
            val = ws_out.cell(row=r, column=col).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws_out.column_dimensions[get_column_letter(col)].width = max(max_len + 3, 8)

    # Freeze first row
    ws_out.freeze_panes = "A2"

write_top_sheet(wb, "Top 5 - Previsione 1 Giorno", "1GG", 5, is_first=True)
write_top_sheet(wb, "Top 10 - Previsione 1 Giorno", "1GG", 10)

wb.save(out_path)
print(f"\nFile salvato: {out_path}")
print(f"Città trovate: {len(data)}")
for city in sorted(data.keys()):
    m1 = data[city]["1GG"]
    if m1:
        print(f"  {city}: Top 1GG = {m1[0][0]} ({m1[0][1]}%F)")
