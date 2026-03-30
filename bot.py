"""
Polymarket Weather Trading Bot - Daemon 24/7

Il bot gira in background e, per ogni mercato di temperatura nelle 20 citta:

1. DOPPIO SNAPSHOT (16:00 e 20:00 locali, giorno prima della risoluzione):
   - Alle 16:00: snapshot "early" (include run 06z di tutti i modelli)
   - Alle 20:00: snapshot "late"  (include run 12z ECMWF IFS, piu accurato)
   - Per ciascuno: cattura odds Polymarket + dati ensemble Open-Meteo
   - Permette di confrontare come cambiano le previsioni tra i due orari

2. RISOLUZIONE (dopo chiusura del mercato):
   - Controlla Polymarket per il bucket vincente
   - Registra il risultato nel file Excel

Uso:
    python bot.py                    # avvia daemon (snapshot alle 16:00 e 20:00 locali)
    python bot.py --hours 18,22      # snapshot alle 18:00 e 22:00 locali
    python bot.py --once             # esegui un solo ciclo e esci
    python bot.py --status           # mostra stato corrente e esci
    python bot.py --days 3           # considera mercati fino a 3 giorni avanti (default: 3)
"""

import requests
import json
import time as time_mod
import logging
import argparse
import numpy as np
import os
import re
import subprocess
import sys
import threading
import traceback
from datetime import datetime, timezone, timedelta, date
from http.server import HTTPServer, BaseHTTPRequestHandler
from pathlib import Path
from zoneinfo import ZoneInfo
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════════════
# CONFIG
# ═══════════════════════════════════════════════════════════════════════════════

# Snapshot: lista di tuple (ora, minuto, mode)
# mode="utc"   -> orario UTC fisso, uguale per tutte le citta
# mode="local"  -> orario locale della citta target
SNAPSHOT_HOURS = [
    (16, 10, "local"),  # 16:10 locale — per comparazione con dati storici + ottimo per Americhe
    (17, 10, "utc"),    # 17:10 UTC — 9/15 modelli 12Z (ottimo per Asia/Europa/Oceania)
    (20, 10, "utc"),    # 20:10 UTC — 15/15 modelli 12Z, benchmark
]
CHECK_INTERVAL = 60         # Secondi tra ogni check del loop principale
EVENT_REFRESH_INTERVAL = 1800  # Refresh lista mercati ogni 30 minuti
SNAPSHOT_WINDOW = 7200      # Finestra di cattura: se il bot parte in ritardo, cattura
                            # se siamo entro 2 ore dopo l'orario previsto
DAYS_AHEAD = 3              # Considera mercati fino a N giorni avanti

GAMMA_API = "https://gamma-api.polymarket.com"
ENSEMBLE_API = "https://ensemble-api.open-meteo.com/v1/ensemble"
DETERMINISTIC_API = "https://api.open-meteo.com/v1/forecast"

BASE_DIR = Path(__file__).parent
DATA_DIR = Path(os.environ.get("DATA_DIR", str(BASE_DIR)))
DATA_DIR.mkdir(parents=True, exist_ok=True)

# Se DATA_DIR diversa da BASE_DIR (es. volume Railway), copia i file dal repo
# al volume al primo avvio (cosi' non si parte da zero)
import shutil
if str(DATA_DIR) != str(BASE_DIR):
    for _fname in ["dati_combinati.xlsx", "dati_meteo.xlsx", "bot_state.json"]:
        _src = BASE_DIR / _fname
        _dst = DATA_DIR / _fname
        if _src.exists() and not _dst.exists():
            shutil.copy2(_src, _dst)

STATE_FILE = DATA_DIR / "bot_state.json"
EXCEL_FILE = DATA_DIR / "dati_meteo.xlsx"
EXCEL_COMBINED = DATA_DIR / "dati_combinati.xlsx"
OPTIMIZATION_FILE = BASE_DIR / "Ottimizzazione Ensemble Stagioni.xlsx"
TOP_MODELS_FILE = BASE_DIR / "Top 5 Modelli per Citta.xlsx"
TOP_MODELS_FILE_V2 = BASE_DIR / "Top Modelli Deterministici per Citta.xlsx"
MODEL_STATS_FILE = BASE_DIR / "model_error_stats.pkl"
LOG_FILE = DATA_DIR / "bot.log"

GIT_AUTO_PUSH = os.environ.get("GIT_AUTO_PUSH", "1") == "1"
GIT_PUSH_FILES = ["dati_combinati.xlsx", "dati_meteo.xlsx", "bot_state.json"]

# ── Tutti i modelli deterministici per il mixture model ──────────────────────

ALL_DETERMINISTIC_MODELS = [
    "best_match",
    "bom_access_global",
    "ecmwf_aifs025_single", "ecmwf_ifs025",
    "gem_global", "gem_regional",
    "gfs_global", "gfs_graphcast025", "gfs_hrrr",
    "icon_d2", "icon_eu", "icon_global",
    "jma_gsm",
    "kma_gdps",
    "knmi_harmonie_arome_europe", "knmi_harmonie_arome_netherlands",
    "dmi_harmonie_arome_europe",
    "meteofrance_arome_france", "meteofrance_arome_france_hd",
    "meteofrance_arpege_europe", "meteofrance_arpege_world",
    "metno_seamless",
    "ukmo_global_deterministic_10km", "ukmo_uk_deterministic_2km",
]

# ── Coordinate stazioni aeroportuali + timezone ──────────────────────────────

CITY_DATA = {
    "Ankara":       {"lat": 40.128082,  "lon": 32.995083,   "tz": "Europe/Istanbul"},
    "Atlanta":      {"lat": 33.62972,   "lon": -84.44224,   "tz": "America/New_York"},
    "Buenos Aires": {"lat": -34.822222, "lon": -58.535833,  "tz": "America/Argentina/Buenos_Aires"},
    "Chicago":      {"lat": 41.96019,   "lon": -87.93162,   "tz": "America/Chicago"},
    "Dallas":       {"lat": 32.8519,    "lon": -96.8555,    "tz": "America/Chicago"},
    "London":       {"lat": 51.505278,  "lon": 0.055278,    "tz": "Europe/London"},
    "Lucknow":      {"lat": 26.7606,    "lon": 80.8893,     "tz": "Asia/Kolkata"},
    "Miami":        {"lat": 25.78805,   "lon": -80.31694,   "tz": "America/New_York"},
    "Monaco":       {"lat": 48.353783,  "lon": 11.786086,   "tz": "Europe/Berlin"},
    "Munich":       {"lat": 48.353783,  "lon": 11.786086,   "tz": "Europe/Berlin"},
    "New York":     {"lat": 40.77945,   "lon": -73.88027,   "tz": "America/New_York"},
    "NYC":          {"lat": 40.77945,   "lon": -73.88027,   "tz": "America/New_York"},
    "Paris":        {"lat": 49.012779,  "lon": 2.55,        "tz": "Europe/Paris"},
    "Sao Paulo":    {"lat": -23.432075, "lon": -46.469511,  "tz": "America/Sao_Paulo"},
    "São Paulo":    {"lat": -23.432075, "lon": -46.469511,  "tz": "America/Sao_Paulo"},
    "Seattle":      {"lat": 47.4444,    "lon": -122.3138,   "tz": "America/Los_Angeles"},
    "Seoul":        {"lat": 37.469075,  "lon": 126.450517,  "tz": "Asia/Seoul"},
    "Shanghai":     {"lat": 31.143378,  "lon": 121.805214,  "tz": "Asia/Shanghai"},
    "Singapore":    {"lat": 1.350189,   "lon": 103.994433,  "tz": "Asia/Singapore"},
    "Tel Aviv":     {"lat": 32.011389,  "lon": 34.886667,   "tz": "Asia/Jerusalem"},
    "Tokyo":        {"lat": 35.552258,  "lon": 139.779694,  "tz": "Asia/Tokyo"},
    "Toronto":      {"lat": 43.677223,  "lon": -79.630556,  "tz": "America/Toronto"},
    "Wellington":   {"lat": -41.333333, "lon": 174.8,       "tz": "Pacific/Auckland"},
    # Nuove città
    "Austin":       {"lat": 30.18311,   "lon": -97.67989,   "tz": "America/Chicago"},
    "Beijing":      {"lat": 40.080111,  "lon": 116.584556,  "tz": "Asia/Shanghai"},
    "Chengdu":      {"lat": 30.578528,  "lon": 103.947086,  "tz": "Asia/Shanghai"},
    "Chongqing":    {"lat": 29.719217,  "lon": 106.641678,  "tz": "Asia/Shanghai"},
    "Denver":       {"lat": 39.84657,   "lon": -104.65623,  "tz": "America/Denver"},
    "Houston":      {"lat": 29.64586,   "lon": -95.28212,   "tz": "America/Chicago"},
    "Madrid":       {"lat": 40.493556,  "lon": -3.566764,   "tz": "Europe/Madrid"},
    "Milan":        {"lat": 45.630606,  "lon": 8.728111,    "tz": "Europe/Rome"},
    "Milano":       {"lat": 45.630606,  "lon": 8.728111,    "tz": "Europe/Rome"},
    "San Francisco":{"lat": 37.61962,   "lon": -122.36562,  "tz": "America/Los_Angeles"},
    "Shenzhen":     {"lat": 22.639258,  "lon": 113.810664,  "tz": "Asia/Shanghai"},
    "Taipei":       {"lat": 25.077731,  "lon": 121.232822,  "tz": "Asia/Taipei"},
    "Warsaw":       {"lat": 52.16575,   "lon": 20.967122,   "tz": "Europe/Warsaw"},
    "Wuhan":        {"lat": 30.783758,  "lon": 114.2081,    "tz": "Asia/Shanghai"},
}

# ── Mapping nomi citta: bot -> file ottimizzazione ──────────────────────────

CITY_NAME_TO_OPT = {
    "Ankara": "Ankara", "Atlanta": "Atlanta", "Buenos Aires": "BuenosAires",
    "Chicago": "Chicago", "Dallas": "Dallas", "London": "Londra",
    "Lucknow": "Lucknow", "Miami": "Miami", "Munich": "Monaco",
    "Monaco": "Monaco", "New York": "New York", "NYC": "New York",
    "Paris": "Parigi", "Sao Paulo": "SaoPaulo", "São Paulo": "SaoPaulo",
    "Seattle": "Seattle", "Seoul": "Seoul", "Shanghai": "Shanghai",
    "Singapore": "Singapore", "Tel Aviv": "TelAviv", "Tokyo": "Tokyo",
    "Toronto": "Toronto", "Wellington": "Wellington",
    # Nuove città
    "Austin": "Austin", "Beijing": "Beijing", "Chengdu": "Chengdu",
    "Chongqing": "Chongqing", "Denver": "Denver", "Houston": "Houston",
    "Madrid": "Madrid", "Milan": "Milano", "Milano": "Milano",
    "San Francisco": "SanFrancisco", "Shenzhen": "Shenzhen",
    "Taipei": "Taipei", "Warsaw": "Warsaw", "Wuhan": "Wuhan",
}

# ── Citta emisfero sud (stagioni invertite) ─────────────────────────────────

SOUTHERN_HEMISPHERE_CITIES = {"BuenosAires", "SaoPaulo", "Wellington"}

def get_season(month: int, opt_city: str) -> str:
    """Ritorna la stagione (Inverno/Primavera/Estate/Autunno) per il mese dato."""
    if opt_city in SOUTHERN_HEMISPHERE_CITIES:
        # Emisfero sud: stagioni invertite
        if month in (12, 1, 2):
            return "Estate"
        elif month in (3, 4, 5):
            return "Autunno"
        elif month in (6, 7, 8):
            return "Inverno"
        else:
            return "Primavera"
    else:
        # Emisfero nord
        if month in (12, 1, 2):
            return "Inverno"
        elif month in (3, 4, 5):
            return "Primavera"
        elif month in (6, 7, 8):
            return "Estate"
        else:
            return "Autunno"


# ── Caricamento config ottimizzazione deterministici ────────────────────────

_det_config_cache = None

def load_deterministic_config() -> dict:
    """
    Carica la config dal foglio 'Riepilogo' del file ottimizzazione (1GG).
    Ritorna dict: {(opt_city, season): {method, models, corrections, mae, verde}}.
    """
    global _det_config_cache
    if _det_config_cache is not None:
        return _det_config_cache

    # Usa la stessa logica di load_deterministic_config_v2
    _det_config_cache = load_deterministic_config_v2()
    return _det_config_cache


# ── Caricamento config deterministici v2 (da Riepilogo aggiornato) ───────

_det_config_v2_cache = None

def load_deterministic_config_v2() -> dict:
    """
    Carica la config dal foglio 'Riepilogo' del file ottimizzazione stagioni.
    Filtra per Orizzonte == '1GG'.
    Ritorna dict: {(opt_city, season): {method, models, corrections, mae, verde}}.
    """
    global _det_config_v2_cache
    if _det_config_v2_cache is not None:
        return _det_config_v2_cache

    if not OPTIMIZATION_FILE.exists():
        log.warning(f"File ottimizzazione non trovato: {OPTIMIZATION_FILE}")
        _det_config_v2_cache = {}
        return _det_config_v2_cache

    wb = openpyxl.load_workbook(OPTIMIZATION_FILE, read_only=True, data_only=True)
    if "Riepilogo" not in wb.sheetnames:
        log.warning("Foglio 'Riepilogo' non trovato in ottimizzazione")
        _det_config_v2_cache = {}
        wb.close()
        return _det_config_v2_cache

    ws = wb["Riepilogo"]

    config = {}
    for r in range(2, ws.max_row + 1):
        city = ws.cell(r, 1).value
        if not city:
            continue

        orizzonte = ws.cell(r, 4).value
        if orizzonte != "1GG":
            continue

        season = ws.cell(r, 3).value
        method = ws.cell(r, 5).value
        verde = ws.cell(r, 7).value or 0
        mae = ws.cell(r, 8).value or 1.0
        models_str = ws.cell(r, 13).value or ""

        models = [m.strip() for m in models_str.split(",")
                  if m.strip() and m.strip() != "cma_grapes_global"]

        config[(str(city).strip(), str(season).strip())] = {
            "method": method,
            "models": models,
            "corrections": {},
            "mae": float(mae),
            "verde": float(verde),
        }

    wb.close()
    _det_config_v2_cache = config
    log.info(f"Config deterministici v2 caricata: {len(config)} combinazioni citta/stagione")
    return config


# ── Caricamento statistiche errore per mixture model ─────────────────────────

_model_stats_cache = None

def load_model_stats() -> dict:
    """
    Carica model_error_stats.pkl (pre-computato da precompute_model_stats.py).
    Contiene: stats, iso_base, iso_enhanced per il mixture model.
    """
    global _model_stats_cache
    if _model_stats_cache is not None:
        return _model_stats_cache

    if not MODEL_STATS_FILE.exists():
        log.warning(f"Model stats non trovato: {MODEL_STATS_FILE} - mixture model disabilitato")
        _model_stats_cache = {}
        return _model_stats_cache

    import pickle
    with open(MODEL_STATS_FILE, "rb") as f:
        _model_stats_cache = pickle.load(f)

    n_entries = len(_model_stats_cache.get("stats", {}))
    log.info(f"Model stats caricati: {n_entries} combinazioni modello/citta/stagione")
    return _model_stats_cache


# ── Caricamento top N modelli per citta (da Top 5 Modelli per Citta.xlsx) ────

_top_models_cache = None

def load_top_models() -> dict:
    """
    Carica i top modelli per citta. Usa il file aggiornato (v2).
    Ritorna dict: {opt_city: {"top5": [model1, ...], "top10": [model1, ...]}}.
    """
    global _top_models_cache
    if _top_models_cache is not None:
        return _top_models_cache

    # Usa la stessa logica di load_top_models_v2
    _top_models_cache = load_top_models_v2()
    return _top_models_cache


# ── Caricamento top modelli v2 (classifica aggiornata) ───────────────────

_top_models_v2_cache = None

def load_top_models_v2() -> dict:
    """
    Carica i top modelli per citta dal file 'Top Modelli Deterministici per Citta.xlsx'.
    Ritorna dict: {opt_city: {"top5": [model1, ...], "top10": [model1, ...]}}.
    """
    global _top_models_v2_cache
    if _top_models_v2_cache is not None:
        return _top_models_v2_cache

    if not TOP_MODELS_FILE_V2.exists():
        log.warning(f"File top modelli v2 non trovato: {TOP_MODELS_FILE_V2}")
        _top_models_v2_cache = {}
        return _top_models_v2_cache

    wb = openpyxl.load_workbook(TOP_MODELS_FILE_V2, read_only=True, data_only=True)
    result = {}

    # Sheet names contengono °C - cerca per pattern
    sheet_map = {}
    for sn in wb.sheetnames:
        if "Top 5" in sn and "1GG" in sn:
            sheet_map["top5"] = sn
        elif "Top 10" in sn and "1GG" in sn:
            sheet_map["top10"] = sn

    for key, sheet_name in sheet_map.items():
        ws = wb[sheet_name]
        for r in range(2, ws.max_row + 1):
            city = ws.cell(r, 1).value
            if not city:
                continue
            city = str(city).strip()
            if city not in result:
                result[city] = {"top5": [], "top10": []}

            models = []
            col = 2
            while True:
                model = ws.cell(r, col).value
                if not model:
                    break
                models.append(str(model).strip())
                col += 3  # salta V%C e V%F

            result[city][key] = models

    wb.close()
    total = len(result)
    log.info(f"Top modelli v2 caricati: {total} citta")
    _top_models_v2_cache = result
    return _top_models_v2_cache


# ── Modelli ensemble ─────────────────────────────────────────────────────────
# Run times UTC e ritardo disponibilita approssimativo (ore dopo il run):
# ECMWF IFS ENS:   00z, 12z         | +6-8h  | 51 membri
# ECMWF AIFS:      00z, 06z, 12z, 18z | +4-5h | 51 membri
# ICON Global EPS: 00z, 06z, 12z, 18z | +4-5h | 40 membri
# ICON EU EPS:     00z, 06z, 12z, 18z | +3-4h | 40 membri
# ICON D2 EPS:     00z, 03z, ..., 21z | +2-3h | 20 membri
# GEFS:            00z, 06z, 12z, 18z | +5-6h | 31 membri
# GEM ENS:         00z, 12z          | +7h    | 21 membri
# UKMO:            00z, 06z, 12z, 18z | +6-8h | 18 membri
# BOM ACCESS:      00z, 12z          | +6-8h | 18 membri

ENSEMBLE_MODELS = [
    "ecmwf_ifs025_ensemble",
    "ecmwf_aifs025_ensemble",
    "icon_global_eps",
    "icon_eu_eps",
    "icon_d2_eps",
    "ncep_gefs025",
    "ncep_gefs05",
    "ncep_aigefs025",
    "gem_global_ensemble",
    "bom_access_global_ensemble",
    "ukmo_global_ensemble_20km",
    "ukmo_uk_ensemble_2km",
]

# ═══════════════════════════════════════════════════════════════════════════════
# LOGGING
# ═══════════════════════════════════════════════════════════════════════════════

def setup_logging():
    logger = logging.getLogger("bot")
    logger.setLevel(logging.INFO)
    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", datefmt="%Y-%m-%d %H:%M:%S")

    # Console
    ch = logging.StreamHandler(sys.stdout)
    ch.setFormatter(fmt)
    logger.addHandler(ch)

    # File
    fh = logging.FileHandler(LOG_FILE, encoding="utf-8")
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    return logger

log = setup_logging()


# ═══════════════════════════════════════════════════════════════════════════════
# STATE MANAGEMENT
# ═══════════════════════════════════════════════════════════════════════════════

def load_state() -> dict:
    if STATE_FILE.exists():
        with open(STATE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"snapshots_done": {}, "resolutions_done": {}, "last_event_refresh": None}


def save_state(state: dict):
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(state, f, indent=2, ensure_ascii=False, default=str)


# ═══════════════════════════════════════════════════════════════════════════════
# POLYMARKET - FETCH & PARSE
# ═══════════════════════════════════════════════════════════════════════════════

def fetch_temp_events(closed: bool = False) -> list[dict]:
    """Scarica eventi temperatura da Polymarket con paginazione."""
    all_events = []
    offset = 0
    try:
        while True:
            resp = requests.get(
                f"{GAMMA_API}/events",
                params={"tag_slug": "weather", "closed": str(closed).lower(),
                        "limit": 200, "offset": offset},
                timeout=20,
            )
            resp.raise_for_status()
            batch = resp.json()
            temp_events = [e for e in batch if "highest temperature" in e.get("title", "").lower()]
            all_events.extend(temp_events)
            if len(batch) < 200:
                break
            offset += 200
        return all_events
    except requests.RequestException as e:
        log.warning(f"Gamma API error: {e}")
        return all_events


def city_from_title(title: str) -> str:
    m = re.search(r"Highest temperature in (.+?) on ", title, re.IGNORECASE)
    return m.group(1).strip() if m else ""


def date_from_title(title: str, end_date_hint: str | None = None) -> str | None:
    """
    Estrae data dal titolo (es. 'on March 19' -> '2026-03-19').
    Se end_date_hint (endDate ISO dall'API) e' disponibile, usa quell'anno.
    """
    m = re.search(r"on (\w+ \d+)", title, re.IGNORECASE)
    if not m:
        return None

    # Se abbiamo endDate dall'API, usa quell'anno
    if end_date_hint:
        try:
            hint_year = int(end_date_hint[:4])
            dt = datetime.strptime(f"{m.group(1)} {hint_year}", "%B %d %Y")
            return dt.strftime("%Y-%m-%d")
        except (ValueError, IndexError):
            pass

    now = datetime.now()
    for year in [now.year, now.year + 1, now.year - 1]:
        try:
            dt = datetime.strptime(f"{m.group(1)} {year}", "%B %d %Y")
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def parse_end_date(s: str) -> datetime | None:
    if not s:
        return None
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except ValueError:
        return None


def temp_sort_key(label: str) -> float:
    nums = re.findall(r"-?\d+\.?\d*", label)
    if not nums:
        return 0.0
    val = float(nums[0])
    lbl = label.lower()
    if "or higher" in lbl or "or above" in lbl:
        return val + 0.5
    if "or below" in lbl or "or lower" in lbl:
        return val - 0.5
    return val


def extract_bucket_temp(label: str) -> int | None:
    nums = re.findall(r"-?\d+", label)
    return int(nums[0]) if nums else None


def parse_bucket(label: str) -> dict | None:
    """
    Parsa un bucket Polymarket in struttura.
    Gestisce sia Celsius singoli ("11C") che Fahrenheit a range ("72-73F").
    Ritorna {unit, low, high, is_lower_bound, is_upper_bound}.
    """
    lbl = label.lower()

    # Detect unit: cerca °F, ºF, o un F subito dopo un numero
    unit = "F" if re.search(r'[°ºo]F|°f|\d\s*F\b', label) else "C"
    is_lower = "or below" in lbl or "or lower" in lbl or "less than" in lbl
    is_upper = "or higher" in lbl or "or above" in lbl

    # Range "72-73" o "72–73" (il trattino NON e' segno negativo)
    range_match = re.search(r'(\d+)\s*[-–]\s*(\d+)', label)

    # Numeri singoli (per temp negative o singoli bucket)
    nums = re.findall(r"-?\d+", label)
    if not nums:
        return None

    if is_lower:
        return {"unit": unit, "low": None, "high": int(nums[0]),
                "is_lower": True, "is_upper": False, "label": label}
    elif is_upper:
        return {"unit": unit, "low": int(nums[0]), "high": None,
                "is_lower": False, "is_upper": True, "label": label}
    elif range_match:
        return {"unit": unit, "low": int(range_match.group(1)),
                "high": int(range_match.group(2)),
                "is_lower": False, "is_upper": False, "label": label}
    else:
        val = int(nums[0])
        return {"unit": unit, "low": val, "high": val,
                "is_lower": False, "is_upper": False, "label": label}


def calc_bucket_prob_from_celsius(values_c: np.ndarray, bucket: dict) -> float:
    """
    Calcola probabilita che i valori ensemble (in Celsius) cadano in un bucket.
    Converte in Fahrenheit se necessario, arrotonda al grado intero,
    poi verifica appartenenza al bucket.
    """
    if len(values_c) == 0:
        return 0.0

    if bucket["unit"] == "F":
        values = values_c * 9.0 / 5.0 + 32.0
    else:
        values = values_c

    rounded = np.round(values).astype(int)
    total = len(rounded)

    if bucket["is_lower"]:
        return int(np.sum(rounded <= bucket["high"])) / total
    elif bucket["is_upper"]:
        return int(np.sum(rounded >= bucket["low"])) / total
    else:
        mask = (rounded >= bucket["low"]) & (rounded <= bucket["high"])
        return int(np.sum(mask)) / total


IGNORED_CITIES = {"Hong Kong", "Los Angeles", "Istanbul"}

def match_city(city_name: str) -> dict | None:
    if city_name in IGNORED_CITIES:
        return None
    if city_name in CITY_DATA:
        return CITY_DATA[city_name]
    for key, data in CITY_DATA.items():
        if key.lower() == city_name.lower():
            return data
    for key, data in CITY_DATA.items():
        if key.lower() in city_name.lower() or city_name.lower() in key.lower():
            return data
    return None


def parse_markets_from_events(events: list[dict], days_ahead: int) -> list[dict]:
    """
    Dalle eventi Polymarket, estrai una lista di mercati strutturati.
    Ogni mercato = {city, target_date, end_dt, buckets: [{label, prob, temp}]}
    """
    now = datetime.now(timezone.utc)
    cutoff = now + timedelta(days=days_ahead)
    markets = {}

    for event in events:
        title = event.get("title", "")
        city = city_from_title(title)
        event_end = event.get("endDate", "")
        target_date = date_from_title(title, end_date_hint=event_end)

        if not city or not target_date:
            continue

        for m in event.get("markets") or []:
            end = parse_end_date(m.get("endDate") or event_end)
            if not end:
                continue
            # Includi solo mercati futuri (non ancora chiusi) entro il cutoff
            if end < now or end > cutoff:
                continue

            label = m.get("groupItemTitle") or m.get("question", "")
            prob_raw = m.get("bestAsk") or m.get("lastTradePrice") or 0
            try:
                prob = float(prob_raw)
            except (ValueError, TypeError):
                prob = 0.0

            parsed = parse_bucket(label)
            bucket_temp = extract_bucket_temp(label)
            key = f"{city}_{target_date}"

            if key not in markets:
                markets[key] = {
                    "city": city,
                    "target_date": target_date,
                    "end_dt": end,
                    "event_slug": event.get("slug", ""),
                    "buckets": [],
                }

            markets[key]["buckets"].append({
                "label": label,
                "prob": prob,
                "temp": bucket_temp,
                "parsed": parsed,
            })

    # Ordina bucket per temperatura
    for key in markets:
        markets[key]["buckets"].sort(key=lambda x: temp_sort_key(x["label"]))

    return list(markets.values())


RESOLUTION_START_DATE = "2026-03-27"  # Registra risoluzioni solo da questa data

def check_resolutions(state: dict) -> list[dict]:
    """Controlla mercati (aperti e chiusi) per trovare risoluzioni delle 20 citta.
    I mercati 2026 restano closed=False nell'API anche dopo la risoluzione,
    quindi cerchiamo in TUTTI gli eventi quelli con outcomePrices ~1.0."""
    # Cerca sia negli eventi aperti che chiusi
    open_events = fetch_temp_events(closed=False)
    closed_events = fetch_temp_events(closed=True)
    all_events = open_events + closed_events
    results = []
    today = date.today().isoformat()

    for event in all_events:
        title = event.get("title", "")
        city = city_from_title(title)
        event_end = event.get("endDate", "")
        target_date = date_from_title(title, end_date_hint=event_end)

        if not city or not target_date:
            continue

        # Solo citta nelle nostre 20
        if not match_city(city):
            continue

        # Solo date passate (la temperatura deve essersi gia verificata)
        if target_date > today:
            continue

        # Solo date recenti (da RESOLUTION_START_DATE in poi)
        if target_date < RESOLUTION_START_DATE:
            continue

        key = f"{city}_{target_date}"

        # Gia registrato?
        if key in state.get("resolutions_done", {}):
            continue

        # Trova il bucket vincente usando outcomePrices (non lastTradePrice)
        # outcomePrices = ["yes_price", "no_price"] — il vincente ha yes_price ~ 1.0
        winner = None
        for m in event.get("markets") or []:
            try:
                outcome_prices = json.loads(m.get("outcomePrices", "[]"))
                if outcome_prices and float(outcome_prices[0]) >= 0.995:
                    winner = m.get("groupItemTitle") or m.get("question", "")
                    break
            except (json.JSONDecodeError, ValueError, IndexError):
                continue

        if winner:
            results.append({
                "city": city,
                "target_date": target_date,
                "key": key,
                "winner": winner,
                "winner_temp": extract_bucket_temp(winner),
            })

    return results



# Correzioni manuali per risoluzioni sbagliate non piu' verificabili via API.
# Formato: {"citta_data": "bucket vincente corretto"}
# Rimuovere le voci dopo che sono state applicate con successo.
RESOLUTION_OVERRIDES = {
    "Wellington_2026-03-23": "19°C",
}


def recheck_past_resolutions(state: dict):
    """Ri-verifica tutte le risoluzioni passate con la soglia attuale (0.995).
    Rimuove dallo state e dall'Excel quelle che non sono piu' valide,
    cosi' il ciclo normale potra' ricalcolarle quando saranno davvero risolte.
    Applica anche le correzioni manuali da RESOLUTION_OVERRIDES."""
    resolutions_done = state.get("resolutions_done", {})
    if not resolutions_done and not RESOLUTION_OVERRIDES:
        log.info("Nessuna risoluzione passata da ri-verificare.")
        return

    log.info(f"Ri-verifica di {len(resolutions_done)} risoluzioni passate...")

    # 1. Applica correzioni manuali (RESOLUTION_OVERRIDES)
    keys_to_update = {}
    for key, correct_winner in RESOLUTION_OVERRIDES.items():
        old_winner = resolutions_done.get(key)
        if old_winner is None:
            continue
        if old_winner.strip() != correct_winner.strip():
            keys_to_update[key] = correct_winner
            log.info(f"  OVERRIDE: {key}: {old_winner} -> {correct_winner}")

    for key, new_winner in keys_to_update.items():
        resolutions_done[key] = new_winner
        city, target_date = key.rsplit("_", 1)
        _update_resolution_in_excel(city, target_date, new_winner)
        _update_resolution_in_combined_excel(city, target_date, new_winner)

    if keys_to_update:
        save_state(state)
        log.info(f"  Override applicati: {len(keys_to_update)}")

    # 2. Ri-verifica via API per eventi ancora disponibili
    open_events = fetch_temp_events(closed=False)
    closed_events = fetch_temp_events(closed=True)
    all_events = open_events + closed_events

    # Mappa key -> evento per lookup veloce
    event_map = {}
    for event in all_events:
        title = event.get("title", "")
        city = city_from_title(title)
        event_end = event.get("endDate", "")
        target_date = date_from_title(title, end_date_hint=event_end)
        if city and target_date:
            key = f"{city}_{target_date}"
            event_map[key] = event

    keys_to_remove = []
    keys_to_update_api = {}

    for key, old_winner in list(resolutions_done.items()):
        event = event_map.get(key)
        if not event:
            # Evento non trovato nell'API, lo lasciamo com'e'
            continue

        # Ri-verifica con la soglia attuale
        new_winner = None
        for m in event.get("markets") or []:
            try:
                outcome_prices = json.loads(m.get("outcomePrices", "[]"))
                if outcome_prices and float(outcome_prices[0]) >= 0.995:
                    new_winner = m.get("groupItemTitle") or m.get("question", "")
                    break
            except (json.JSONDecodeError, ValueError, IndexError):
                continue

        if new_winner is None:
            # Non supera piu' la soglia: rimuovere
            keys_to_remove.append(key)
            log.info(f"  RIMOSSA risoluzione {key} (vecchio winner: {old_winner}) — sotto soglia 0.995")
        elif new_winner.strip() != old_winner.strip():
            # Winner diverso: aggiornare
            keys_to_update_api[key] = new_winner
            log.info(f"  AGGIORNATA risoluzione {key}: {old_winner} -> {new_winner}")

    # Applica rimozioni
    for key in keys_to_remove:
        del resolutions_done[key]
        city, target_date = key.rsplit("_", 1)
        _remove_resolution_from_excel(city, target_date)
        _remove_resolution_from_combined_excel(city, target_date)

    # Applica aggiornamenti da API
    for key, new_winner in keys_to_update_api.items():
        resolutions_done[key] = new_winner
        city, target_date = key.rsplit("_", 1)
        _update_resolution_in_excel(city, target_date, new_winner)
        _update_resolution_in_combined_excel(city, target_date, new_winner)

    if keys_to_remove or keys_to_update_api:
        save_state(state)
        log.info(f"  Ri-verifica API completata: {len(keys_to_remove)} rimosse, {len(keys_to_update_api)} aggiornate.")
    else:
        log.info("  Ri-verifica API completata: tutte le risoluzioni sono corrette.")


def _remove_resolution_from_excel(city: str, target_date: str):
    """Rimuove una riga di risoluzione dal foglio Risoluzioni dell'Excel principale."""
    if not EXCEL_FILE.exists():
        return
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
    except Exception:
        return
    if "Risoluzioni" not in wb.sheetnames:
        return
    ws = wb["Risoluzioni"]
    for r in range(ws.max_row, 1, -1):
        if ws.cell(r, 1).value == city and ws.cell(r, 2).value == target_date:
            ws.delete_rows(r, 1)
            break
    try:
        wb.save(EXCEL_FILE)
    except Exception as e:
        log.warning(f"Errore salvataggio Excel dopo rimozione risoluzione: {e}")


def _update_resolution_in_excel(city: str, target_date: str, new_winner: str):
    """Aggiorna il bucket vincente nel foglio Risoluzioni dell'Excel principale."""
    if not EXCEL_FILE.exists():
        return
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
    except Exception:
        return
    if "Risoluzioni" not in wb.sheetnames:
        return
    ws = wb["Risoluzioni"]
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value == city and ws.cell(r, 2).value == target_date:
            ws.cell(r, 3, new_winner)
            ws.cell(r, 4, datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"))
            break
    try:
        wb.save(EXCEL_FILE)
    except Exception as e:
        log.warning(f"Errore salvataggio Excel dopo aggiornamento risoluzione: {e}")


def _remove_resolution_from_combined_excel(city: str, target_date: str):
    """Rimuove le righe RISOLUZIONE dal file Excel combinato."""
    if not EXCEL_COMBINED.exists():
        return
    try:
        wb = openpyxl.load_workbook(EXCEL_COMBINED)
    except Exception:
        return
    section_title = f"{city} \u2014 {target_date}"
    for sheet_name in ["Prob Deterministici", "Prob Ensemble", "Prob Combinate", "Prob Mixture Raw", "Prob Comb 2", "Prob Comb 3", "Prob Comb 4"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for r in range(ws.max_row, 0, -1):
            if ws.cell(r, 1).value == "RISOLUZIONE":
                # Verifica che sia nella sezione giusta cercando il titolo sopra
                for r2 in range(r - 1, 0, -1):
                    cell_val = ws.cell(r2, 1).value
                    if isinstance(cell_val, str) and " \u2014 " in cell_val:
                        if cell_val == section_title:
                            ws.delete_rows(r, 1)
                        break
    try:
        wb.save(EXCEL_COMBINED)
    except Exception as e:
        log.warning(f"Errore salvataggio combinato dopo rimozione risoluzione: {e}")


def _update_resolution_in_combined_excel(city: str, target_date: str, new_winner: str):
    """Aggiorna il bucket vincente nelle righe RISOLUZIONE del file Excel combinato."""
    if not EXCEL_COMBINED.exists():
        return
    try:
        wb = openpyxl.load_workbook(EXCEL_COMBINED)
    except Exception:
        return
    section_title = f"{city} \u2014 {target_date}"
    for sheet_name in ["Prob Deterministici", "Prob Ensemble", "Prob Combinate", "Prob Mixture Raw", "Prob Comb 2", "Prob Comb 3", "Prob Comb 4"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for r in range(ws.max_row, 0, -1):
            if ws.cell(r, 1).value == "RISOLUZIONE":
                # Verifica che sia nella sezione giusta
                in_section = False
                for r2 in range(r - 1, 0, -1):
                    cell_val = ws.cell(r2, 1).value
                    if isinstance(cell_val, str) and " \u2014 " in cell_val:
                        if cell_val == section_title:
                            in_section = True
                        break
                if not in_section:
                    continue
                # Leggi bucket labels dall'header (sezione_row + 1)
                header_row = None
                for r2 in range(r - 1, 0, -1):
                    cell_val = ws.cell(r2, 1).value
                    if isinstance(cell_val, str) and " \u2014 " in cell_val:
                        header_row = r2 + 1
                        break
                if header_row:
                    col = 4
                    while ws.cell(header_row, col).value is not None:
                        lbl = str(ws.cell(header_row, col).value)
                        if lbl.strip() == new_winner.strip():
                            ws.cell(r, col, "VINCENTE")
                            ws.cell(r, col).fill = GREEN_FILL
                            ws.cell(r, col).font = Font(bold=True, color="006100", size=10)
                        else:
                            ws.cell(r, col, "")
                            ws.cell(r, col).fill = PatternFill("solid", fgColor="E2EFDA")
                            ws.cell(r, col).font = Font(bold=True, size=10)
                        col += 1
    try:
        wb.save(EXCEL_COMBINED)
    except Exception as e:
        log.warning(f"Errore salvataggio combinato dopo aggiornamento risoluzione: {e}")


# ═══════════════════════════════════════════════════════════════════════════════
# ENSEMBLE - FETCH & CALC
# ═══════════════════════════════════════════════════════════════════════════════

def fetch_ensemble_for_city(lat: float, lon: float, forecast_date: str) -> dict[str, list[float]]:
    """
    Chiama Open-Meteo ensemble API per ogni modello.
    Ritorna {model_name: [T_max per ogni membro]}.
    """
    all_members = {}

    for model in ENSEMBLE_MODELS:
        try:
            r = requests.get(ENSEMBLE_API, params={
                "latitude": lat,
                "longitude": lon,
                "daily": "temperature_2m_max",
                "models": model,
                "timezone": "auto",
                "start_date": forecast_date,
                "end_date": forecast_date,
            }, timeout=30)

            if r.status_code != 200:
                continue

            data = r.json()
            daily = data.get("daily", {})
            times = daily.get("time", [])

            if forecast_date not in times:
                continue

            idx = times.index(forecast_date)
            values = []

            # Control run (member00)
            if "temperature_2m_max" in daily:
                v = daily["temperature_2m_max"][idx]
                if v is not None:
                    values.append(v)

            # Members
            for dkey in sorted(daily.keys()):
                if dkey.startswith("temperature_2m_max_member"):
                    v = daily[dkey][idx]
                    if v is not None:
                        values.append(v)

            if values:
                all_members[model] = values

            time_mod.sleep(0.3)

        except Exception as e:
            log.debug(f"  {model}: {e}")
            continue

    return all_members


def calc_ensemble_stats(all_members: dict[str, list[float]],
                        parsed_buckets: list[dict]) -> dict:
    """
    Calcola statistiche e probabilita dai dati ensemble.
    parsed_buckets = lista di bucket parsati (con unit, low, high).
    I valori ensemble sono sempre in Celsius; la conversione a F
    avviene dentro calc_bucket_prob_from_celsius se il bucket e' in F.
    """
    if not all_members:
        return {
            "combined_probs": {},
            "combined_stats": {},
            "per_model": {},
            "all_values": [],
            "n_total": 0,
            "n_models": 0,
        }

    def _bucket_probs(values, buckets):
        arr = np.array(values)
        probs = {}
        for b in buckets:
            if b is None:
                continue
            key = b["label"]
            probs[key] = calc_bucket_prob_from_celsius(arr, b)
        return probs

    # Per modello
    per_model = {}
    for model, vals in all_members.items():
        arr = np.array(vals)
        per_model[model] = {
            "n": len(vals),
            "mean": float(np.mean(arr)),
            "median": float(np.median(arr)),
            "std": float(np.std(arr)),
            "min": float(np.min(arr)),
            "max": float(np.max(arr)),
            "probs": _bucket_probs(vals, parsed_buckets),
            "raw": vals,
        }

    # Combinato (tutti i membri insieme)
    all_vals = []
    for vals in all_members.values():
        all_vals.extend(vals)

    arr_all = np.array(all_vals)

    # Temperatura media anche in F per citta US
    unit = parsed_buckets[0]["unit"] if parsed_buckets and parsed_buckets[0] else "C"
    mean_c = float(np.mean(arr_all))
    median_c = float(np.median(arr_all))

    std_c = float(np.std(arr_all))
    min_c = float(np.min(arr_all))
    max_c = float(np.max(arr_all))

    def _to_display(val_c):
        return val_c * 9/5 + 32 if unit == "F" else val_c

    return {
        "combined_probs": _bucket_probs(all_vals, parsed_buckets),
        "combined_stats": {
            "mean": mean_c,
            "median": median_c,
            "std": std_c,
            "min": min_c,
            "max": max_c,
            "mean_display": _to_display(mean_c),
            "median_display": _to_display(median_c),
            "std_display": std_c * 9/5 if unit == "F" else std_c,
            "min_display": _to_display(min_c),
            "max_display": _to_display(max_c),
            "unit": unit,
        },
        "per_model": per_model,
        "all_values": all_vals,
        "n_total": len(all_vals),
        "n_models": len(all_members),
    }


# ═══════════════════════════════════════════════════════════════════════════════
# DETERMINISTICI - FETCH, BC, AGGREGAZIONE, GAUSSIANA
# ═══════════════════════════════════════════════════════════════════════════════

def fetch_deterministic_for_city(lat: float, lon: float, forecast_date: str,
                                  models: list[str]) -> dict[str, float | None]:
    """
    Chiama Open-Meteo forecast API per i modelli deterministici.
    Ritorna {model_name: T_max_celsius} per la data richiesta.
    """
    if not models:
        return {}

    results = {}
    # Batch fino a 10 modelli per chiamata per evitare URL troppo lunghi
    batch_size = 10
    for i in range(0, len(models), batch_size):
        batch = models[i:i + batch_size]
        try:
            r = requests.get(DETERMINISTIC_API, params={
                "latitude": lat,
                "longitude": lon,
                "daily": "temperature_2m_max",
                "models": ",".join(batch),
                "start_date": forecast_date,
                "end_date": forecast_date,
                "timezone": "UTC",
            }, timeout=30)

            if r.status_code != 200:
                log.warning(f"  Deterministic API {r.status_code} per batch {batch}")
                continue

            data = r.json()
            daily = data.get("daily", {})
            times = daily.get("time", [])

            if forecast_date not in times:
                continue
            idx = times.index(forecast_date)

            if len(batch) == 1:
                # Singolo modello: chiave e' "temperature_2m_max"
                key = "temperature_2m_max"
                if key in daily and daily[key][idx] is not None:
                    results[batch[0]] = daily[key][idx]
            else:
                # Multi modello: chiave e' "temperature_2m_max_{model}"
                for model in batch:
                    key = f"temperature_2m_max_{model}"
                    if key in daily and daily[key][idx] is not None:
                        results[model] = daily[key][idx]

            time_mod.sleep(0.3)

        except Exception as e:
            log.debug(f"  Deterministic batch {batch}: {e}")
            continue

    return results


def apply_bc_and_aggregate(raw_forecasts: dict[str, float | None],
                            corrections: dict[str, int],
                            method: str) -> dict:
    """
    Applica bias correction e aggrega le previsioni deterministiche.
    method: 'Media', 'Mediana', 'Media +BC', 'Mediana +BC', 'Singolo'
    Ritorna {forecast_c, per_model: {model: {raw, bc, corrected}}, n_models}.
    """
    use_bc = "+BC" in method

    per_model = {}
    corrected_values = []

    for model, raw_c in raw_forecasts.items():
        if raw_c is None:
            continue
        bc = corrections.get(model, 0) if use_bc else 0
        corrected = round(raw_c) + bc
        per_model[model] = {"raw": raw_c, "bc": bc, "corrected": corrected}
        corrected_values.append(corrected)

    if not corrected_values:
        return {"forecast_c": None, "per_model": per_model, "n_models": 0}

    arr = np.array(corrected_values)

    if "Mediana" in method or "Singolo" in method:
        forecast_c = float(np.median(arr))
    else:
        forecast_c = float(np.mean(arr))

    return {
        "forecast_c": forecast_c,
        "per_model": per_model,
        "n_models": len(corrected_values),
    }


def gaussian_bucket_probs(forecast_c: float, mae: float,
                           parsed_buckets: list[dict]) -> dict[str, float]:
    """
    Converte un forecast puntuale in distribuzione di probabilita per bucket
    usando una Gaussiana centrata sul forecast con std = MAE.
    """
    from scipy.stats import norm

    if not parsed_buckets:
        return {}

    unit = parsed_buckets[0]["unit"] if parsed_buckets[0] else "C"

    # Converti forecast e MAE in unita del bucket se necessario
    if unit == "F":
        center = forecast_c * 9 / 5 + 32
        sigma = mae * 9 / 5
    else:
        center = forecast_c
        sigma = mae

    # Sigma minimo per evitare distribuzione degenere
    sigma = max(sigma, 0.3)

    dist = norm(loc=center, scale=sigma)
    probs = {}

    for b in parsed_buckets:
        if b is None:
            continue

        if b["is_lower"]:
            # "X or below" -> P(T <= X + 0.5)
            p = dist.cdf(b["high"] + 0.5)
        elif b["is_upper"]:
            # "X or higher" -> P(T >= X - 0.5)
            p = 1 - dist.cdf(b["low"] - 0.5)
        else:
            # Range bucket: P(low - 0.5 <= T <= high + 0.5)
            p = dist.cdf(b["high"] + 0.5) - dist.cdf(b["low"] - 0.5)

        probs[b["label"]] = max(0.0, p)

    # Normalizza a 1.0
    total = sum(probs.values())
    if total > 0:
        probs = {k: v / total for k, v in probs.items()}

    return probs


def combine_probabilities(ens_probs: dict[str, float],
                           det_probs: dict[str, float],
                           verde_pct: float) -> dict[str, float]:
    """
    Combina probabilita ensemble e deterministiche.
    alpha = 0.75 * verde_pct / 100 (peso deterministici)
    P_combined = alpha * P_det + (1-alpha) * P_ens
    """
    alpha = 0.75 * verde_pct / 100.0
    alpha = min(alpha, 0.95)  # cap massimo

    combined = {}
    all_labels = set(list(ens_probs.keys()) + list(det_probs.keys()))

    for label in all_labels:
        p_ens = ens_probs.get(label, 0.0)
        p_det = det_probs.get(label, 0.0)
        combined[label] = alpha * p_det + (1 - alpha) * p_ens

    # Normalizza
    total = sum(combined.values())
    if total > 0:
        combined = {k: v / total for k, v in combined.items()}

    return combined


def do_deterministic_forecast(city: str, target_date: str,
                               parsed_buckets: list[dict]) -> dict | None:
    """
    Esegue il forecast deterministico completo per una citta/data.
    Ritorna dict con forecast, gaussian probs, config, o None se non disponibile.
    """
    opt_city = CITY_NAME_TO_OPT.get(city)
    if not opt_city:
        log.debug(f"  Det: citta '{city}' non mappata a config ottimizzazione")
        return None

    config = load_deterministic_config()
    month = int(target_date.split("-")[1])
    season = get_season(month, opt_city)
    key = (opt_city, season)

    if key not in config:
        log.debug(f"  Det: config non trovata per {opt_city}/{season}")
        return None

    cfg = config[key]
    city_data = match_city(city)
    if not city_data:
        return None

    log.info(f"    Deterministici [{opt_city}/{season}]: {len(cfg['models'])} modelli, "
             f"metodo={cfg['method']}, verde={cfg['verde']}%")

    # Fetch
    raw = fetch_deterministic_for_city(
        city_data["lat"], city_data["lon"], target_date, cfg["models"])
    log.info(f"    Deterministici: {len(raw)}/{len(cfg['models'])} modelli ricevuti")

    if not raw:
        return None

    # BC + aggregazione
    agg = apply_bc_and_aggregate(raw, cfg["corrections"], cfg["method"])

    if agg["forecast_c"] is None:
        return None

    # Probabilita Gaussiana
    det_probs = gaussian_bucket_probs(agg["forecast_c"], cfg["mae"], parsed_buckets)

    # Display value
    unit = parsed_buckets[0]["unit"] if parsed_buckets and parsed_buckets[0] else "C"
    forecast_display = agg["forecast_c"] * 9 / 5 + 32 if unit == "F" else agg["forecast_c"]

    log.info(f"    Deterministici: forecast={forecast_display:.1f}°{unit} "
             f"(MAE={cfg['mae']:.2f}°C, metodo={cfg['method']})")

    return {
        "forecast_c": agg["forecast_c"],
        "forecast_display": forecast_display,
        "unit": unit,
        "per_model": agg["per_model"],
        "n_models": agg["n_models"],
        "method": cfg["method"],
        "mae": cfg["mae"],
        "verde": cfg["verde"],
        "season": season,
        "gaussian_probs": det_probs,
        "config": cfg,
    }



# ═══════════════════════════════════════════════════════════════════════════════
# MIXTURE MODEL - Probabilita calibrate multi-modello
# ═══════════════════════════════════════════════════════════════════════════════

def mixture_bucket_probs(raw_forecasts: dict[str, float],
                          city_opt: str, season: str,
                          parsed_buckets: list[dict],
                          lead_time: str = "D1") -> dict | None:
    """
    Calcola probabilita per bucket usando mixture-of-normals calibrato.

    Per ogni modello con stats storici, costruisce N(forecast - bias, sigma²),
    pesa con inverse-MAE, modula sigma per spread inter-modello,
    e calibra con isotonic regression.

    Ritorna dict con mixture_probs, dettagli, o None se stats non disponibili.
    """
    from scipy.stats import norm

    model_stats_data = load_model_stats()
    if not model_stats_data or "stats" not in model_stats_data:
        return None

    all_stats = model_stats_data["stats"]
    iso_model = model_stats_data.get("iso_enhanced")

    if not parsed_buckets:
        return None

    unit = parsed_buckets[0]["unit"] if parsed_buckets[0] else "C"

    # Raccogli stats per ogni modello disponibile
    models_used = []
    for model_name, forecast_c in raw_forecasts.items():
        if forecast_c is None:
            continue
        key = (city_opt, model_name, season, lead_time)
        if key not in all_stats:
            continue
        s = all_stats[key]
        models_used.append({
            "model": model_name,
            "forecast_c": forecast_c,
            "bias": s["bias"],
            "sigma": s["sigma"],
            "mae": s["mae"],
            "weight": s["weight"],
            "mu_c": forecast_c - s["bias"],  # valore atteso temperatura reale (°C)
        })

    if len(models_used) < 3:
        return None

    # Normalizza pesi
    total_w = sum(m["weight"] for m in models_used)
    for m in models_used:
        m["weight_norm"] = m["weight"] / total_w

    # Spread inter-modello (dispersione dei valori attesi corretti)
    corrected_means = [m["mu_c"] for m in models_used]
    inter_model_spread = float(np.std(corrected_means))

    # Sigma potenziata: sqrt(sigma_storico² + spread²)
    for m in models_used:
        m["sigma_enh"] = np.sqrt(m["sigma"]**2 + inter_model_spread**2)

    # Media pesata
    weighted_mean_c = sum(m["mu_c"] * m["weight_norm"] for m in models_used)

    # Calcola probabilita per ogni bucket Polymarket
    probs = {}
    for b in parsed_buckets:
        if b is None:
            continue

        p = 0.0
        for m in models_used:
            # Converti in unita del bucket se necessario
            if unit == "F":
                mu = m["mu_c"] * 9 / 5 + 32
                sigma = m["sigma_enh"] * 9 / 5
            else:
                mu = m["mu_c"]
                sigma = m["sigma_enh"]

            sigma = max(sigma, 0.3)
            w = m["weight_norm"]

            if b["is_lower"]:
                p += w * norm.cdf((b["high"] + 0.5 - mu) / sigma)
            elif b["is_upper"]:
                p += w * (1 - norm.cdf((b["low"] - 0.5 - mu) / sigma))
            else:
                p += w * (norm.cdf((b["high"] + 0.5 - mu) / sigma) -
                          norm.cdf((b["low"] - 0.5 - mu) / sigma))

        probs[b["label"]] = max(0.0, p)

    # Normalizza
    total = sum(probs.values())
    if total > 0:
        probs = {k: v / total for k, v in probs.items()}

    # Calibrazione isotonica
    probs_raw = dict(probs)
    if iso_model is not None:
        labels = list(probs.keys())
        raw_p = np.array([probs[l] for l in labels])
        cal_p = iso_model.predict(raw_p)
        cal_p = cal_p / cal_p.sum()
        probs = dict(zip(labels, cal_p))

    return {
        "mixture_probs": probs,
        "mixture_probs_raw": probs_raw,
        "n_models_used": len(models_used),
        "inter_model_spread": inter_model_spread,
        "weighted_mean_c": weighted_mean_c,
        "models_detail": models_used,
    }


def do_mixture_forecast(city: str, target_date: str,
                         parsed_buckets: list[dict]) -> dict | None:
    """
    Esegue il forecast mixture completo: scarica TUTTI i modelli deterministici
    e calcola le probabilita calibrate per bucket.
    """
    opt_city = CITY_NAME_TO_OPT.get(city)
    if not opt_city:
        return None

    # Verifica che model stats siano disponibili
    model_stats_data = load_model_stats()
    if not model_stats_data or "stats" not in model_stats_data:
        log.debug(f"  Mixture: stats non disponibili")
        return None

    city_data = match_city(city)
    if not city_data:
        return None

    month = int(target_date.split("-")[1])
    season = get_season(month, opt_city)

    # Scarica TUTTI i modelli deterministici
    log.info(f"    Mixture: fetching {len(ALL_DETERMINISTIC_MODELS)} modelli deterministici...")
    raw = fetch_deterministic_for_city(
        city_data["lat"], city_data["lon"], target_date, ALL_DETERMINISTIC_MODELS)
    log.info(f"    Mixture: {len(raw)}/{len(ALL_DETERMINISTIC_MODELS)} modelli ricevuti")

    if len(raw) < 3:
        log.info(f"    Mixture: troppi pochi modelli, skip")
        return None

    # Calcola probabilita mixture
    result = mixture_bucket_probs(raw, opt_city, season, parsed_buckets)

    if result:
        log.info(f"    Mixture: {result['n_models_used']} modelli usati, "
                 f"spread={result['inter_model_spread']:.2f}°C, "
                 f"media={result['weighted_mean_c']:.1f}°C")

    return result


def do_mixture_forecast_optimized(city: str, target_date: str,
                                   parsed_buckets: list[dict]) -> dict | None:
    """
    Come do_mixture_forecast ma usa SOLO i modelli selezionati
    dall'ottimizzazione per stagione (file Ottimizzazione Ensemble Stagioni.xlsx).
    """
    opt_city = CITY_NAME_TO_OPT.get(city)
    if not opt_city:
        return None

    model_stats_data = load_model_stats()
    if not model_stats_data or "stats" not in model_stats_data:
        return None

    city_data = match_city(city)
    if not city_data:
        return None

    month = int(target_date.split("-")[1])
    season = get_season(month, opt_city)

    # Carica config ottimizzazione per sapere quali modelli usare
    config = load_deterministic_config()
    key = (opt_city, season)
    if key not in config:
        log.debug(f"  Mixture Opt: config non trovata per {opt_city}/{season}")
        return None

    cfg = config[key]
    optimized_models = cfg["models"]

    log.info(f"    Mixture Opt: fetching {len(optimized_models)} modelli ottimizzati...")
    raw = fetch_deterministic_for_city(
        city_data["lat"], city_data["lon"], target_date, optimized_models)
    log.info(f"    Mixture Opt: {len(raw)}/{len(optimized_models)} modelli ricevuti")

    if len(raw) < 2:
        log.info(f"    Mixture Opt: troppi pochi modelli, skip")
        return None

    result = mixture_bucket_probs(raw, opt_city, season, parsed_buckets)

    if result:
        log.info(f"    Mixture Opt: {result['n_models_used']} modelli usati, "
                 f"spread={result['inter_model_spread']:.2f}°C, "
                 f"media={result['weighted_mean_c']:.1f}°C")

    return result



def _do_mixture_forecast_topN(city: str, target_date: str,
                               parsed_buckets: list[dict],
                               top_key: str) -> dict | None:
    """
    Mixture calibrato usando i top N modelli per Verde% storica.
    top_key: "top5" o "top10".
    """
    opt_city = CITY_NAME_TO_OPT.get(city)
    if not opt_city:
        return None

    model_stats_data = load_model_stats()
    if not model_stats_data or "stats" not in model_stats_data:
        return None

    city_data = match_city(city)
    if not city_data:
        return None

    month = int(target_date.split("-")[1])
    season = get_season(month, opt_city)

    # Carica top modelli — il file usa i nomi "confronto" (= opt_city)
    top_models_data = load_top_models()
    city_top = top_models_data.get(opt_city)
    if not city_top or not city_top.get(top_key):
        log.debug(f"  Mixture {top_key}: nessun modello trovato per {opt_city}")
        return None

    models = city_top[top_key]
    label = top_key.upper()

    log.info(f"    Mixture {label}: fetching {len(models)} modelli...")
    raw = fetch_deterministic_for_city(
        city_data["lat"], city_data["lon"], target_date, models)
    log.info(f"    Mixture {label}: {len(raw)}/{len(models)} modelli ricevuti")

    if len(raw) < 2:
        log.info(f"    Mixture {label}: troppi pochi modelli, skip")
        return None

    result = mixture_bucket_probs(raw, opt_city, season, parsed_buckets)

    if result:
        log.info(f"    Mixture {label}: {result['n_models_used']} modelli usati, "
                 f"spread={result['inter_model_spread']:.2f}°C, "
                 f"media={result['weighted_mean_c']:.1f}°C")

    return result


def do_mixture_forecast_top5(city: str, target_date: str,
                              parsed_buckets: list[dict]) -> dict | None:
    """Mixture calibrato con i top 5 modelli per Verde% storica."""
    return _do_mixture_forecast_topN(city, target_date, parsed_buckets, "top5")


def do_mixture_forecast_top10(city: str, target_date: str,
                               parsed_buckets: list[dict]) -> dict | None:
    """Mixture calibrato con i top 10 modelli per Verde% storica."""
    return _do_mixture_forecast_topN(city, target_date, parsed_buckets, "top10")




# ═══════════════════════════════════════════════════════════════════════════════
# SCHEDULING
# ═══════════════════════════════════════════════════════════════════════════════

def get_snapshot_utc(city: str, target_date: str, hour: int, minute: int = 0,
                     mode: str = "local") -> datetime | None:
    """
    Calcola il momento UTC dello snapshot per una citta.
    mode="local": hour:minute nel fuso orario della citta (giorno prima del target)
    mode="utc":   hour:minute UTC fisso (giorno prima del target)
    """
    target = datetime.strptime(target_date, "%Y-%m-%d").date()
    snapshot_day = target - timedelta(days=1)

    if mode == "utc":
        return datetime(snapshot_day.year, snapshot_day.month, snapshot_day.day,
                        hour, minute, 0, tzinfo=timezone.utc)

    # mode == "local"
    city_data = match_city(city)
    if not city_data:
        return None
    tz = ZoneInfo(city_data["tz"])
    snapshot_local = datetime(snapshot_day.year, snapshot_day.month, snapshot_day.day,
                              hour, minute, 0, tzinfo=tz)
    return snapshot_local.astimezone(timezone.utc)


def get_pending_snapshots(markets: list[dict], state: dict, hours: list) -> list[dict]:
    """Trova mercati per cui e' ora di fare lo snapshot (per ciascun orario).
    hours: lista di tuple (hour, minute, mode) dove mode e' 'utc' o 'local'."""
    now = datetime.now(timezone.utc)
    pending = []

    for mkt in markets:
        for entry_hours in hours:
            hour, minute = entry_hours[0], entry_hours[1]
            mode = entry_hours[2] if len(entry_hours) > 2 else "local"
            mode_tag = "u" if mode == "utc" else ""
            key = f"{mkt['city']}_{mkt['target_date']}_h{hour}{minute:02d}{mode_tag}"

            if key in state.get("snapshots_done", {}):
                continue

            snap_utc = get_snapshot_utc(mkt["city"], mkt["target_date"], hour, minute, mode)
            if snap_utc is None:
                continue

            # Snapshot se siamo nella finestra [snap_utc, snap_utc + SNAPSHOT_WINDOW]
            diff = (now - snap_utc).total_seconds()
            if 0 <= diff <= SNAPSHOT_WINDOW:
                entry = dict(mkt)
                entry["snapshot_utc"] = snap_utc
                entry["key"] = key
                entry["snapshot_hour"] = hour
                entry["snapshot_minute"] = minute
                entry["snapshot_mode"] = mode
                pending.append(entry)

    return pending


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL - STRUTTURA E SCRITTURA
# ═══════════════════════════════════════════════════════════════════════════════

# Stili
HDR_FILL    = PatternFill("solid", fgColor="1F4E79")
HDR_FONT    = Font(bold=True, color="FFFFFF", size=10)
CITY_FILL   = PatternFill("solid", fgColor="2E75B6")
CITY_FONT   = Font(bold=True, color="FFFFFF", size=11)
GREEN_FILL  = PatternFill("solid", fgColor="C6EFCE")
RED_FILL    = PatternFill("solid", fgColor="FFC7CE")
YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
BLUE_FILL   = PatternFill("solid", fgColor="D6E4F0")
GRAY_FILL   = PatternFill("solid", fgColor="F2F2F2")
CENTER      = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        c = ws.cell(row, col)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = CENTER
        c.border = THIN_BORDER


def _style_row(ws, row, max_col, fill=None, font=None):
    for col in range(1, max_col + 1):
        c = ws.cell(row, col)
        c.border = THIN_BORDER
        c.alignment = CENTER
        if fill:
            c.fill = fill
        if font:
            c.font = font


def get_or_create_sheet(wb, name, headers):
    """Ritorna il foglio esistente o ne crea uno nuovo con headers."""
    if name in wb.sheetnames:
        return wb[name]
    ws = wb.create_sheet(title=name)
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)
    _style_header(ws, 1, len(headers))
    ws.freeze_panes = "A2"
    return ws


def init_workbook() -> openpyxl.Workbook:
    """Inizializza o carica il workbook Excel."""
    if EXCEL_FILE.exists():
        try:
            return openpyxl.load_workbook(EXCEL_FILE)
        except Exception as e:
            log.warning(f"Excel corrotto, ricreo: {e}")
            EXCEL_FILE.unlink()

    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    return wb


def write_snapshot_to_excel(city: str, target_date: str, snapshot_time: str,
                             snapshot_hour: int, snapshot_minute: int,
                             polymarket_buckets: list[dict],
                             ensemble: dict, snapshot_mode: str = "local"):
    """Scrive i dati di uno snapshot nel file Excel."""
    wb = init_workbook()

    # ── Foglio 1: Confronto ──────────────────────────────────────────────
    headers_confronto = [
        "Citta", "Data Target", "Ora Snapshot", "Snapshot UTC", "Bucket",
        "Polymarket %", "Ensemble %", "Edge (pp)", "Segnale",
        "Ens Media", "Ens Mediana", "Ens Std", "Ens Min", "Ens Max",
        "N Membri", "N Modelli",
    ]
    ws1 = get_or_create_sheet(wb, "Confronto", headers_confronto)

    # Deduplicazione Confronto: rimuovi righe esistenti per stessa citta/data/ora
    utc_suffix = " UTC" if snapshot_mode == "utc" else ""
    ora_snap_tag = f"{snapshot_hour}:{snapshot_minute:02d}{utc_suffix}"
    rows_to_del = []
    for r in range(2, ws1.max_row + 1):
        if (ws1.cell(r, 1).value == city and
            ws1.cell(r, 2).value == target_date and
            ws1.cell(r, 3).value == ora_snap_tag):
            rows_to_del.append(r)
    for r in reversed(rows_to_del):
        ws1.delete_rows(r)

    row = ws1.max_row + 1

    combined_probs = ensemble.get("combined_probs", {})
    stats = ensemble.get("combined_stats", {})

    for b in polymarket_buckets:
        label = b["label"]
        pm_prob = b["prob"]
        ens_prob = combined_probs.get(label, 0.0)
        edge = (ens_prob - pm_prob) * 100

        if edge > 5:
            signal = "BUY"
        elif edge < -5:
            signal = "SELL"
        else:
            signal = "-"

        ws1.cell(row, 1, city)
        ws1.cell(row, 2, target_date)
        ws1.cell(row, 3, f"{snapshot_hour}:{snapshot_minute:02d}")
        ws1.cell(row, 4, snapshot_time)
        ws1.cell(row, 5, b["label"])
        ws1.cell(row, 6, round(pm_prob * 100, 1))
        ws1.cell(row, 7, round(ens_prob * 100, 1))
        ws1.cell(row, 8, round(edge, 1))
        ws1.cell(row, 9, signal)
        ws1.cell(row, 10, round(stats.get("mean_display", stats.get("mean", 0)), 2))
        ws1.cell(row, 11, round(stats.get("median_display", stats.get("median", 0)), 2))
        ws1.cell(row, 12, round(stats.get("std_display", stats.get("std", 0)), 2))
        ws1.cell(row, 13, round(stats.get("min_display", stats.get("min", 0)), 2))
        ws1.cell(row, 14, round(stats.get("max_display", stats.get("max", 0)), 2))
        ws1.cell(row, 15, ensemble.get("n_total", 0))
        ws1.cell(row, 16, ensemble.get("n_models", 0))

        # Colorazione segnale
        if signal == "BUY":
            _style_row(ws1, row, 16, fill=GREEN_FILL)
        elif signal == "SELL":
            _style_row(ws1, row, 16, fill=RED_FILL)
        else:
            _style_row(ws1, row, 16)

        row += 1

    # Larghezze
    widths = [14, 12, 13, 20, 22, 13, 12, 10, 9, 10, 11, 9, 9, 9, 10, 10]
    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── Foglio 2: Ensemble Modelli ───────────────────────────────────────
    headers_modelli = [
        "Citta", "Data Target", "Ora Snapshot", "Modello", "N Membri",
        "Media", "Mediana", "Std", "Min", "Max",
    ]
    ws2 = get_or_create_sheet(wb, "Ensemble Modelli", headers_modelli)

    # Deduplicazione Ensemble Modelli
    rows_to_del = []
    for r in range(2, ws2.max_row + 1):
        if (ws2.cell(r, 1).value == city and
            ws2.cell(r, 2).value == target_date and
            ws2.cell(r, 3).value == ora_snap_tag):
            rows_to_del.append(r)
    for r in reversed(rows_to_del):
        ws2.delete_rows(r)

    row2 = ws2.max_row + 1

    unit = stats.get("unit", "C")
    def _to_disp(val_c):
        return val_c * 9/5 + 32 if unit == "F" else val_c

    for model, mdata in ensemble.get("per_model", {}).items():
        short = model.replace("_ensemble", "").replace("_seamless", "")
        ws2.cell(row2, 1, city)
        ws2.cell(row2, 2, target_date)
        ws2.cell(row2, 3, f"{snapshot_hour}:{snapshot_minute:02d}")
        ws2.cell(row2, 4, short)
        ws2.cell(row2, 5, mdata["n"])
        ws2.cell(row2, 6, round(_to_disp(mdata["mean"]), 2))
        ws2.cell(row2, 7, round(_to_disp(mdata["median"]), 2))
        ws2.cell(row2, 8, round(mdata["std"] * 9/5 if unit == "F" else mdata["std"], 2))
        ws2.cell(row2, 9, round(_to_disp(mdata["min"]), 2))
        ws2.cell(row2, 10, round(_to_disp(mdata["max"]), 2))
        _style_row(ws2, row2, 10)
        row2 += 1

    for i, w in enumerate([14, 12, 13, 26, 10, 8, 9, 7, 7, 7], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Foglio 3: Ensemble Raw ───────────────────────────────────────────
    headers_raw = ["Citta", "Data Target", "Ora Snapshot", "Modello", "Membro", "T Max (C)"]
    ws3 = get_or_create_sheet(wb, "Ensemble Raw", headers_raw)

    # Deduplicazione Ensemble Raw
    rows_to_del = []
    for r in range(2, ws3.max_row + 1):
        if (ws3.cell(r, 1).value == city and
            ws3.cell(r, 2).value == target_date and
            ws3.cell(r, 3).value == ora_snap_tag):
            rows_to_del.append(r)
    for r in reversed(rows_to_del):
        ws3.delete_rows(r)

    row3 = ws3.max_row + 1

    for model, mdata in ensemble.get("per_model", {}).items():
        short = model.replace("_ensemble", "").replace("_seamless", "")
        for i, val in enumerate(mdata["raw"]):
            ws3.cell(row3, 1, city)
            ws3.cell(row3, 2, target_date)
            ws3.cell(row3, 3, f"{snapshot_hour}:{snapshot_minute:02d}")
            ws3.cell(row3, 4, short)
            ws3.cell(row3, 5, f"member{i:02d}")
            ws3.cell(row3, 6, round(val, 2))
            _style_row(ws3, row3, 6)
            row3 += 1

    for i, w in enumerate([14, 12, 13, 26, 10, 10], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # ── Foglio 4: Probabilita per Modello ────────────────────────────────
    # Ogni citta/data ha i PROPRI bucket come header (sezione separata)
    sheet_name = "Probabilita per Modello"
    if sheet_name not in wb.sheetnames:
        ws4 = wb.create_sheet(title=sheet_name)
        ws4.freeze_panes = None
    else:
        ws4 = wb[sheet_name]

    # Deduplicazione: controlla se esiste gia' una sezione per questa citta/data/ora
    ora_snap_check = f"{snapshot_hour}:{snapshot_minute:02d}{utc_suffix}"
    section_title = f"{city} — {target_date}"
    already_exists = False
    for r in range(1, ws4.max_row + 1):
        cell_val = ws4.cell(r, 1).value
        if cell_val == section_title:
            # Controlla se l'ora snapshot nella riga POLYMARKET (r+2) e' la stessa
            poly_ora = ws4.cell(r + 2, 3).value
            if poly_ora == ora_snap_check:
                already_exists = True
                break
    if already_exists:
        log.info(f"    Probabilita per Modello: sezione {section_title} ({ora_snap_check}) gia' presente, skip")
        wb.save(EXCEL_FILE)
        return

    row4 = ws4.max_row + 1
    bucket_labels = [b["label"] for b in polymarket_buckets]
    n_buckets = len(bucket_labels)
    n_cols = 3 + n_buckets  # Modello, N Membri, Ora + buckets

    # Riga titolo sezione: "Citta - Data Target"
    ws4.cell(row4, 1, f"{city} — {target_date}")
    _style_row(ws4, row4, n_cols, fill=CITY_FILL, font=CITY_FONT)
    row4 += 1

    # Mini-header con bucket labels
    ws4.cell(row4, 1, "Modello")
    ws4.cell(row4, 2, "N Membri")
    ws4.cell(row4, 3, "Ora Snapshot")
    for ci, lbl in enumerate(bucket_labels, 4):
        ws4.cell(row4, ci, lbl)
    _style_header(ws4, row4, n_cols + 1)
    row4 += 1

    ora_snap = f"{snapshot_hour}:{snapshot_minute:02d}{utc_suffix}"

    # Riga Polymarket
    ws4.cell(row4, 1, "POLYMARKET")
    ws4.cell(row4, 2, "-")
    ws4.cell(row4, 3, ora_snap)
    for ci, b in enumerate(polymarket_buckets, 4):
        ws4.cell(row4, ci, f"{b['prob']*100:.1f}%")
    _style_row(ws4, row4, n_cols + 1, fill=YELLOW_FILL, font=Font(bold=True, size=10))
    row4 += 1

    # Riga Ensemble combinato
    ws4.cell(row4, 1, "ENSEMBLE (tutti)")
    ws4.cell(row4, 2, ensemble.get("n_total", 0))
    ws4.cell(row4, 3, ora_snap)
    for ci, b in enumerate(polymarket_buckets, 4):
        ep = combined_probs.get(b["label"], 0)
        ws4.cell(row4, ci, f"{ep*100:.1f}%")
    _style_row(ws4, row4, n_cols + 1, fill=BLUE_FILL, font=Font(bold=True, size=10))
    row4 += 1

    # Righe per modello
    for model, mdata in ensemble.get("per_model", {}).items():
        short = model.replace("_ensemble", "").replace("_seamless", "")
        ws4.cell(row4, 1, short)
        ws4.cell(row4, 2, mdata["n"])
        ws4.cell(row4, 3, ora_snap)
        for ci, b in enumerate(polymarket_buckets, 4):
            mp = mdata["probs"].get(b["label"], 0)
            ws4.cell(row4, ci, f"{mp*100:.1f}%")
        _style_row(ws4, row4, n_cols + 1)
        row4 += 1

    # Riga vuota separatrice
    row4 += 1

    # Larghezze
    ws4.column_dimensions["A"].width = 26
    ws4.column_dimensions["B"].width = 10
    ws4.column_dimensions["C"].width = 13
    for i in range(4, n_cols + 2):
        ws4.column_dimensions[get_column_letter(i)].width = 14

    # Salva
    wb.save(EXCEL_FILE)


def write_resolution_to_excel(city: str, target_date: str, winner: str, timestamp: str):
    """Scrive la risoluzione nel file Excel."""
    wb = init_workbook()

    headers = ["Citta", "Data Target", "Bucket Vincente", "Timestamp Risoluzione"]
    ws = get_or_create_sheet(wb, "Risoluzioni", headers)

    # Deduplicazione: controlla se citta+data gia' presenti
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value == city and ws.cell(r, 2).value == target_date:
            return  # gia' registrata

    row = ws.max_row + 1

    ws.cell(row, 1, city)
    ws.cell(row, 2, target_date)
    ws.cell(row, 3, winner)
    ws.cell(row, 4, timestamp)
    _style_row(ws, row, 4, fill=GREEN_FILL)

    for i, w in enumerate([14, 12, 24, 22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(EXCEL_FILE)


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL COMBINATO (Deterministici + Ensemble + Combinate)
# ═══════════════════════════════════════════════════════════════════════════════

def init_combined_workbook():
    """Inizializza o carica il workbook combinato."""
    if EXCEL_COMBINED.exists():
        try:
            return openpyxl.load_workbook(EXCEL_COMBINED)
        except Exception as e:
            log.warning(f"Excel combinato corrotto, ricreo: {e}")
            EXCEL_COMBINED.unlink()
    wb = openpyxl.Workbook()
    # Rimuovi foglio default
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    return wb


def write_combined_to_excel(city: str, target_date: str, snapshot_time: str,
                             snapshot_hour: int, snapshot_minute: int,
                             polymarket_buckets: list[dict],
                             ensemble: dict, deterministic: dict | None,
                             mixture: dict | None = None,
                             snapshot_mode: str = "local",
                             mixture_opt: dict | None = None,
                             mixture_top5: dict | None = None,
                             mixture_top10: dict | None = None,
):
    """Scrive nel file Excel combinato con fogli probabilita.
    Raggruppa i diversi orari nella stessa sezione per citta/data."""
    wb = init_combined_workbook()
    utc_suffix = " UTC" if snapshot_mode == "utc" else ""
    ora_snap = f"{snapshot_hour}:{snapshot_minute:02d}{utc_suffix}"
    bucket_labels = [b["label"] for b in polymarket_buckets]
    n_buckets = len(bucket_labels)
    n_cols = 3 + n_buckets  # Fonte, Info, Ora + bucket columns

    ens_probs = ensemble.get("combined_probs", {})
    det_probs = deterministic.get("gaussian_probs", {}) if deterministic else {}
    verde = deterministic.get("verde", 0) if deterministic else 0
    mix_probs = mixture.get("mixture_probs", {}) if mixture else {}

    # Prob Combinate: solo mixture (deterministici con calibrazione isotonica)
    comb_probs = mix_probs

    section_title = f"{city} — {target_date}"

    def _find_section(ws):
        """Trova sezione esistente. Ritorna (section_row, end_row) o (None, None)."""
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == section_title:
                end_row = ws.max_row
                for r2 in range(r + 1, ws.max_row + 2):
                    cell_val = ws.cell(r2, 1).value
                    if cell_val is None or str(cell_val).strip() == "":
                        end_row = r2 - 1
                        break
                    if r2 > r + 1 and isinstance(cell_val, str) and " \u2014 " in cell_val and cell_val != section_title:
                        end_row = r2 - 1
                        break
                return r, end_row
        return None, None

    def _time_exists_in_section(ws, section_row, end_row):
        """Controlla se questo ora_snap e' gia' presente nella sezione."""
        for r in range(section_row + 1, end_row + 1):
            if ws.cell(r, 3).value == ora_snap and ws.cell(r, 1).value == "POLYMARKET":
                return True
        return False

    def _write_prob_sheet(sheet_name: str, probs: dict, extra_info: str = "",
                          det_models: dict | None = None, det_unit: str = "C"):
        """Scrive un foglio di probabilita con sezioni raggruppate per citta/data.
        Quando arriva un secondo orario, lo accoda nella stessa sezione."""
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(title=sheet_name)
        else:
            ws = wb[sheet_name]

        section_row, end_row = _find_section(ws)

        if section_row is not None:
            # Sezione esistente - controlla se il tempo e' gia' presente
            if _time_exists_in_section(ws, section_row, end_row):
                return  # gia' presente

            # Trova punto di inserimento: prima di RISOLUZIONE se esiste,
            # altrimenti dopo l'ultima riga dati della sezione
            insert_at = end_row + 1
            for r in range(section_row + 1, end_row + 1):
                if ws.cell(r, 1).value == "RISOLUZIONE":
                    insert_at = r
                    break

            # Calcola righe da inserire: 3 base + eventuali dettagli modello
            n_insert = 3  # POLYMARKET, PROB, EDGE
            if det_models:
                n_insert += 1 + len(det_models)
            ws.insert_rows(insert_at, n_insert)
            row = insert_at
        else:
            # Nuova sezione alla fine
            row = ws.max_row + 1
            if row > 1:
                row += 1  # riga vuota separatrice

            # Titolo sezione
            ws.cell(row, 1, section_title)
            if extra_info:
                ws.cell(row, 2, extra_info)
            _style_row(ws, row, n_cols, fill=CITY_FILL, font=CITY_FONT)
            row += 1

            # Header con bucket labels
            ws.cell(row, 1, "Fonte")
            ws.cell(row, 2, "Info")
            ws.cell(row, 3, "Ora Snapshot")
            for ci, lbl in enumerate(bucket_labels, 4):
                ws.cell(row, ci, lbl)
            _style_header(ws, row, n_cols + 1)
            row += 1

        # Riga POLYMARKET
        ws.cell(row, 1, "POLYMARKET")
        ws.cell(row, 2, "-")
        ws.cell(row, 3, ora_snap)
        for ci, b in enumerate(polymarket_buckets, 4):
            ws.cell(row, ci, f"{b['prob']*100:.1f}%")
        _style_row(ws, row, n_cols + 1, fill=YELLOW_FILL, font=Font(bold=True, size=10))
        row += 1

        # Riga probabilita
        ws.cell(row, 1, sheet_name.upper())
        ws.cell(row, 2, extra_info)
        ws.cell(row, 3, ora_snap)
        for ci, b in enumerate(polymarket_buckets, 4):
            p = probs.get(b["label"], 0)
            ws.cell(row, ci, f"{p*100:.1f}%")
        _style_row(ws, row, n_cols + 1, fill=BLUE_FILL, font=Font(bold=True, size=10))
        row += 1

        # Riga EDGE
        ws.cell(row, 1, "EDGE (pp)")
        ws.cell(row, 2, "-")
        ws.cell(row, 3, ora_snap)
        for ci, b in enumerate(polymarket_buckets, 4):
            p = probs.get(b["label"], 0)
            edge = (p - b["prob"]) * 100
            ws.cell(row, ci, f"{edge:+.1f}")
            if edge > 5:
                ws.cell(row, ci).fill = GREEN_FILL
            elif edge < -5:
                ws.cell(row, ci).fill = RED_FILL
        _style_row(ws, row, n_cols + 1, font=Font(italic=True, size=9))
        row += 1

        # Dettaglio modelli deterministici (solo per Prob Deterministici)
        if det_models:
            ws.cell(row, 1, f"Dettaglio ({ora_snap}):")
            ws.cell(row, 1).font = Font(bold=True, size=9)
            row += 1
            for model, mdata in det_models.items():
                short = model.replace("_deterministic", "").replace("_seamless", "")
                bc_str = f"{mdata['bc']:+d}\u00b0C" if mdata["bc"] != 0 else "\u2014"
                corr_disp = mdata["corrected"] * 9 / 5 + 32 if det_unit == "F" else mdata["corrected"]
                raw_disp = mdata["raw"] * 9 / 5 + 32 if det_unit == "F" else mdata["raw"]
                ws.cell(row, 1, short)
                ws.cell(row, 2, f"Raw={raw_disp:.1f}\u00b0{det_unit} BC={bc_str} \u2192 {corr_disp:.1f}\u00b0{det_unit}")
                _style_row(ws, row, n_cols + 1)
                row += 1

        # Larghezze colonne
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 13
        for i in range(4, n_cols + 2):
            ws.column_dimensions[get_column_letter(i)].width = 14

    # ── Foglio 1: Probabilita Deterministici ──────────────────────────────
    if deterministic and det_probs:
        det_info = (f"Metodo={deterministic['method']} | "
                    f"Forecast={deterministic['forecast_display']:.1f}\u00b0{deterministic['unit']} | "
                    f"MAE={deterministic['mae']:.2f} | "
                    f"Verde={deterministic['verde']:.1f}% | "
                    f"Stagione={deterministic['season']}")
        _write_prob_sheet("Prob Deterministici", det_probs, det_info,
                          det_models=deterministic.get("per_model"),
                          det_unit=deterministic.get("unit", "C"))
    else:
        if "Prob Deterministici" not in wb.sheetnames:
            wb.create_sheet(title="Prob Deterministici")

    # ── Foglio 2: Probabilita Ensemble ────────────────────────────────────
    ens_info = f"N={ensemble.get('n_total', 0)} membri da {ensemble.get('n_models', 0)} modelli"
    _write_prob_sheet("Prob Ensemble", ens_probs, ens_info)

    # ── Foglio 3: Probabilita Combinate (usa mixture se disponibile) ────
    if mix_probs:
        comb_info = (f"MIXTURE CALIBRATO | {mixture['n_models_used']} modelli | "
                     f"spread={mixture['inter_model_spread']:.2f}°C | "
                     f"media={mixture['weighted_mean_c']:.1f}°C")
    elif deterministic and det_probs:
        alpha = 0.75 * verde / 100.0
        alpha = min(alpha, 0.95)
        comb_info = f"\u03b1={alpha:.2f} (75%\u00d7Verde) | Det\u00d7{alpha:.0%} + Ens\u00d7{1-alpha:.0%}"
    else:
        comb_info = "Solo ensemble (deterministici non disponibili)"
    _write_prob_sheet("Prob Combinate", comb_probs, comb_info)

    # ── Foglio 4: Prob Mixture dettaglio (se disponibile) ────────────────
    if mix_probs:
        mix_raw = mixture.get("mixture_probs_raw", mix_probs)
        mix_detail_info = (f"Pre-calibrazione | {mixture['n_models_used']} modelli | "
                           f"spread={mixture['inter_model_spread']:.2f}°C")
        _write_prob_sheet("Prob Mixture Raw", mix_raw, mix_detail_info)

    # ── Foglio 5: Prob Comb 2 (mixture con modelli ottimizzati per stagione) ──
    if mixture_opt:
        mix_opt_probs = mixture_opt.get("mixture_probs", {})
        if mix_opt_probs:
            mix_opt_info = (f"MIXTURE OPT | {mixture_opt['n_models_used']} modelli ottimizzati | "
                            f"spread={mixture_opt['inter_model_spread']:.2f}°C | "
                            f"media={mixture_opt['weighted_mean_c']:.1f}°C")
            _write_prob_sheet("Prob Comb 2", mix_opt_probs, mix_opt_info)

    # ── Foglio 6: Prob Comb 3 (mixture con top 5 modelli per Verde%) ──
    if mixture_top5:
        mix_t5_probs = mixture_top5.get("mixture_probs", {})
        if mix_t5_probs:
            mix_t5_info = (f"MIXTURE TOP5 | {mixture_top5['n_models_used']} modelli top Verde% | "
                           f"spread={mixture_top5['inter_model_spread']:.2f}°C | "
                           f"media={mixture_top5['weighted_mean_c']:.1f}°C")
            _write_prob_sheet("Prob Comb 3", mix_t5_probs, mix_t5_info)

    # ── Foglio 7: Prob Comb 4 (mixture con top 10 modelli per Verde%) ──
    if mixture_top10:
        mix_t10_probs = mixture_top10.get("mixture_probs", {})
        if mix_t10_probs:
            mix_t10_info = (f"MIXTURE TOP10 | {mixture_top10['n_models_used']} modelli top Verde% | "
                            f"spread={mixture_top10['inter_model_spread']:.2f}°C | "
                            f"media={mixture_top10['weighted_mean_c']:.1f}°C")
            _write_prob_sheet("Prob Comb 4", mix_t10_probs, mix_t10_info)

    wb.save(EXCEL_COMBINED)


def write_resolution_to_combined_excel(city: str, target_date: str, winner: str, timestamp: str):
    """Scrive la risoluzione nei fogli del file Excel combinato."""
    if not EXCEL_COMBINED.exists():
        return

    try:
        wb = openpyxl.load_workbook(EXCEL_COMBINED)
    except Exception:
        return

    section_title = f"{city} \u2014 {target_date}"
    RES_FILL = PatternFill("solid", fgColor="E2EFDA")

    for sheet_name in ["Prob Deterministici", "Prob Ensemble", "Prob Combinate", "Prob Mixture Raw", "Prob Comb 2", "Prob Comb 3", "Prob Comb 4"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        # Trova sezione
        section_row = None
        end_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == section_title:
                section_row = r
                end_row = ws.max_row
                for r2 in range(r + 1, ws.max_row + 2):
                    cell_val = ws.cell(r2, 1).value
                    if cell_val is None or str(cell_val).strip() == "":
                        end_row = r2 - 1
                        break
                    if r2 > r + 1 and isinstance(cell_val, str) and " \u2014 " in cell_val:
                        end_row = r2 - 1
                        break
                break

        if section_row is None:
            continue

        # Controlla se risoluzione gia' presente
        already_resolved = False
        for r in range(section_row, end_row + 1):
            if ws.cell(r, 1).value == "RISOLUZIONE":
                already_resolved = True
                break
        if already_resolved:
            continue

        # Leggi bucket labels dall'header (section_row + 1)
        header_row = section_row + 1
        bucket_labels = []
        col = 4
        while ws.cell(header_row, col).value is not None:
            bucket_labels.append(str(ws.cell(header_row, col).value))
            col += 1

        n_cols_res = 3 + len(bucket_labels)

        # Inserisci riga RISOLUZIONE alla fine della sezione
        res_row = end_row + 1
        ws.insert_rows(res_row, 1)

        ws.cell(res_row, 1, "RISOLUZIONE")
        ws.cell(res_row, 2, timestamp)
        ws.cell(res_row, 3, "-")
        _style_row(ws, res_row, n_cols_res + 1, fill=RES_FILL, font=Font(bold=True, size=10))

        # Evidenzia il bucket vincente
        for ci, lbl in enumerate(bucket_labels, 4):
            if lbl.strip() == winner.strip():
                ws.cell(res_row, ci, "VINCENTE")
                ws.cell(res_row, ci).fill = GREEN_FILL
                ws.cell(res_row, ci).font = Font(bold=True, color="006100", size=10)
            else:
                ws.cell(res_row, ci, "")

    try:
        wb.save(EXCEL_COMBINED)
    except Exception as e:
        log.warning(f"Errore salvataggio risoluzioni combinato: {e}")


# ═══════════════════════════════════════════════════════════════════════════════
# SNAPSHOT - CATTURA COMPLETA
# ═══════════════════════════════════════════════════════════════════════════════

def do_snapshot(market: dict, state: dict):
    """Esegue lo snapshot completo: Polymarket + Ensemble."""
    city = market["city"]
    target_date = market["target_date"]
    key = market["key"]
    snapshot_hour = market.get("snapshot_hour", 20)
    snapshot_minute = market.get("snapshot_minute", 10)
    snapshot_mode = market.get("snapshot_mode", "local")

    city_data = match_city(city)
    if not city_data:
        log.warning(f"  Coordinate non trovate per: {city}")
        return

    mode_label = " UTC" if snapshot_mode == "utc" else ""
    log.info(f"  SNAPSHOT [{snapshot_hour}:{snapshot_minute:02d}{mode_label}]: {city} - {target_date}")

    # 1. Dati Polymarket (gia' nei buckets del mercato)
    buckets = market["buckets"]
    parsed_buckets = [b.get("parsed") for b in buckets if b.get("parsed")]
    log.info(f"    Polymarket: {len(buckets)} bucket")

    # 2. Dati Ensemble
    log.info(f"    Ensemble: fetching da {len(ENSEMBLE_MODELS)} modelli...")
    members = fetch_ensemble_for_city(city_data["lat"], city_data["lon"], target_date)
    ensemble = calc_ensemble_stats(members, parsed_buckets)
    log.info(f"    Ensemble: {ensemble['n_total']} membri, {ensemble['n_models']} modelli")

    # 3. Dati Deterministici (subset ottimizzato - per backward compat)
    log.info(f"    Deterministici: fetching...")
    deterministic = do_deterministic_forecast(city, target_date, parsed_buckets)

    # 3b. Mixture Model (tutti i modelli deterministici + calibrazione)
    log.info(f"    Mixture model: avvio...")
    mixture = do_mixture_forecast(city, target_date, parsed_buckets)

    # 3c. Mixture Ottimizzato (solo modelli selezionati dall'ottimizzazione stagionale)
    log.info(f"    Mixture ottimizzato: avvio...")
    mixture_opt = do_mixture_forecast_optimized(city, target_date, parsed_buckets)

    # 3d. Mixture Top 5 (top 5 modelli per Verde% storica)
    log.info(f"    Mixture top5: avvio...")
    mixture_top5 = do_mixture_forecast_top5(city, target_date, parsed_buckets)

    # 3e. Mixture Top 10 (top 10 modelli per Verde% storica)
    log.info(f"    Mixture top10: avvio...")
    mixture_top10 = do_mixture_forecast_top10(city, target_date, parsed_buckets)

    # 4. Scrivi Excel originale
    snapshot_time = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    write_snapshot_to_excel(city, target_date, snapshot_time, snapshot_hour, snapshot_minute,
                             buckets, ensemble, snapshot_mode)
    log.info(f"    Excel dati_meteo aggiornato")

    # 5. Scrivi Excel combinato
    write_combined_to_excel(city, target_date, snapshot_time, snapshot_hour, snapshot_minute,
                             buckets, ensemble, deterministic, mixture, snapshot_mode,
                             mixture_opt=mixture_opt,
                             mixture_top5=mixture_top5,
                             mixture_top10=mixture_top10)
    log.info(f"    Excel dati_combinati aggiornato")

    # 6. Aggiorna stato
    state.setdefault("snapshots_done", {})[key] = snapshot_time
    save_state(state)

    # 7. Stampa riepilogo a console
    ens_probs = ensemble.get("combined_probs", {})
    stats = ensemble.get("combined_stats", {})
    unit = stats.get("unit", "C")
    mean_d = stats.get("mean_display", stats.get("mean", 0))
    median_d = stats.get("median_display", stats.get("median", 0))
    log.info(f"    Media ensemble: {mean_d:.1f}{unit} | "
             f"Mediana: {median_d:.1f}{unit} | "
             f"Std: {stats.get('std', 0):.1f}C")

    if deterministic:
        log.info(f"    Forecast det: {deterministic['forecast_display']:.1f}°{deterministic['unit']} "
                 f"({deterministic['method']}, verde={deterministic['verde']:.1f}%)")

    # Probabilita finali: solo mixture (deterministici con calibrazione isotonica)
    if mixture and mixture.get("mixture_probs"):
        comb_probs = mixture["mixture_probs"]
        prob_source = "MIXTURE"
    else:
        comb_probs = {}
        prob_source = "SKIP"

    for b in buckets:
        label = b["label"]
        pm = b["prob"]
        cp = comb_probs.get(label, 0)
        edge = (cp - pm) * 100
        if abs(edge) > 5:
            signal = "BUY" if edge > 0 else "SELL"
            log.info(f"    >>> [{prob_source}] {label}: Polym={pm:.0%} Mio={cp:.0%} Edge={edge:+.1f}pp {signal}")


# ═══════════════════════════════════════════════════════════════════════════════
# STATUS
# ═══════════════════════════════════════════════════════════════════════════════

def show_status(state: dict, hours: list):
    """Mostra lo stato corrente del bot."""
    def _fh(e):
        h, m = e[0], e[1]
        mode = e[2] if len(e) > 2 else "local"
        return f"{h}:{m:02d}{' UTC' if mode == 'utc' else ' loc'}"
    hours_str = ", ".join(_fh(e) for e in hours)
    print(f"\n{'='*60}")
    print(f"  POLYMARKET WEATHER BOT - STATUS")
    print(f"  Snapshot hours: {hours_str} locale")
    print(f"  Excel: {EXCEL_FILE}")
    print(f"{'='*60}")

    snapshots = state.get("snapshots_done", {})
    resolutions = state.get("resolutions_done", {})

    if snapshots:
        print(f"\n  Snapshot completati ({len(snapshots)}):")
        for key, ts in sorted(snapshots.items()):
            # Estrai base key (senza _hXX) per check risoluzione
            base_key = re.sub(r'_h\d+$', '', key)
            resolved = "RISOLTO" if base_key in resolutions else "in attesa"
            winner = resolutions.get(base_key, "")
            if winner:
                resolved += f" -> {winner}"
            print(f"    {key}: {ts} [{resolved}]")
    else:
        print("\n  Nessuno snapshot completato ancora.")

    if resolutions:
        print(f"\n  Risoluzioni ({len(resolutions)}):")
        for key, winner in sorted(resolutions.items()):
            print(f"    {key}: {winner}")

    # Prossimi snapshot
    print(f"\n  Prossimi snapshot programmati:")
    events = fetch_temp_events(closed=False)
    markets = parse_markets_from_events(events, DAYS_AHEAD)
    now = datetime.now(timezone.utc)

    upcoming = []
    for mkt in markets:
        for entry_hours in hours:
            hour, minute = entry_hours[0], entry_hours[1]
            key = f"{mkt['city']}_{mkt['target_date']}_h{hour}{minute:02d}"
            if key in snapshots:
                continue
            snap_utc = get_snapshot_utc(mkt["city"], mkt["target_date"], hour, minute)
            if snap_utc and snap_utc > now:
                upcoming.append((snap_utc, mkt["city"], mkt["target_date"], hour, minute))

    upcoming.sort()
    for snap_utc, city, target, hour, minute in upcoming[:30]:
        diff = snap_utc - now
        hrs = diff.total_seconds() / 3600
        local_tz = match_city(city)
        tz_name = local_tz["tz"] if local_tz else "UTC"
        local_time = snap_utc.astimezone(ZoneInfo(tz_name)).strftime("%H:%M %Z")
        print(f"    {city} ({target}) {hour}:{minute:02d}: {snap_utc.strftime('%Y-%m-%d %H:%M UTC')} "
              f"({local_time}) - tra {hrs:.1f}h")

    print()


# ═══════════════════════════════════════════════════════════════════════════════
# GIT AUTO-PUSH
# ═══════════════════════════════════════════════════════════════════════════════

_last_git_push = 0

def git_push_data():
    """Placeholder - git push disabilitato. I file sono serviti via web endpoint."""
    pass


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN LOOP
# ═══════════════════════════════════════════════════════════════════════════════

def run_cycle(state: dict, hours: list[int], days_ahead: int) -> int:
    """
    Esegue un ciclo completo di check. Ritorna il numero di azioni eseguite.
    """
    actions = 0

    # 1. Fetch mercati aperti
    log.info("Checking mercati Polymarket...")
    events = fetch_temp_events(closed=False)
    if not events:
        log.info("  Nessun evento trovato (API non disponibile?)")
        return 0

    markets = parse_markets_from_events(events, days_ahead)
    log.info(f"  {len(events)} eventi, {len(markets)} mercati attivi")

    # 2. Check snapshot pendenti (per ciascun orario)
    pending = get_pending_snapshots(markets, state, hours)
    if pending:
        log.info(f"  {len(pending)} snapshot da eseguire!")
        for mkt in pending:
            try:
                do_snapshot(mkt, state)
                actions += 1
            except Exception as e:
                log.error(f"  Errore snapshot {mkt['key']}: {e}")
                log.debug(traceback.format_exc())

    # 3. Check risoluzioni
    try:
        resolutions = check_resolutions(state)
        for res in resolutions:
            log.info(f"  RISOLUZIONE: {res['city']} {res['target_date']} -> {res['winner']}")
            res_timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
            write_resolution_to_excel(
                res["city"], res["target_date"], res["winner"], res_timestamp,
            )
            write_resolution_to_combined_excel(
                res["city"], res["target_date"], res["winner"], res_timestamp,
            )
            state.setdefault("resolutions_done", {})[res["key"]] = res["winner"]
            save_state(state)
            actions += 1
    except Exception as e:
        log.error(f"  Errore check risoluzioni: {e}")

    if actions > 0:
        git_push_data()

    return actions


def main_loop(hours: list[tuple[int, int]], days_ahead: int):
    """Loop principale del daemon.
    Refresh lista mercati ogni EVENT_REFRESH_INTERVAL,
    ma controlla snapshot/risoluzioni ogni CHECK_INTERVAL."""
    state = load_state()
    last_event_refresh = 0
    cached_markets = []

    # Ri-verifica risoluzioni passate con la soglia attuale all'avvio
    try:
        recheck_past_resolutions(state)
    except Exception as e:
        log.error(f"Errore ri-verifica risoluzioni: {e}")

    def _fmt_hour(entry):
        h, m = entry[0], entry[1]
        mode = entry[2] if len(entry) > 2 else "local"
        return f"{h}:{m:02d}{' UTC' if mode == 'utc' else ' loc'}"
    hours_str = ", ".join(_fmt_hour(e) for e in hours)
    log.info(f"Bot avviato | snapshot alle {hours_str} | check ogni {CHECK_INTERVAL}s")
    log.info(f"Excel: {EXCEL_FILE}")
    log.info(f"State: {STATE_FILE}")

    cycle = 0
    while True:
        try:
            now_ts = time_mod.time()

            # 1. Refresh lista mercati ogni EVENT_REFRESH_INTERVAL (o primo ciclo)
            if now_ts - last_event_refresh >= EVENT_REFRESH_INTERVAL or cycle == 0:
                log.info("Checking mercati Polymarket...")
                events = fetch_temp_events(closed=False)
                if events:
                    cached_markets = parse_markets_from_events(events, days_ahead)
                    log.info(f"  {len(events)} eventi, {len(cached_markets)} mercati attivi")
                else:
                    log.info("  Nessun evento trovato (API non disponibile?)")
                last_event_refresh = now_ts

            # 2. Check snapshot pendenti ad ogni ciclo (usa cached_markets)
            actions = 0
            pending = get_pending_snapshots(cached_markets, state, hours)
            if pending:
                log.info(f"  {len(pending)} snapshot da eseguire!")
                for mkt in pending:
                    try:
                        do_snapshot(mkt, state)
                        actions += 1
                    except Exception as e:
                        log.error(f"  Errore snapshot {mkt['key']}: {e}")
                        log.debug(traceback.format_exc())

            # 3. Check risoluzioni ad ogni ciclo
            try:
                resolutions = check_resolutions(state)
                for res in resolutions:
                    log.info(f"  RISOLUZIONE: {res['city']} {res['target_date']} -> {res['winner']}")
                    res_timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
                    write_resolution_to_excel(
                        res["city"], res["target_date"], res["winner"], res_timestamp,
                    )
                    write_resolution_to_combined_excel(
                        res["city"], res["target_date"], res["winner"], res_timestamp,
                    )
                    state.setdefault("resolutions_done", {})[res["key"]] = res["winner"]
                    save_state(state)
                    actions += 1
            except Exception as e:
                log.error(f"  Errore check risoluzioni: {e}")

            # 3b. Git push se ci sono state azioni
            if actions > 0:
                git_push_data()

            # 4. Info prossimo snapshot (solo se nessuna azione e non primo ciclo)
            if actions == 0 and cycle > 0 and cycle % 10 == 0:
                now = datetime.now(timezone.utc)
                next_snap = None
                for mkt in cached_markets:
                    for entry_hours in hours:
                        hour, minute = entry_hours[0], entry_hours[1]
                        key = f"{mkt['city']}_{mkt['target_date']}_h{hour}{minute:02d}"
                        if key in state.get("snapshots_done", {}):
                            continue
                        snap_utc = get_snapshot_utc(mkt["city"], mkt["target_date"], hour, minute)
                        if snap_utc and snap_utc > now:
                            if next_snap is None or snap_utc < next_snap[0]:
                                next_snap = (snap_utc, mkt["city"], mkt["target_date"], hour, minute)

                if next_snap:
                    diff_h = (next_snap[0] - now).total_seconds() / 3600
                    log.info(f"Prossimo snapshot: {next_snap[1]} ({next_snap[2]}) "
                             f"{next_snap[3]}:{next_snap[4]:02d} tra {diff_h:.1f}h "
                             f"({next_snap[0].strftime('%H:%M UTC')})")
                else:
                    log.info("Nessun snapshot programmato. In attesa di nuovi mercati...")

            cycle += 1
            time_mod.sleep(CHECK_INTERVAL)

        except KeyboardInterrupt:
            log.info("Bot interrotto dall'utente (Ctrl+C)")
            break
        except Exception as e:
            log.error(f"Errore nel loop principale: {e}")
            log.debug(traceback.format_exc())
            time_mod.sleep(CHECK_INTERVAL)


# ═══════════════════════════════════════════════════════════════════════════════
# WEB SERVER (per Railway - scarica Excel da browser)
# ═══════════════════════════════════════════════════════════════════════════════

class FileHandler(BaseHTTPRequestHandler):
    """Mini server per scaricare i file Excel dal browser."""

    ALLOWED_FILES = {
        "/dati_meteo.xlsx": EXCEL_FILE,
        "/dati_combinati.xlsx": EXCEL_COMBINED,
    }

    def do_GET(self):
        if self.path == "/":
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.end_headers()
            html = "<h1>Weather Trading Bot</h1><ul>"
            for name, fpath in self.ALLOWED_FILES.items():
                ok = "&#9989;" if fpath.exists() else "&#10060;"
                html += f'<li>{ok} <a href="{name}">{name.lstrip("/")}</a></li>'
            html += '<li><a href="/status">Stato bot (JSON)</a></li>'
            html += '<li><a href="/log">Ultimi log</a></li>'
            html += "</ul>"
            self.wfile.write(html.encode())

        elif self.path in self.ALLOWED_FILES:
            fpath = self.ALLOWED_FILES[self.path]
            if fpath.exists():
                self.send_response(200)
                self.send_header("Content-Type",
                                 "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition",
                                 f'attachment; filename="{fpath.name}"')
                self.end_headers()
                self.wfile.write(fpath.read_bytes())
            else:
                self.send_error(404, "File non ancora generato")

        elif self.path == "/status":
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.end_headers()
            state = load_state()
            self.wfile.write(json.dumps(state, indent=2, ensure_ascii=False).encode())

        elif self.path == "/log":
            self.send_response(200)
            self.send_header("Content-Type", "text/plain; charset=utf-8")
            self.end_headers()
            if LOG_FILE.exists():
                lines = LOG_FILE.read_text(encoding="utf-8", errors="replace").splitlines()[-100:]
                self.wfile.write("\n".join(lines).encode())
            else:
                self.wfile.write(b"Nessun log disponibile")
        else:
            self.send_error(404)

    def log_message(self, format, *args):
        pass  # No spam nei log


def start_web_server():
    port = int(os.environ.get("PORT", 8080))
    server = HTTPServer(("0.0.0.0", port), FileHandler)
    log.info(f"Web server avviato su porta {port}")
    server.serve_forever()


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    def _fmt_default(entry):
        h, m = entry[0], entry[1]
        mode = entry[2] if len(entry) > 2 else "local"
        suffix = "u" if mode == "utc" else ""
        return f"{h}:{m:02d}{suffix}"
    default_hours = ",".join(_fmt_default(e) for e in SNAPSHOT_HOURS)

    parser = argparse.ArgumentParser(
        description="Polymarket Weather Trading Bot - Daemon 24/7",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Esempi:
  python bot.py                       # avvia con orari default (UTC fissi)
  python bot.py --hours 17:10u,20:10u # snapshot alle 17:10 e 20:10 UTC
  python bot.py --hours 16:10,20:10   # snapshot alle 16:10 e 20:10 locali
  python bot.py --once                # esegui un solo ciclo
  python bot.py --status              # mostra stato corrente
  Suffisso 'u' = UTC fisso, senza suffisso = ora locale della citta
        """,
    )
    parser.add_argument("--hours", type=str, default=default_hours,
                        help=f"Orari snapshot HH:MM[u] separati da virgola, u=UTC (default: {default_hours})")
    parser.add_argument("--once", action="store_true",
                        help="Esegui un solo ciclo e esci")
    parser.add_argument("--status", action="store_true",
                        help="Mostra stato corrente e esci")
    parser.add_argument("--days", type=int, default=DAYS_AHEAD,
                        help=f"Mercati fino a N giorni avanti (default: {DAYS_AHEAD})")
    args = parser.parse_args()

    hours = []
    for part in args.hours.split(","):
        part = part.strip()
        mode = "local"
        if part.endswith("u"):
            mode = "utc"
            part = part[:-1]
        if ":" in part:
            h, m = part.split(":")
            hours.append((int(h), int(m), mode))
        else:
            hours.append((int(part), 0, mode))
    hours.sort()
    def _fmt(e):
        h, m = e[0], e[1]
        md = e[2] if len(e) > 2 else "local"
        return f"{h}:{m:02d}{' UTC' if md == 'utc' else ' loc'}"
    hours_str = ", ".join(_fmt(e) for e in hours)

    state = load_state()

    if args.status:
        show_status(state, hours)
        return

    if args.once:
        log.info(f"=== Modalita singolo ciclo (--once) | orari={hours_str} ===")
        run_cycle(state, hours, args.days)
        return

    # Avvia web server in background se su Railway (PORT impostata)
    if os.environ.get("PORT"):
        threading.Thread(target=start_web_server, daemon=True).start()

    # Daemon mode
    log.info("=" * 60)
    log.info("  POLYMARKET WEATHER TRADING BOT")
    log.info(f"  Snapshot: {hours_str} locale, giorno prima del mercato")
    log.info(f"  Mercati: fino a {args.days} giorni avanti")
    log.info(f"  Check interval: ogni {CHECK_INTERVAL}s")
    if os.environ.get("DATA_DIR"):
        log.info(f"  Data dir: {DATA_DIR}")
    log.info("=" * 60)
    main_loop(hours, args.days)


if __name__ == "__main__":
    main()
