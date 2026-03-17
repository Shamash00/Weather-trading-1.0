import pandas as pd
import requests
import time
import os
from datetime import datetime, timedelta
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
storico_path = os.path.join(BASE_DIR, "Temperature Storiche", "Temperature Storiche FM-15 Tutte le Citta 2021-2025 Daily Max.xlsx")
out_path = os.path.join(BASE_DIR, "Confronto Tutti Modelli vs Storiche Daily Max.xlsx")

CITIES = {
    "Ankara":      {"lat": 40.128082, "lon": 32.995083},
    "Atlanta":     {"lat": 33.62972,  "lon": -84.44224},
    "BuenosAires": {"lat": -34.822222,"lon": -58.535833},
    "Chicago":     {"lat": 41.96019,  "lon": -87.93162},
    "Dallas":      {"lat": 32.8519,   "lon": -96.8555},
    "Londra":      {"lat": 51.505278, "lon": 0.055278},
    "Lucknow":     {"lat": 26.7606,   "lon": 80.8893},
    "Miami":       {"lat": 25.78805,  "lon": -80.31694},
    "Monaco":      {"lat": 48.353783, "lon": 11.786086},
    "New York":    {"lat": 40.77945,  "lon": -73.88027},
    "Parigi":      {"lat": 49.012779, "lon": 2.55},
    "SaoPaulo":    {"lat": -23.432075,"lon": -46.469511},
    "Seattle":     {"lat": 47.4444,   "lon": -122.3138},
    "Seoul":       {"lat": 37.469075, "lon": 126.450517},
    "Shanghai":    {"lat": 31.143378, "lon": 121.805214},
    "Singapore":   {"lat": 1.350189,  "lon": 103.994433},
    "TelAviv":     {"lat": 32.011389, "lon": 34.886667},
    "Tokyo":       {"lat": 35.552258, "lon": 139.779694},
    "Toronto":     {"lat": 43.677223, "lon": -79.630556},
    "Wellington":  {"lat": -41.3333333,"lon": 174.8},
}

MODELS = ("ecmwf_ifs,ecmwf_ifs025,ecmwf_aifs025_single,cma_grapes_global,"
          "bom_access_global,gfs_seamless,gfs_global,gfs_hrrr,ncep_nbm_conus,"
          "ncep_nam_conus,gfs_graphcast025,ncep_aigfs025,ncep_hgefs025_ensemble_mean,"
          "jma_seamless,jma_msm,jma_gsm,kma_seamless,kma_ldps,kma_gdps,"
          "italia_meteo_arpae_icon_2i,meteoswiss_icon_seamless,meteoswiss_icon_ch1,"
          "meteoswiss_icon_ch2,meteofrance_seamless,meteofrance_arpege_world,"
          "meteofrance_arpege_europe,meteofrance_arome_france,meteofrance_arome_france_hd,"
          "ukmo_seamless,ukmo_global_deterministic_10km,gem_seamless,"
          "ukmo_uk_deterministic_2km,gem_global,gem_regional,gem_hrdps_continental,"
          "gem_hrdps_west,knmi_seamless,knmi_harmonie_arome_europe,"
          "knmi_harmonie_arome_netherlands,dmi_seamless,dmi_harmonie_arome_europe,"
          "metno_nordic,icon_eu,icon_global,icon_seamless,metno_seamless,icon_d2,"
          "best_match")

API_URL = "https://previous-runs-api.open-meteo.com/v1/forecast"
CHUNK_DAYS = 500
CACHE_DIR = os.path.join(BASE_DIR, "_cache_modelli")
os.makedirs(CACHE_DIR, exist_ok=True)

green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

api_call_count = 0


def fetch_city_data(city_name, lat, lon, start_date, end_date):
    """Fetch all model forecasts for a city in chunks, with caching."""
    global api_call_count
    cache_file = os.path.join(CACHE_DIR, f"{city_name}.pkl")
    if os.path.exists(cache_file):
        print(f"  Cache trovata, caricamento da disco...")
        return pd.read_pickle(cache_file)

    all_chunks = []
    current = start_date
    chunk_num = 0
    total_chunks = ((end_date - start_date).days // CHUNK_DAYS) + 1

    while current <= end_date:
        chunk_end = min(current + timedelta(days=CHUNK_DAYS - 1), end_date)
        chunk_num += 1
        s = current.strftime("%Y-%m-%d")
        e = chunk_end.strftime("%Y-%m-%d")

        success = False
        for attempt in range(2):
            try:
                api_call_count += 1
                r = requests.get(API_URL, params={
                    "latitude": lat, "longitude": lon,
                    "hourly": "temperature_2m_previous_day1,temperature_2m_previous_day2",
                    "models": MODELS,
                    "timezone": "auto",
                    "temperature_unit": "fahrenheit",
                    "start_date": s, "end_date": e,
                }, timeout=240)
                if r.status_code == 429:
                    if attempt == 0:
                        print(f"    chunk {chunk_num}/{total_chunks} ({s}->{e}): rate limit (call #{api_call_count}), attendo 10 min...")
                        time.sleep(600)
                        continue
                    else:
                        print(f"    chunk {chunk_num}/{total_chunks}: ancora rate limit, salvo progresso e esco.")
                        # Save what we have so far
                        if all_chunks:
                            partial = pd.concat(all_chunks, ignore_index=True)
                            partial.to_pickle(cache_file + ".partial")
                            print(f"    Salvato parziale: {cache_file}.partial")
                        return None
                r.raise_for_status()
                data = r.json()
                chunk_df = pd.DataFrame(data["hourly"])
                chunk_df["time"] = pd.to_datetime(chunk_df["time"])
                all_chunks.append(chunk_df)
                print(f"    chunk {chunk_num}/{total_chunks} ({s}->{e}): OK (call #{api_call_count})")
                success = True
                break
            except requests.exceptions.HTTPError:
                print(f"    chunk {chunk_num}/{total_chunks}: HTTP {r.status_code} (call #{api_call_count})")
                if attempt == 0:
                    time.sleep(300)
                    continue
                break
            except Exception as ex:
                print(f"    chunk {chunk_num}/{total_chunks}: {ex} (call #{api_call_count})")
                if attempt == 0:
                    time.sleep(60)
                    continue
                break

        if not success:
            print(f"    SALTATO chunk {chunk_num}")

        # Rate limit management: pause every 8 calls
        if api_call_count > 0 and api_call_count % 8 == 0:
            print(f"    === Pausa preventiva dopo {api_call_count} chiamate (5 min) ===")
            time.sleep(300)
        else:
            time.sleep(20)
        current = chunk_end + timedelta(days=1)

    if not all_chunks:
        return None

    result = pd.concat(all_chunks, ignore_index=True)
    result.to_pickle(cache_file)
    print(f"  Salvato cache: {cache_file}")
    return result


def compute_daily_max(hourly_df):
    """From hourly data, compute daily max for each model column."""
    hourly_df["Data"] = hourly_df["time"].dt.date

    prev1_cols = [c for c in hourly_df.columns
                  if c.startswith("temperature_2m_previous_day1_")]
    prev2_cols = [c for c in hourly_df.columns
                  if c.startswith("temperature_2m_previous_day2_")]

    agg_dict = {c: "max" for c in prev1_cols + prev2_cols}
    daily = hourly_df.groupby("Data").agg(agg_dict).reset_index()

    rename = {}
    for c in prev1_cols:
        rename[c] = f"1GG_{c.replace('temperature_2m_previous_day1_', '')}"
    for c in prev2_cols:
        rename[c] = f"2GG_{c.replace('temperature_2m_previous_day2_', '')}"
    daily.rename(columns=rename, inplace=True)

    return daily


# ── Main ──
print("Caricamento temperature storiche...")
xl_st = pd.ExcelFile(storico_path)

city_results = {}

for city_name, coords in CITIES.items():
    print(f"\n{'='*60}")
    print(f"{city_name} (lat={coords['lat']}, lon={coords['lon']})")
    print(f"{'='*60}")

    if city_name not in xl_st.sheet_names:
        print(f"  ATTENZIONE: {city_name} non presente nei dati storici, saltato.")
        continue

    df_st = pd.read_excel(xl_st, sheet_name=city_name)
    df_st["Data"] = pd.to_datetime(df_st["Data"]).dt.date
    start_date = min(df_st["Data"])
    end_date = max(df_st["Data"])
    print(f"  Periodo storico: {start_date} -> {end_date}")

    print(f"  Fetching API data...")
    raw_df = fetch_city_data(city_name, coords["lat"], coords["lon"],
                              datetime.strptime(str(start_date), "%Y-%m-%d"),
                              datetime.strptime(str(end_date), "%Y-%m-%d"))

    if raw_df is None:
        print(f"  ERRORE: nessun dato ricevuto, saltato.")
        continue

    daily = compute_daily_max(raw_df)

    merged = pd.merge(daily, df_st[["Data", "Max_Temperatura_F"]], on="Data", how="inner")

    model_1gg = sorted([c for c in merged.columns if c.startswith("1GG_")])
    model_2gg = sorted([c for c in merged.columns if c.startswith("2GG_")])

    min_valid = len(merged) * 0.10
    model_1gg = [c for c in model_1gg if merged[c].notna().sum() >= min_valid]
    model_2gg = [c for c in model_2gg if merged[c].notna().sum() >= min_valid]

    out = pd.DataFrame()
    out["Data"] = merged["Data"]

    # ── Colonne Celsius ──
    reg_c = ((merged["Max_Temperatura_F"] - 32) * 5 / 9).round(0)
    out["Registrata_C"] = reg_c

    for c in model_1gg:
        model_name = c.replace("1GG_", "")
        stima_c = ((merged[c].round(0) - 32) * 5 / 9).round(0)
        out[f"D1c_{model_name}"] = stima_c - reg_c

    for c in model_2gg:
        model_name = c.replace("2GG_", "")
        stima_c = ((merged[c].round(0) - 32) * 5 / 9).round(0)
        out[f"D2c_{model_name}"] = stima_c - reg_c

    # ── Colonne Fahrenheit ──
    out["Registrata_F"] = merged["Max_Temperatura_F"]

    for c in model_1gg:
        model_name = c.replace("1GG_", "")
        forecast_val = merged[c].round(0)
        out[f"D1_{model_name}"] = forecast_val - merged["Max_Temperatura_F"]

    for c in model_2gg:
        model_name = c.replace("2GG_", "")
        forecast_val = merged[c].round(0)
        out[f"D2_{model_name}"] = forecast_val - merged["Max_Temperatura_F"]

    city_results[city_name] = out
    n_models_1 = len(model_1gg)
    n_models_2 = len(model_2gg)
    print(f"  {len(out)} giorni, {n_models_1} modelli 1GG, {n_models_2} modelli 2GG")


# ── Write Excel ──
if not city_results:
    print("\nNessun risultato da scrivere.")
else:
    print(f"\n{'='*60}")
    print("Scrittura Excel...")
    bold = Font(bold=True)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for city_name, out in city_results.items():
            print(f"  {city_name}...")
            out.to_excel(writer, sheet_name=city_name, index=False)

            ws = writer.sheets[city_name]
            max_row = ws.max_row
            n_cols = ws.max_column

            # Identifica colonne delta F e C
            delta_f_cols = []
            delta_c_cols = []
            for col_idx in range(3, n_cols + 1):
                header = ws.cell(row=1, column=col_idx).value
                if header and (header.startswith("D1_") or header.startswith("D2_")):
                    delta_f_cols.append(col_idx)
                elif header and (header.startswith("D1c_") or header.startswith("D2c_")):
                    delta_c_cols.append(col_idx)

            # Conditional formatting Fahrenheit: verde ≤1, rosso >1
            for ci in delta_f_cols:
                col_letter = get_column_letter(ci)
                rng = f"{col_letter}2:{col_letter}{max_row}"
                ws.conditional_formatting.add(rng, FormulaRule(
                    formula=[f'AND({col_letter}2<>"", ABS({col_letter}2)<=1)'],
                    fill=green_fill, stopIfTrue=True))
                ws.conditional_formatting.add(rng, FormulaRule(
                    formula=[f'AND({col_letter}2<>"", ABS({col_letter}2)>1)'],
                    fill=red_fill, stopIfTrue=True))

            # Conditional formatting Celsius: verde ≤1, rosso >1
            for ci in delta_c_cols:
                col_letter = get_column_letter(ci)
                rng = f"{col_letter}2:{col_letter}{max_row}"
                ws.conditional_formatting.add(rng, FormulaRule(
                    formula=[f'AND({col_letter}2<>"", ABS({col_letter}2)<=1)'],
                    fill=green_fill, stopIfTrue=True))
                ws.conditional_formatting.add(rng, FormulaRule(
                    formula=[f'AND({col_letter}2<>"", ABS({col_letter}2)>1)'],
                    fill=red_fill, stopIfTrue=True))

            # Statistiche riassuntive
            summary_row = max_row + 2
            for i, label in enumerate(["Verde %", "Rosso %",
                                        "Delta Medio", "Errore Medio Abs"]):
                cell = ws.cell(row=summary_row + i, column=1, value=label)
                cell.font = bold

            # Stats Fahrenheit (soglie: verde ≤1, rosso >1)
            for ci in delta_f_cols:
                col_name = ws.cell(row=1, column=ci).value
                if col_name not in out.columns:
                    continue
                vals = out[col_name].dropna()
                if len(vals) == 0:
                    continue
                n = len(vals)
                abs_vals = vals.abs()
                g = (abs_vals <= 1).sum() / n * 100
                r = (abs_vals > 1).sum() / n * 100
                mean_d = round(float(vals.mean()), 1)
                mae = round(float(abs_vals.mean()), 1)

                cg = ws.cell(row=summary_row, column=ci, value=round(g, 1))
                cg.fill = green_fill; cg.font = bold; cg.number_format = '0.0"%"'
                cr = ws.cell(row=summary_row + 1, column=ci, value=round(r, 1))
                cr.fill = red_fill; cr.font = bold; cr.number_format = '0.0"%"'
                cm = ws.cell(row=summary_row + 2, column=ci, value=mean_d)
                cm.font = bold; cm.number_format = '0.0'
                ca = ws.cell(row=summary_row + 3, column=ci, value=mae)
                ca.font = bold; ca.number_format = '0.0'

            # Stats Celsius (soglie: verde ≤1, rosso >1)
            for ci in delta_c_cols:
                col_name = ws.cell(row=1, column=ci).value
                if col_name not in out.columns:
                    continue
                vals = out[col_name].dropna()
                if len(vals) == 0:
                    continue
                n = len(vals)
                abs_vals = vals.abs()
                g = (abs_vals <= 1).sum() / n * 100
                r = (abs_vals > 1).sum() / n * 100
                mean_d = round(float(vals.mean()), 1)
                mae = round(float(abs_vals.mean()), 1)

                cg = ws.cell(row=summary_row, column=ci, value=round(g, 1))
                cg.fill = green_fill; cg.font = bold; cg.number_format = '0.0"%"'
                cr = ws.cell(row=summary_row + 1, column=ci, value=round(r, 1))
                cr.fill = red_fill; cr.font = bold; cr.number_format = '0.0"%"'
                cm = ws.cell(row=summary_row + 2, column=ci, value=mean_d)
                cm.font = bold; cm.number_format = '0.0'
                ca = ws.cell(row=summary_row + 3, column=ci, value=mae)
                ca.font = bold; ca.number_format = '0.0'

        # ── Foglio Riepilogo ──
        print("  Riepilogo...")
        ws_r = writer.book.create_sheet("Riepilogo")
        row_r = 1
        all_years = [2021, 2022, 2023, 2024, 2025]

        for city_name, out in city_results.items():
            dates = pd.to_datetime(out["Data"])
            out_years = dates.dt.year

            # City header
            cell = ws_r.cell(row=row_r, column=1, value=city_name)
            cell.font = Font(bold=True, size=14)
            row_r += 1

            d1c = [c for c in out.columns if c.startswith("D1c_")]
            d2c = [c for c in out.columns if c.startswith("D2c_")]
            d1f = [c for c in out.columns if c.startswith("D1_")]
            d2f = [c for c in out.columns if c.startswith("D2_")]

            for section, cols_c, cols_f in [("1GG", d1c, d1f), ("2GG", d2c, d2f)]:
                if not cols_c and not cols_f:
                    continue

                label = "Previsione 1 Giorno" if section == "1GG" else "Previsione 2 Giorni"
                cell = ws_r.cell(row=row_r, column=1, value=label)
                cell.font = Font(bold=True, size=11)
                row_r += 1

                # Headers
                headers = ["Modello"]
                for y in all_years:
                    headers.append(f"V%C {y}")
                headers += ["V%C Tot", "R%C Tot"]
                for y in all_years:
                    headers.append(f"V%F {y}")
                headers += ["V%F Tot", "R%F Tot"]

                for j, h in enumerate(headers):
                    cell = ws_r.cell(row=row_r, column=j + 1, value=h)
                    cell.font = bold
                row_r += 1

                # Model names
                prefix_c = "D1c_" if section == "1GG" else "D2c_"
                prefix_f = "D1_" if section == "1GG" else "D2_"
                models_c = {c[len(prefix_c):] for c in cols_c}
                models_f = {c[len(prefix_f):] for c in cols_f}
                all_models = sorted(models_c | models_f)

                model_data = []
                for model in all_models:
                    md = {"model": model}

                    col_c = f"{prefix_c}{model}"
                    if col_c in out.columns:
                        vals = out[col_c]
                        for y in all_years:
                            mask = (out_years == y) & vals.notna()
                            yv = vals[mask]
                            if len(yv) > 0:
                                md[f"VC_{y}"] = round((yv.abs() <= 1).sum() / len(yv) * 100, 1)
                        all_valid = vals.dropna()
                        if len(all_valid) > 0:
                            md["VC_Tot"] = round((all_valid.abs() <= 1).sum() / len(all_valid) * 100, 1)
                            md["RC_Tot"] = round(100 - md["VC_Tot"], 1)

                    col_f = f"{prefix_f}{model}"
                    if col_f in out.columns:
                        vals = out[col_f]
                        for y in all_years:
                            mask = (out_years == y) & vals.notna()
                            yv = vals[mask]
                            if len(yv) > 0:
                                md[f"VF_{y}"] = round((yv.abs() <= 1).sum() / len(yv) * 100, 1)
                        all_valid = vals.dropna()
                        if len(all_valid) > 0:
                            md["VF_Tot"] = round((all_valid.abs() <= 1).sum() / len(all_valid) * 100, 1)
                            md["RF_Tot"] = round(100 - md["VF_Tot"], 1)

                    model_data.append(md)

                model_data.sort(key=lambda x: (x.get("VC_Tot", 0), x.get("VF_Tot", 0)), reverse=True)

                for md in model_data:
                    ws_r.cell(row=row_r, column=1, value=md["model"])
                    col_idx = 2
                    # V%C per year
                    for y in all_years:
                        val = md.get(f"VC_{y}")
                        if val is not None:
                            c = ws_r.cell(row=row_r, column=col_idx, value=val)
                            c.number_format = '0.0'
                        col_idx += 1
                    # V%C Tot
                    val = md.get("VC_Tot")
                    if val is not None:
                        c = ws_r.cell(row=row_r, column=col_idx, value=val)
                        c.number_format = '0.0'; c.font = bold; c.fill = green_fill
                    col_idx += 1
                    # R%C Tot
                    val = md.get("RC_Tot")
                    if val is not None:
                        c = ws_r.cell(row=row_r, column=col_idx, value=val)
                        c.number_format = '0.0'; c.font = bold; c.fill = red_fill
                    col_idx += 1
                    # V%F per year
                    for y in all_years:
                        val = md.get(f"VF_{y}")
                        if val is not None:
                            c = ws_r.cell(row=row_r, column=col_idx, value=val)
                            c.number_format = '0.0'
                        col_idx += 1
                    # V%F Tot
                    val = md.get("VF_Tot")
                    if val is not None:
                        c = ws_r.cell(row=row_r, column=col_idx, value=val)
                        c.number_format = '0.0'; c.font = bold; c.fill = green_fill
                    col_idx += 1
                    # R%F Tot
                    val = md.get("RF_Tot")
                    if val is not None:
                        c = ws_r.cell(row=row_r, column=col_idx, value=val)
                        c.number_format = '0.0'; c.font = bold; c.fill = red_fill
                    row_r += 1

                row_r += 1  # blank row between 1GG/2GG
            row_r += 1  # blank row between cities

    print(f"\nFile salvato: {out_path}")
