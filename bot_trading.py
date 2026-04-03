"""
Polymarket Weather Trading Bot - LIVE TRADING
Strategia: COMBtop2

Usa il mixture model calibrato (deterministici + isotonica) per calcolare
le probabilita reali di ogni bucket di temperatura. Compra YES solo su
bucket sottovalutati dal mercato, applicando filtri rigorosi:

  - p <= 10%     : solo bucket con prezzo Polymarket <= 10 centesimi
  - mxP < 55%    : la prob massima del modello deve essere < 55%
  - dist >= 1    : almeno 1 posizione di distanza dal picco del modello
  - spr >= 0.6   : spread inter-modello >= 0.6°C
  - Top 2        : max 2 scommesse per mercato (i 2 edge piu alti)

Uso:
    python bot_trading.py                     # avvia in dry-run (default)
    python bot_trading.py --live              # piazza ordini reali
    python bot_trading.py --once              # un solo ciclo e esci
    python bot_trading.py --min-edge 8        # edge minimo 8pp (default: 5)
    python bot_trading.py --max-bet 50        # puntata max $50 (default: $20)

Variabili d'ambiente (richieste solo per --live):
    POLY_API_KEY        API key Polymarket
    POLY_API_SECRET     API secret Polymarket
    POLY_PASSPHRASE     Passphrase Polymarket
    POLY_PRIVATE_KEY    Private key wallet Polygon
"""

import requests
import json
import time as time_mod
import logging
import argparse
import numpy as np
import os
import re
import sys
import traceback
from datetime import datetime, timezone, timedelta, date
from pathlib import Path
from scipy.stats import norm

# ═══════════════════════════════════════════════════════════════════════════════
# CONFIG
# ═══════════════════════════════════════════════════════════════════════════════

TRADE_HOUR_UTC = 17         # Ora UTC in cui fare previsioni e piazzare ordini
TRADE_MINUTE_UTC = 10       # Minuto UTC
DAYS_AHEAD = 1              # Considera solo mercati del giorno dopo (domani)

# ── Strategia COMBtop2 ───────────────────────────────────────────────────────
TOP_N = 2                   # Massimo N bucket per mercato su cui puntare
MAX_MARKET_PRICE = 0.10     # Prezzo Polymarket massimo (p <= 10%)
MAX_POLY_PEAK = 0.55        # Se bestAsk piu alta di Polymarket >= 55%, skippa mercato
SKIP_IF_SAME_PEAK = True    # Skippa mercato se picco modello = picco Polymarket
MIN_SPREAD = 0.6            # Spread inter-modello minimo in °C

# ── Limiti di rischio ────────────────────────────────────────────────────────
MIN_EDGE_PP = 0.7           # Edge minimo 0.7pp per scommettere
MIN_MODEL_PROB = 0.015      # Prob minima del modello (1.5%) per selezionare un bucket
BET_SIZE_USD = 1.0          # Puntata fissa per ogni mercato ($)
PRICE_TOLERANCE_HIGH = 0.01 # Tolleranza prezzo: +1% se prezzo mercato > 5%
PRICE_TOLERANCE_LOW = 0.005 # Tolleranza prezzo: +0.5% se prezzo mercato <= 5%
PRICE_TOLERANCE_THRESHOLD = 0.05  # Soglia per decidere quale tolleranza applicare
ORDER_EXPIRATION_SECS = 6 * 3600        # Scadenza limit order: 6 ore (America/Europa)
ORDER_EXPIRATION_SECS_ASIA = int(2.5 * 3600)  # Scadenza limit order: 2.5 ore (Asia)

# API endpoints
GAMMA_API = "https://gamma-api.polymarket.com"
CLOB_API = "https://clob.polymarket.com"
DETERMINISTIC_API = "https://api.open-meteo.com/v1/forecast"

BASE_DIR = Path(__file__).parent
OPTIMIZATION_FILE = BASE_DIR / "Ottimizzazione Ensemble Stagioni.xlsx"
TOP_MODELS_FILE_V2 = BASE_DIR / "Top Modelli Deterministici per Citta.xlsx"
MODEL_STATS_FILE = BASE_DIR / "model_error_stats.pkl"
STATE_FILE = BASE_DIR / "trading_state.json"

# ── Modelli ──────────────────────────────────────────────────────────────────

ALL_DETERMINISTIC_MODELS = [
    "best_match",
    "bom_access_global",
    "ecmwf_aifs025_single", "ecmwf_ifs025",
    "gem_global", "gem_regional",
    "gfs_global", "gfs_graphcast025", "gfs_hrrr",
    "icon_d2", "icon_eu", "icon_global",
    "jma_gsm",
    "kma_gdps",
    "knmi_harmonie_arome_europe", "knmi_harmonie_arome_netherlands",
    "dmi_harmonie_arome_europe",
    "meteofrance_arome_france", "meteofrance_arome_france_hd",
    "meteofrance_arpege_europe", "meteofrance_arpege_world",
    "metno_seamless",
    "ukmo_global_deterministic_10km", "ukmo_uk_deterministic_2km",
]

# ── Coordinate stazioni + timezone ───────────────────────────────────────────

CITY_DATA = {
    "Ankara":       {"lat": 40.128082,  "lon": 32.995083,   "tz": "Europe/Istanbul"},
    "Atlanta":      {"lat": 33.62972,   "lon": -84.44224,   "tz": "America/New_York"},
    "Austin":       {"lat": 30.18311,   "lon": -97.67989,   "tz": "America/Chicago"},
    "Beijing":      {"lat": 40.080111,  "lon": 116.584556,  "tz": "Asia/Shanghai"},
    "Buenos Aires": {"lat": -34.822222, "lon": -58.535833,  "tz": "America/Argentina/Buenos_Aires"},
    "Chengdu":      {"lat": 30.578528,  "lon": 103.947086,  "tz": "Asia/Shanghai"},
    "Chicago":      {"lat": 41.96019,   "lon": -87.93162,   "tz": "America/Chicago"},
    "Chongqing":    {"lat": 29.719217,  "lon": 106.641678,  "tz": "Asia/Shanghai"},
    "Dallas":       {"lat": 32.8519,    "lon": -96.8555,    "tz": "America/Chicago"},
    "Denver":       {"lat": 39.84657,   "lon": -104.65623,  "tz": "America/Denver"},
    "Houston":      {"lat": 29.64586,   "lon": -95.28212,   "tz": "America/Chicago"},
    "London":       {"lat": 51.505278,  "lon": 0.055278,    "tz": "Europe/London"},
    "Lucknow":      {"lat": 26.7606,    "lon": 80.8893,     "tz": "Asia/Kolkata"},
    "Madrid":       {"lat": 40.493556,  "lon": -3.566764,   "tz": "Europe/Madrid"},
    "Miami":        {"lat": 25.78805,   "lon": -80.31694,   "tz": "America/New_York"},
    "Milan":        {"lat": 45.630606,  "lon": 8.728111,    "tz": "Europe/Rome"},
    "Milano":       {"lat": 45.630606,  "lon": 8.728111,    "tz": "Europe/Rome"},
    "Monaco":       {"lat": 48.353783,  "lon": 11.786086,   "tz": "Europe/Berlin"},
    "Munich":       {"lat": 48.353783,  "lon": 11.786086,   "tz": "Europe/Berlin"},
    "New York":     {"lat": 40.77945,   "lon": -73.88027,   "tz": "America/New_York"},
    "NYC":          {"lat": 40.77945,   "lon": -73.88027,   "tz": "America/New_York"},
    "Paris":        {"lat": 49.012779,  "lon": 2.55,        "tz": "Europe/Paris"},
    "San Francisco":{"lat": 37.61962,   "lon": -122.36562,  "tz": "America/Los_Angeles"},
    "Sao Paulo":    {"lat": -23.432075, "lon": -46.469511,  "tz": "America/Sao_Paulo"},
    "São Paulo":    {"lat": -23.432075, "lon": -46.469511,  "tz": "America/Sao_Paulo"},
    "Seattle":      {"lat": 47.4444,    "lon": -122.3138,   "tz": "America/Los_Angeles"},
    "Seoul":        {"lat": 37.469075,  "lon": 126.450517,  "tz": "Asia/Seoul"},
    "Shanghai":     {"lat": 31.143378,  "lon": 121.805214,  "tz": "Asia/Shanghai"},
    "Shenzhen":     {"lat": 22.639258,  "lon": 113.810664,  "tz": "Asia/Shanghai"},
    "Singapore":    {"lat": 1.350189,   "lon": 103.994433,  "tz": "Asia/Singapore"},
    "Taipei":       {"lat": 25.077731,  "lon": 121.232822,  "tz": "Asia/Taipei"},
    "Tel Aviv":     {"lat": 32.011389,  "lon": 34.886667,   "tz": "Asia/Jerusalem"},
    "Tokyo":        {"lat": 35.552258,  "lon": 139.779694,  "tz": "Asia/Tokyo"},
    "Toronto":      {"lat": 43.677223,  "lon": -79.630556,  "tz": "America/Toronto"},
    "Warsaw":       {"lat": 52.16575,   "lon": 20.967122,   "tz": "Europe/Warsaw"},
    "Wellington":   {"lat": -41.333333, "lon": 174.8,       "tz": "Pacific/Auckland"},
    "Wuhan":        {"lat": 30.783758,  "lon": 114.2081,    "tz": "Asia/Shanghai"},
}

CITY_NAME_TO_OPT = {
    "Ankara": "Ankara", "Atlanta": "Atlanta", "Austin": "Austin",
    "Beijing": "Beijing", "Buenos Aires": "BuenosAires",
    "Chengdu": "Chengdu", "Chicago": "Chicago", "Chongqing": "Chongqing",
    "Dallas": "Dallas", "Denver": "Denver", "Houston": "Houston",
    "London": "Londra", "Lucknow": "Lucknow", "Madrid": "Madrid",
    "Miami": "Miami", "Milan": "Milano", "Milano": "Milano",
    "Munich": "Monaco", "Monaco": "Monaco",
    "New York": "New York", "NYC": "New York",
    "Paris": "Parigi",
    "San Francisco": "SanFrancisco",
    "Sao Paulo": "SaoPaulo", "São Paulo": "SaoPaulo",
    "Seattle": "Seattle", "Seoul": "Seoul",
    "Shanghai": "Shanghai", "Shenzhen": "Shenzhen",
    "Singapore": "Singapore", "Taipei": "Taipei",
    "Tel Aviv": "TelAviv", "Tokyo": "Tokyo",
    "Toronto": "Toronto", "Warsaw": "Warsaw",
    "Wellington": "Wellington", "Wuhan": "Wuhan",
}

SOUTHERN_HEMISPHERE_CITIES = {"BuenosAires", "SaoPaulo", "Wellington"}
IGNORED_CITIES = {"Hong Kong", "Los Angeles"}


# ═══════════════════════════════════════════════════════════════════════════════
# LOGGING
# ═══════════════════════════════════════════════════════════════════════════════

def setup_logging():
    logger = logging.getLogger("trading")
    logger.setLevel(logging.INFO)
    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", datefmt="%Y-%m-%d %H:%M:%S")
    ch = logging.StreamHandler(sys.stdout)
    ch.setFormatter(fmt)
    logger.addHandler(ch)
    fh = logging.FileHandler(BASE_DIR / "trading.log", encoding="utf-8")
    fh.setFormatter(fmt)
    logger.addHandler(fh)
    return logger

log = setup_logging()


# ═══════════════════════════════════════════════════════════════════════════════
# STATE
# ═══════════════════════════════════════════════════════════════════════════════

def load_state() -> dict:
    if STATE_FILE.exists():
        with open(STATE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"orders": [], "total_exposure": 0.0, "pnl": 0.0}


def save_state(state: dict):
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(state, f, indent=2, ensure_ascii=False, default=str)


# ═══════════════════════════════════════════════════════════════════════════════
# UTILS
# ═══════════════════════════════════════════════════════════════════════════════

def get_season(month: int, opt_city: str) -> str:
    if opt_city in SOUTHERN_HEMISPHERE_CITIES:
        if month in (12, 1, 2): return "Estate"
        elif month in (3, 4, 5): return "Autunno"
        elif month in (6, 7, 8): return "Inverno"
        else: return "Primavera"
    else:
        if month in (12, 1, 2): return "Inverno"
        elif month in (3, 4, 5): return "Primavera"
        elif month in (6, 7, 8): return "Estate"
        else: return "Autunno"


def match_city(city_name: str) -> dict | None:
    if city_name in IGNORED_CITIES:
        return None
    if city_name in CITY_DATA:
        return CITY_DATA[city_name]
    for key, data in CITY_DATA.items():
        if key.lower() == city_name.lower():
            return data
    for key, data in CITY_DATA.items():
        if key.lower() in city_name.lower() or city_name.lower() in key.lower():
            return data
    return None


# ═══════════════════════════════════════════════════════════════════════════════
# CONFIG LOADERS (stesse del bot.py)
# ═══════════════════════════════════════════════════════════════════════════════

_det_config_cache = None

def load_deterministic_config() -> dict:
    global _det_config_cache
    if _det_config_cache is not None:
        return _det_config_cache

    import openpyxl
    if not OPTIMIZATION_FILE.exists():
        log.warning(f"File ottimizzazione non trovato: {OPTIMIZATION_FILE}")
        _det_config_cache = {}
        return _det_config_cache

    wb = openpyxl.load_workbook(OPTIMIZATION_FILE, read_only=True, data_only=True)
    if "Riepilogo" not in wb.sheetnames:
        _det_config_cache = {}
        wb.close()
        return _det_config_cache

    ws = wb["Riepilogo"]
    config = {}
    for r in range(2, ws.max_row + 1):
        city = ws.cell(r, 1).value
        if not city:
            continue
        orizzonte = ws.cell(r, 4).value
        if orizzonte != "1GG":
            continue
        season = ws.cell(r, 3).value
        method = ws.cell(r, 5).value
        verde = ws.cell(r, 7).value or 0
        mae = ws.cell(r, 8).value or 1.0
        models_str = ws.cell(r, 13).value or ""
        models = [m.strip() for m in models_str.split(",")
                  if m.strip() and m.strip() != "cma_grapes_global"]
        config[(str(city).strip(), str(season).strip())] = {
            "method": method, "models": models, "corrections": {},
            "mae": float(mae), "verde": float(verde),
        }
    wb.close()
    _det_config_cache = config
    log.info(f"Config deterministici caricata: {len(config)} combinazioni")
    return config


_model_stats_cache = None

def load_model_stats() -> dict:
    global _model_stats_cache
    if _model_stats_cache is not None:
        return _model_stats_cache
    if not MODEL_STATS_FILE.exists():
        log.warning(f"Model stats non trovato: {MODEL_STATS_FILE}")
        _model_stats_cache = {}
        return _model_stats_cache
    import pickle
    with open(MODEL_STATS_FILE, "rb") as f:
        _model_stats_cache = pickle.load(f)
    log.info(f"Model stats caricati: {len(_model_stats_cache.get('stats', {}))} combinazioni")
    return _model_stats_cache


_top_models_cache = None

def load_top_models() -> dict:
    global _top_models_cache
    if _top_models_cache is not None:
        return _top_models_cache
    import openpyxl
    if not TOP_MODELS_FILE_V2.exists():
        _top_models_cache = {}
        return _top_models_cache

    wb = openpyxl.load_workbook(TOP_MODELS_FILE_V2, read_only=True, data_only=True)
    result = {}
    sheet_map = {}
    for sn in wb.sheetnames:
        if "Top 5" in sn and "1GG" in sn:
            sheet_map["top5"] = sn
        elif "Top 10" in sn and "1GG" in sn:
            sheet_map["top10"] = sn

    for key, sheet_name in sheet_map.items():
        ws = wb[sheet_name]
        for r in range(2, ws.max_row + 1):
            city = ws.cell(r, 1).value
            if not city:
                continue
            city = str(city).strip()
            if city not in result:
                result[city] = {"top5": [], "top10": []}
            models = []
            col = 2
            while True:
                model = ws.cell(r, col).value
                if not model:
                    break
                models.append(str(model).strip())
                col += 3
            result[city][key] = models
    wb.close()
    _top_models_cache = result
    log.info(f"Top modelli caricati: {len(result)} citta")
    return result


# ═══════════════════════════════════════════════════════════════════════════════
# POLYMARKET - FETCH MERCATI
# ═══════════════════════════════════════════════════════════════════════════════

def fetch_temp_events(closed: bool = False) -> list[dict]:
    all_events = []
    offset = 0
    try:
        while True:
            resp = requests.get(
                f"{GAMMA_API}/events",
                params={"tag_slug": "weather", "closed": str(closed).lower(),
                        "limit": 200, "offset": offset},
                timeout=20,
            )
            resp.raise_for_status()
            batch = resp.json()
            temp_events = [e for e in batch if "highest temperature" in e.get("title", "").lower()]
            all_events.extend(temp_events)
            if len(batch) < 200:
                break
            offset += 200
        return all_events
    except requests.RequestException as e:
        log.warning(f"Gamma API error: {e}")
        return all_events


def city_from_title(title: str) -> str:
    m = re.search(r"Highest temperature in (.+?) on ", title, re.IGNORECASE)
    return m.group(1).strip() if m else ""


def date_from_title(title: str, end_date_hint: str | None = None) -> str | None:
    m = re.search(r"on (\w+ \d+)", title, re.IGNORECASE)
    if not m:
        return None
    if end_date_hint:
        try:
            hint_year = int(end_date_hint[:4])
            dt = datetime.strptime(f"{m.group(1)} {hint_year}", "%B %d %Y")
            return dt.strftime("%Y-%m-%d")
        except (ValueError, IndexError):
            pass
    now = datetime.now()
    for year in [now.year, now.year + 1, now.year - 1]:
        try:
            dt = datetime.strptime(f"{m.group(1)} {year}", "%B %d %Y")
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def parse_bucket(label: str) -> dict | None:
    lbl = label.lower()
    unit = "F" if re.search(r'[°ºo]F|°f|\d\s*F\b', label) else "C"
    is_lower = "or below" in lbl or "or lower" in lbl or "less than" in lbl
    is_upper = "or higher" in lbl or "or above" in lbl
    range_match = re.search(r'(\d+)\s*[-–]\s*(\d+)', label)
    nums = re.findall(r"-?\d+", label)
    if not nums:
        return None
    if is_lower:
        return {"unit": unit, "low": None, "high": int(nums[0]),
                "is_lower": True, "is_upper": False, "label": label}
    elif is_upper:
        return {"unit": unit, "low": int(nums[0]), "high": None,
                "is_lower": False, "is_upper": True, "label": label}
    elif range_match:
        return {"unit": unit, "low": int(range_match.group(1)),
                "high": int(range_match.group(2)),
                "is_lower": False, "is_upper": False, "label": label}
    else:
        val = int(nums[0])
        return {"unit": unit, "low": val, "high": val,
                "is_lower": False, "is_upper": False, "label": label}


def temp_sort_key(label: str) -> float:
    nums = re.findall(r"-?\d+\.?\d*", label)
    if not nums:
        return 0.0
    val = float(nums[0])
    lbl = label.lower()
    if "or higher" in lbl or "or above" in lbl:
        return val + 0.5
    if "or below" in lbl or "or lower" in lbl:
        return val - 0.5
    return val


def parse_markets_from_events(events: list[dict], days_ahead: int) -> list[dict]:
    now = datetime.now(timezone.utc)
    # Solo mercati per domani (target_date = oggi + 1)
    tomorrow = (now + timedelta(days=1)).strftime("%Y-%m-%d")
    markets = {}

    for event in events:
        title = event.get("title", "")
        city = city_from_title(title)
        event_end = event.get("endDate", "")
        target_date = date_from_title(title, end_date_hint=event_end)
        if not city or not target_date:
            continue

        # Filtra: solo mercati per domani
        if target_date != tomorrow:
            continue

        for m in event.get("markets") or []:
            end_str = m.get("endDate") or event_end
            end = None
            if end_str:
                try:
                    end = datetime.fromisoformat(end_str.replace("Z", "+00:00"))
                except ValueError:
                    pass
            if not end or end < now:
                continue

            label = m.get("groupItemTitle") or m.get("question", "")
            # Usa bestAsk se disponibile e > 0, altrimenti lastTradePrice
            best_ask = m.get("bestAsk")
            last_trade = m.get("lastTradePrice")
            try:
                best_ask_f = float(best_ask) if best_ask is not None else 0.0
            except (ValueError, TypeError):
                best_ask_f = 0.0
            try:
                last_trade_f = float(last_trade) if last_trade is not None else 0.0
            except (ValueError, TypeError):
                last_trade_f = 0.0
            prob = best_ask_f if best_ask_f > 0 else last_trade_f

            parsed = parse_bucket(label)
            key = f"{city}_{target_date}"

            if key not in markets:
                markets[key] = {
                    "city": city, "target_date": target_date, "end_dt": end,
                    "event_slug": event.get("slug", ""),
                    "buckets": [],
                }

            markets[key]["buckets"].append({
                "label": label, "prob": prob, "parsed": parsed,
                "token_id": m.get("clobTokenIds"),
                "condition_id": m.get("conditionId"),
                "market_slug": m.get("slug", ""),
            })

    result = list(markets.values())
    for mkt in result:
        mkt["buckets"].sort(key=lambda b: temp_sort_key(b["label"]))
    return result


# ═══════════════════════════════════════════════════════════════════════════════
# METEO - FETCH & CALCOLO PROBABILITA
# ═══════════════════════════════════════════════════════════════════════════════

def calc_bucket_prob_from_celsius(values_c: np.ndarray, bucket: dict) -> float:
    if len(values_c) == 0:
        return 0.0
    if bucket["unit"] == "F":
        values = values_c * 9.0 / 5.0 + 32.0
    else:
        values = values_c
    rounded = np.round(values).astype(int)
    total = len(rounded)
    if bucket["is_lower"]:
        return int(np.sum(rounded <= bucket["high"])) / total
    elif bucket["is_upper"]:
        return int(np.sum(rounded >= bucket["low"])) / total
    else:
        mask = (rounded >= bucket["low"]) & (rounded <= bucket["high"])
        return int(np.sum(mask)) / total


def fetch_deterministic_for_city(lat, lon, forecast_date, models) -> dict:
    if not models:
        return {}
    results = {}
    batch_size = 10
    for i in range(0, len(models), batch_size):
        batch = models[i:i + batch_size]
        try:
            r = requests.get(DETERMINISTIC_API, params={
                "latitude": lat, "longitude": lon,
                "daily": "temperature_2m_max",
                "models": ",".join(batch),
                "start_date": forecast_date, "end_date": forecast_date,
                "timezone": "auto",
            }, timeout=30)
            if r.status_code != 200:
                continue
            data = r.json()
            daily = data.get("daily", {})
            times = daily.get("time", [])
            if forecast_date not in times:
                continue
            idx = times.index(forecast_date)
            if len(batch) == 1:
                key = "temperature_2m_max"
                if key in daily and daily[key][idx] is not None:
                    results[batch[0]] = daily[key][idx]
            else:
                for model in batch:
                    key = f"temperature_2m_max_{model}"
                    if key in daily and daily[key][idx] is not None:
                        results[model] = daily[key][idx]
            time_mod.sleep(0.3)
        except Exception:
            continue
    return results


BIAS_THRESHOLD = 1.5  # Se median |bias| citta/stagione > soglia, no bias correction

def consensus_gaussian_probs(raw_forecasts: dict, city_opt: str, season: str,
                              parsed_buckets: list[dict]) -> dict | None:
    """
    Gaussiana del consenso con protezione bias (Fix 4).

    1. Controlla median |bias| della citta/stagione: se > BIAS_THRESHOLD,
       usa i forecast grezzi senza correzione bias.
    2. Calcola media pesata di tutti i modelli (mu del consenso).
    3. Calcola SE = sqrt(spread² + (median_sigma/sqrt(n))²).
    4. Usa una singola gaussiana N(mu, SE²) per calcolare le probabilita.
    """
    model_stats_data = load_model_stats()
    if not model_stats_data or "stats" not in model_stats_data:
        return None

    all_stats = model_stats_data["stats"]

    if not parsed_buckets:
        return None

    unit = parsed_buckets[0]["unit"] if parsed_buckets[0] else "C"

    # Controlla median |bias| per decidere se applicare la correzione
    biases_for_city = []
    for key, s in all_stats.items():
        if key[0] == city_opt and key[2] == season and key[3] == "D1":
            biases_for_city.append(abs(s["bias"]))
    median_bias = float(np.median(biases_for_city)) if biases_for_city else 0
    use_bias = median_bias <= BIAS_THRESHOLD

    models_used = []
    for model_name, forecast_c in raw_forecasts.items():
        if forecast_c is None:
            continue
        key = (city_opt, model_name, season, "D1")
        if key not in all_stats:
            continue
        s = all_stats[key]
        if use_bias:
            mu_c = forecast_c - s["bias"]
        else:
            mu_c = forecast_c
        models_used.append({
            "model": model_name, "forecast_c": forecast_c,
            "bias": s["bias"], "sigma": s["sigma"],
            "weight": s["weight"], "mu_c": mu_c,
        })

    if len(models_used) < 3:
        return None

    total_w = sum(m["weight"] for m in models_used)
    for m in models_used:
        m["weight_norm"] = m["weight"] / total_w

    corrected_means = [m["mu_c"] for m in models_used]
    inter_model_spread = float(np.std(corrected_means))
    weighted_mean_c = sum(m["mu_c"] * m["weight_norm"] for m in models_used)
    median_sigma = float(np.median([m["sigma"] for m in models_used]))
    n = len(models_used)

    # Standard error del consenso
    se_c = np.sqrt(inter_model_spread**2 + (median_sigma / np.sqrt(n))**2)

    if unit == "F":
        mu = weighted_mean_c * 9 / 5 + 32
        se = se_c * 9 / 5
    else:
        mu = weighted_mean_c
        se = se_c
    se = max(se, 0.3)

    probs = {}
    for b in parsed_buckets:
        if b is None:
            continue
        if b["is_lower"]:
            p = norm.cdf((b["high"] + 0.5 - mu) / se)
        elif b["is_upper"]:
            p = 1 - norm.cdf((b["low"] - 0.5 - mu) / se)
        else:
            p = (norm.cdf((b["high"] + 0.5 - mu) / se) -
                 norm.cdf((b["low"] - 0.5 - mu) / se))
        probs[b["label"]] = max(0.0, p)

    total = sum(probs.values())
    if total > 0:
        probs = {k: v / total for k, v in probs.items()}

    bias_note = "con bias" if use_bias else f"NO BIAS (median|b|={median_bias:.1f})"

    return {
        "probs": probs,
        "n_models": len(models_used),
        "spread": inter_model_spread,
        "mean_c": weighted_mean_c,
        "se_c": se_c,
        "use_bias": use_bias,
        "bias_note": bias_note,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# CALCOLO PROBABILITA COMPLETO PER UN MERCATO
# ═══════════════════════════════════════════════════════════════════════════════

def compute_market_probs(city: str, target_date: str,
                          parsed_buckets: list[dict]) -> dict | None:
    """Calcola le probabilita per un mercato usando gaussiana del consenso (Fix 4)."""
    opt_city = CITY_NAME_TO_OPT.get(city)
    if not opt_city:
        return None

    city_data = match_city(city)
    if not city_data:
        return None

    month = int(target_date.split("-")[1])
    season = get_season(month, opt_city)

    log.info(f"  {city} {target_date}: fetching {len(ALL_DETERMINISTIC_MODELS)} modelli...")
    raw = fetch_deterministic_for_city(
        city_data["lat"], city_data["lon"], target_date, ALL_DETERMINISTIC_MODELS)
    log.info(f"  {city} {target_date}: {len(raw)}/{len(ALL_DETERMINISTIC_MODELS)} ricevuti")

    if len(raw) < 3:
        log.warning(f"  {city} {target_date}: troppi pochi modelli, skip")
        return None

    result = consensus_gaussian_probs(raw, opt_city, season, parsed_buckets)
    if result:
        log.info(f"  {city} {target_date}: consensus SE OK - {result['n_models']} modelli, "
                 f"spread={result['spread']:.2f}°C, SE={result['se_c']:.2f}°C, "
                 f"media={result['mean_c']:.1f}°C ({result['bias_note']})")
    return result


# ═══════════════════════════════════════════════════════════════════════════════
# POSITION SIZING (Kelly frazionario)
# ═══════════════════════════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════════════════════════
# POLYMARKET CLOB - PIAZZAMENTO ORDINI
# ═══════════════════════════════════════════════════════════════════════════════

def init_clob_client():
    """Inizializza il client CLOB di Polymarket."""
    try:
        from py_clob_client.client import ClobClient
        from py_clob_client.clob_types import ApiCreds

        api_key = os.environ.get("POLY_API_KEY", "")
        api_secret = os.environ.get("POLY_API_SECRET", "")
        passphrase = os.environ.get("POLY_PASSPHRASE", "")
        private_key = os.environ.get("POLY_PRIVATE_KEY", "")

        if not all([api_key, api_secret, passphrase, private_key]):
            log.warning("Credenziali Polymarket mancanti - ordini disabilitati")
            return None

        creds = ApiCreds(
            api_key=api_key,
            api_secret=api_secret,
            api_passphrase=passphrase,
        )

        client = ClobClient(
            CLOB_API,
            key=private_key,
            chain_id=137,  # Polygon mainnet
            creds=creds,
        )

        log.info("Client CLOB Polymarket inizializzato")
        return client
    except Exception as e:
        log.error(f"Errore init CLOB client: {e}")
        return None


def place_order(client, token_id: str, side: str, size_usd: float,
                market_price: float, city: str = "",
                dry_run: bool = True) -> dict | None:
    """
    Piazza un LIMIT ORDER su Polymarket.
    Prezzo = prezzo di mercato attuale + tolleranza (0.5%).
    Scadenza = 6h (America/Europa) o 2.5h (Asia).
    """
    # Prezzo limite: +1% se mercato > 5%, +0.5% se mercato <= 5%
    tol = PRICE_TOLERANCE_HIGH if market_price >= PRICE_TOLERANCE_THRESHOLD else PRICE_TOLERANCE_LOW
    limit_price = round(min(market_price + tol, 0.99), 2)
    # Polymarket accetta prezzi con 2 decimali (0.01 - 0.99)
    limit_price = max(limit_price, 0.01)

    # Scadenza diversa per mercati asiatici vs americani/europei
    city_data = match_city(city) if city else None
    is_asia = city_data and city_data.get("tz", "").startswith("Asia/")
    exp_secs = ORDER_EXPIRATION_SECS_ASIA if is_asia else ORDER_EXPIRATION_SECS
    exp_label = "2.5h" if is_asia else "6h"
    expiration = int((datetime.now(timezone.utc) + timedelta(seconds=exp_secs)).timestamp())

    if dry_run:
        log.info(f"    [DRY-RUN] LIMIT {side} ${size_usd:.2f} @ {limit_price:.2f} "
                 f"(mkt={market_price:.2f}, exp={exp_label}, token: {token_id[:16]}...)")
        return {"dry_run": True, "side": side, "size": size_usd,
                "limit_price": limit_price, "market_price": market_price}

    if client is None:
        log.warning("    Client CLOB non disponibile, ordine saltato")
        return None

    try:
        from py_clob_client.clob_types import OrderArgs
        from py_clob_client.order_builder.constants import BUY

        order_args = OrderArgs(
            price=limit_price,
            size=size_usd,
            side=BUY,
            token_id=token_id,
            expiration=expiration,
        )

        signed_order = client.create_order(order_args)
        result = client.post_order(signed_order)

        log.info(f"    ORDINE PIAZZATO: LIMIT BUY ${size_usd:.2f} @ {limit_price:.2f} "
                 f"(mkt={market_price:.2f}, scade tra {exp_label})")
        return result

    except Exception as e:
        log.error(f"    Errore piazzamento ordine: {e}")
        return None


# ═══════════════════════════════════════════════════════════════════════════════
# TRADING LOGIC
# ═══════════════════════════════════════════════════════════════════════════════

def find_trades(market: dict, model_result: dict, min_edge: float) -> list[dict]:
    """
    Strategia COMBtop2: trova opportunita di trading in un mercato.

    Filtri:
    - p <= 10%: solo bucket con prezzo Polymarket <= MAX_MARKET_PRICE
    - mxP < 55%: prob massima del modello su un singolo bucket < MAX_POLY_PEAK
    - dist >= 1: bucket ad almeno MIN_DIST_FROM_PEAK posizioni dal picco modello
    - spr >= 0.6: spread inter-modello >= MIN_SPREAD
    - Top 2: prendi i TOP_N bucket con edge positivo piu alto
    """
    model_probs = model_result["probs"]
    spread = model_result["spread"]

    # ── Filtro spread inter-modello ──────────────────────────────────────
    if spread < MIN_SPREAD:
        log.info(f"  SKIP: spread={spread:.2f}°C < {MIN_SPREAD}°C (modelli troppo d'accordo)")
        return []

    # ── Filtro bestAsk massima Polymarket ────────────────────────────────
    max_poly_prob = max(b["prob"] for b in market["buckets"]) if market["buckets"] else 0
    if max_poly_prob >= MAX_POLY_PEAK:
        log.info(f"  SKIP: max bestAsk Polymarket={max_poly_prob:.0%} >= {MAX_POLY_PEAK:.0%} (mercato troppo concentrato)")
        return []

    # ── Filtro same peak: se modello e Polymarket concordano sul favorito → skip ──
    if SKIP_IF_SAME_PEAK:
        model_peak = max(model_probs, key=model_probs.get)
        poly_peak = max(market["buckets"], key=lambda b: b["prob"])["label"]
        if model_peak == poly_peak:
            log.info(f"  SKIP: picco modello = picco Polymarket ({model_peak}), mercato ben prezzato")
            return []

    # ── Valuta ogni bucket ───────────────────────────────────────────────
    candidates = []
    for i, bucket in enumerate(market["buckets"]):
        label = bucket["label"]
        pm_prob = bucket["prob"]
        my_prob = model_probs.get(label, 0.0)
        edge_pp = (my_prob - pm_prob) * 100

        # Solo edge positivo (modello vede piu valore del mercato)
        if edge_pp <= 0:
            continue

        # Filtro prezzo Polymarket: p <= 10%
        if pm_prob > MAX_MARKET_PRICE:
            continue

        # Filtro prob minima modello: almeno 1.5%
        if my_prob < MIN_MODEL_PROB:
            continue

        candidates.append({
            "label": label,
            "side": "BUY",
            "target": "YES",
            "my_prob": my_prob,
            "market_prob": pm_prob,
            "edge_pp": edge_pp,
            "token_id": bucket.get("token_id"),
            "condition_id": bucket.get("condition_id"),
        })

    # ── Top N per edge ───────────────────────────────────────────────────
    candidates.sort(key=lambda t: t["edge_pp"], reverse=True)
    trades = candidates[:TOP_N]

    if candidates and not trades:
        log.info(f"  Nessun candidato passa tutti i filtri")
    elif len(candidates) > TOP_N:
        log.info(f"  {len(candidates)} candidati, selezionati top {TOP_N}")

    return trades


def execute_trades(client, trades: list[dict], state: dict,
                   city: str = "", dry_run: bool = True) -> int:
    """Esegue i trade COMBtop2 (sempre BUY YES, puntata fissa)."""
    executed = 0

    for trade in trades:
        token_ids = trade.get("token_id")
        if not token_ids:
            continue

        try:
            if isinstance(token_ids, str):
                token_ids = json.loads(token_ids)
            token_id = token_ids[0]  # YES token
        except (json.JSONDecodeError, IndexError, TypeError):
            log.warning(f"  Token ID non valido per {trade['label']}")
            continue

        price = trade["market_prob"]

        tol = PRICE_TOLERANCE if price >= PRICE_TOLERANCE_THRESHOLD else 0
        limit_price = round(min(price + tol, 0.99), 2)
        city_data = match_city(city) if city else None
        is_asia = city_data and city_data.get("tz", "").startswith("Asia/")
        exp_label = "2.5h" if is_asia else "6h"
        log.info(f"  >>> BUY YES {trade['label']}: "
                 f"modello={trade['my_prob']:.1%} mercato={price:.1%} "
                 f"edge={trade['edge_pp']:+.1f}pp | ${BET_SIZE_USD:.2f} @ {limit_price:.2f} (exp {exp_label})")

        result = place_order(client, token_id, "BUY", BET_SIZE_USD, price, city, dry_run)

        if result:
            executed += 1
            state.setdefault("orders", []).append({
                "timestamp": datetime.now(timezone.utc).isoformat(),
                "label": trade["label"],
                "my_prob": round(trade["my_prob"], 4),
                "market_prob": round(price, 4),
                "edge_pp": round(trade["edge_pp"], 1),
                "bet_size": BET_SIZE_USD,
                "dry_run": dry_run,
            })

    return executed


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN LOOP
# ═══════════════════════════════════════════════════════════════════════════════

def run_cycle(state: dict, client, min_edge: float,
              days_ahead: int, dry_run: bool) -> int:
    """Esegue un ciclo completo di trading."""
    actions = 0

    log.info("Scarico mercati Polymarket...")
    events = fetch_temp_events(closed=False)
    if not events:
        log.info("  Nessun evento trovato")
        return 0

    markets = parse_markets_from_events(events, days_ahead)
    log.info(f"  {len(events)} eventi, {len(markets)} mercati attivi")

    for mkt in markets:
        city = mkt["city"]
        target_date = mkt["target_date"]

        if not match_city(city):
            continue

        parsed_buckets = [b.get("parsed") for b in mkt["buckets"] if b.get("parsed")]
        if not parsed_buckets:
            continue

        log.info(f"Analisi: {city} - {target_date} ({len(mkt['buckets'])} bucket)")

        # Calcola probabilita
        result = compute_market_probs(city, target_date, parsed_buckets)
        if not result:
            continue

        # Trova trade (strategia COMBtop2)
        trades = find_trades(mkt, result, min_edge)
        if not trades:
            log.info(f"  Nessun edge > {min_edge}pp")
            continue

        log.info(f"  {len(trades)} opportunita trovate!")

        # Esegui trade
        n = execute_trades(client, trades, state, city, dry_run)
        actions += n

    save_state(state)
    return actions


def seconds_until_next_trade() -> float:
    """Calcola i secondi fino alle prossime TRADE_HOUR_UTC:TRADE_MINUTE_UTC."""
    now = datetime.now(timezone.utc)
    target = now.replace(hour=TRADE_HOUR_UTC, minute=TRADE_MINUTE_UTC, second=0, microsecond=0)
    if target <= now:
        target += timedelta(days=1)
    return (target - now).total_seconds()


def main():
    parser = argparse.ArgumentParser(description="Polymarket Weather Trading Bot - COMBtop2")
    parser.add_argument("--dry-run", action="store_true", default=True,
                        help="Simula ordini senza piazzarli (default: True)")
    parser.add_argument("--live", action="store_true",
                        help="Piazza ordini reali (disattiva dry-run)")
    parser.add_argument("--once", action="store_true",
                        help="Esegui un solo ciclo e esci")
    parser.add_argument("--min-edge", type=float, default=MIN_EDGE_PP,
                        help=f"Edge minimo in pp (default: {MIN_EDGE_PP})")
    parser.add_argument("--bet", type=float, default=BET_SIZE_USD,
                        help=f"Puntata fissa $ per mercato (default: {BET_SIZE_USD})")
    parser.add_argument("--days", type=int, default=DAYS_AHEAD,
                        help=f"Giorni avanti (default: {DAYS_AHEAD})")

    args = parser.parse_args()
    dry_run = not args.live

    # Aggiorna config da args
    import bot_trading as _self
    _self.BET_SIZE_USD = args.bet

    mode = "DRY-RUN" if dry_run else "LIVE"
    log.info(f"{'='*60}")
    log.info(f"  POLYMARKET WEATHER TRADING BOT - {mode}")
    log.info(f"  Strategia: Gaussiana SE (Fix4) | Min edge: {args.min_edge}pp")
    log.info(f"  Puntata fissa: ${args.bet} per mercato")
    log.info(f"  Orario trading: {TRADE_HOUR_UTC}:{TRADE_MINUTE_UTC:02d} UTC (ogni giorno)")
    log.info(f"  Bias threshold: {BIAS_THRESHOLD}C | Filtri: p<={MAX_MARKET_PRICE:.0%} mxP<{MAX_POLY_PEAK:.0%} minProb>={MIN_MODEL_PROB:.1%} spr>={MIN_SPREAD}")
    log.info(f"{'='*60}")

    if not dry_run:
        log.warning(">>> MODALITA LIVE - ORDINI REALI <<<")

    # Init CLOB client
    client = None
    if not dry_run:
        client = init_clob_client()
        if client is None:
            log.error("Impossibile inizializzare CLOB client, esco")
            sys.exit(1)

    state = load_state()

    if args.once:
        run_cycle(state, client, args.min_edge, args.days, dry_run)
        return

    # Loop giornaliero: aspetta le 17:10 UTC, esegui, ripeti
    while True:
        wait = seconds_until_next_trade()
        next_time = datetime.now(timezone.utc) + timedelta(seconds=wait)
        log.info(f"Prossimo trading: {next_time.strftime('%Y-%m-%d %H:%M UTC')} (tra {wait/3600:.1f}h)")

        time_mod.sleep(wait)

        log.info(f"{'='*40} TRADING CYCLE {'='*40}")
        try:
            run_cycle(state, client, args.min_edge, args.days, dry_run)
        except Exception as e:
            log.error(f"Errore nel ciclo: {e}")
            log.debug(traceback.format_exc())

        # Pausa 60s per evitare doppio trigger
        time_mod.sleep(60)


if __name__ == "__main__":
    main()
