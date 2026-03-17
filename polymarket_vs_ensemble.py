"""
Confronta probabilita' ensemble multi-modello con prezzi Polymarket
per il mercato "Highest temperature in London".

Risoluzione Polymarket: temperatura massima arrotondata al grado intero
misurata alla London City Airport Station (Weather Underground).
Quindi "11C" = temperatura reale in [10.5, 11.5).
"""

import requests
import numpy as np
import pandas as pd
import os
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUT_PATH = os.path.join(BASE_DIR, "Polymarket vs Ensemble Londra.xlsx")

LAT = 51.505278
LON = 0.055278
API_URL = "https://ensemble-api.open-meteo.com/v1/ensemble"
START = "2026-03-14"
END = "2026-03-17"

MODELS = [
    "icon_seamless_eps",
    "icon_global_eps",
    "icon_eu_eps",
    "icon_d2_eps",
    "ncep_gefs_seamless",
    "ncep_gefs025",
    "ncep_gefs05",
    "ncep_aigefs025",
    "ecmwf_ifs025_ensemble",
    "ecmwf_aifs025_ensemble",
    "gem_global_ensemble",
    "bom_access_global_ensemble",
    "ukmo_global_ensemble_20km",
    "ukmo_uk_ensemble_2km",
]

# Prezzi Polymarket attuali (probabilita' implicite)
# Bucket = temperatura arrotondata al grado intero
POLYMARKET = {
    "2026-03-14": {
        10: 0.001, 11: 0.75, 12: 0.25, 13: 0.01, 14: 0.001,
    },
    "2026-03-15": {
        6: 0.005, 7: 0.01, 8: 0.01, 9: 0.015, 10: 0.11,
        11: 0.48, 12: 0.32, 13: 0.06, 14: 0.03,
    },
    "2026-03-16": {
        5: 0.007, 6: 0.01, 7: 0.01, 8: 0.03, 9: 0.06,
        10: 0.14, 11: 0.33, 12: 0.28, 13: 0.11, 14: 0.03, 15: 0.04,
    },
    "2026-03-17": {
        8: 0.009, 9: 0.01, 10: 0.035, 11: 0.05, 12: 0.14,
        13: 0.28, 14: 0.32, 15: 0.14, 16: 0.07, 17: 0.02, 18: 0.015,
    },
}

# Formattazione
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
BLUE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def fetch_all_ensemble_members(date_str):
    """Scarica TUTTI i membri ensemble da tutti i modelli per una data."""
    import time

    all_members = {}  # model_name -> list of T max values

    for model in MODELS:
        try:
            r = requests.get(API_URL, params={
                "latitude": LAT, "longitude": LON,
                "daily": "temperature_2m_max",
                "models": model,
                "timezone": "auto",
                "start_date": date_str,
                "end_date": date_str,
            }, timeout=60)

            if r.status_code != 200:
                continue

            data = r.json()
            daily = data.get("daily", {})
            times = daily.get("time", [])

            if date_str not in times:
                continue

            idx = times.index(date_str)
            values = []

            # Control run
            if "temperature_2m_max" in daily:
                v = daily["temperature_2m_max"][idx]
                if v is not None:
                    values.append(v)

            # Members
            for key in daily:
                if key.startswith("temperature_2m_max_member"):
                    v = daily[key][idx]
                    if v is not None:
                        values.append(v)

            if values:
                all_members[model] = values

            time.sleep(0.5)

        except Exception as e:
            print(f"  {model}: {e}")
            continue

    return all_members


def calc_bucket_probabilities(values, buckets):
    """
    Calcola probabilita' per ogni bucket di temperatura.
    Bucket "N" = temperatura reale in [N-0.5, N+0.5)
    cioe' arrotondata al grado intero = N.
    """
    rounded = np.round(values).astype(int)
    total = len(rounded)
    probs = {}

    for b in buckets:
        count = np.sum(rounded == b)
        probs[b] = count / total

    # Bucket estremi (es. "6 or below", "14 or higher")
    min_bucket = min(buckets)
    max_bucket = max(buckets)

    # Somma tutto cio' che cade sotto il min_bucket
    below = np.sum(rounded < min_bucket)
    probs[min_bucket] = (probs.get(min_bucket, 0) * total + below) / total

    # Somma tutto cio' che cade sopra il max_bucket
    above = np.sum(rounded > max_bucket)
    probs[max_bucket] = (probs.get(max_bucket, 0) * total + above) / total

    return probs


def main():
    import time as time_mod

    print("=== Polymarket vs Ensemble - Londra ===\n")

    results = {}  # date -> {model_probs, combined_probs, polymarket_probs}

    for date_str in sorted(POLYMARKET.keys()):
        print(f"\n--- {date_str} ---")
        buckets = sorted(POLYMARKET[date_str].keys())

        members = fetch_all_ensemble_members(date_str)

        if not members:
            print("  Nessun dato!")
            continue

        # Probabilita' per modello
        model_probs = {}
        for model, vals in members.items():
            short_name = model.replace("_ensemble", "").replace("_seamless", "")
            probs = calc_bucket_probabilities(np.array(vals), buckets)
            model_probs[short_name] = {"probs": probs, "n": len(vals)}
            print(f"  {short_name}: {len(vals)} membri")

        # Combinazione multi-modello (tutti i membri insieme, peso uguale)
        all_vals = []
        for vals in members.values():
            all_vals.extend(vals)
        combined_equal = calc_bucket_probabilities(np.array(all_vals), buckets)

        # Combinazione pesata per qualita' modello (pesi soggettivi basati su skill noto)
        # ECMWF > ICON > GFS > GEM > BOM/UKMO
        model_weights = {
            "icon_eps": 1.0, "icon_global_eps": 1.0, "icon_eu_eps": 1.2,
            "icon_d2_eps": 1.3,
            "ncep_gefs": 0.8, "ncep_gefs025": 0.9, "ncep_gefs05": 0.7,
            "ncep_aigefs025": 0.9,
            "ecmwf_ifs025": 1.5, "ecmwf_aifs025": 1.4,
            "gem_global": 0.7,
            "bom_access_global": 0.5, "ukmo_global_20km": 0.6,
            "ukmo_uk_2km": 0.8,
        }

        weighted_probs = {b: 0.0 for b in buckets}
        total_weight = 0.0
        for model_short, data in model_probs.items():
            w = model_weights.get(model_short, 0.7)
            for b in buckets:
                weighted_probs[b] += data["probs"].get(b, 0) * w
            total_weight += w

        for b in buckets:
            weighted_probs[b] /= total_weight

        results[date_str] = {
            "buckets": buckets,
            "model_probs": model_probs,
            "combined_equal": combined_equal,
            "combined_weighted": weighted_probs,
            "polymarket": POLYMARKET[date_str],
            "n_total": len(all_vals),
        }

    # ── Scrivi Excel ──
    print("\n\nScrittura Excel...")

    with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:

        for date_str, res in results.items():
            buckets = res["buckets"]
            sheet_name = date_str.replace("2026-", "").replace("-", " Mar ")

            rows = []

            # Header info
            rows.append({"Fonte": f"DATA: {date_str}", **{f"{b}C": "" for b in buckets}})
            rows.append({"Fonte": f"Membri totali: {res['n_total']}", **{f"{b}C": "" for b in buckets}})
            rows.append({"Fonte": "", **{f"{b}C": "" for b in buckets}})

            # Polymarket
            pm_row = {"Fonte": "POLYMARKET"}
            for b in buckets:
                pm_row[f"{b}C"] = f"{res['polymarket'].get(b, 0)*100:.1f}%"
            rows.append(pm_row)

            rows.append({"Fonte": "", **{f"{b}C": "" for b in buckets}})

            # Ensemble combinato (uguale)
            eq_row = {"Fonte": "ENSEMBLE (tutti, peso uguale)"}
            for b in buckets:
                eq_row[f"{b}C"] = f"{res['combined_equal'].get(b, 0)*100:.1f}%"
            rows.append(eq_row)

            # Ensemble combinato (pesato)
            wt_row = {"Fonte": "ENSEMBLE (pesato per skill)"}
            for b in buckets:
                wt_row[f"{b}C"] = f"{res['combined_weighted'].get(b, 0)*100:.1f}%"
            rows.append(wt_row)

            rows.append({"Fonte": "", **{f"{b}C": "" for b in buckets}})

            # Delta (edge) = ensemble - polymarket
            delta_row = {"Fonte": "EDGE (ensemble - market)"}
            for b in buckets:
                ens_p = res["combined_weighted"].get(b, 0)
                mkt_p = res["polymarket"].get(b, 0)
                delta = (ens_p - mkt_p) * 100
                delta_row[f"{b}C"] = f"{delta:+.1f}pp"
            rows.append(delta_row)

            # Segnale
            signal_row = {"Fonte": "SEGNALE"}
            for b in buckets:
                ens_p = res["combined_weighted"].get(b, 0)
                mkt_p = res["polymarket"].get(b, 0)
                delta = ens_p - mkt_p
                if delta > 0.05:
                    signal_row[f"{b}C"] = "BUY"
                elif delta < -0.05:
                    signal_row[f"{b}C"] = "SELL"
                else:
                    signal_row[f"{b}C"] = "-"
            rows.append(signal_row)

            rows.append({"Fonte": "", **{f"{b}C": "" for b in buckets}})
            rows.append({"Fonte": "--- PER MODELLO ---", **{f"{b}C": "" for b in buckets}})

            # Per modello
            for model_short, data in sorted(res["model_probs"].items()):
                m_row = {"Fonte": f"{model_short} ({data['n']}m)"}
                for b in buckets:
                    m_row[f"{b}C"] = f"{data['probs'].get(b, 0)*100:.1f}%"
                rows.append(m_row)

            df = pd.DataFrame(rows)
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Formattazione
            ws = writer.sheets[sheet_name]

            # Header
            for cell in ws[1]:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = CENTER
                cell.border = THIN_BORDER

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                first = row[0].value or ""

                for cell in row:
                    cell.border = THIN_BORDER
                    cell.alignment = CENTER
                    cell.font = Font(size=10)

                if first == "POLYMARKET":
                    for cell in row:
                        cell.fill = YELLOW_FILL
                        cell.font = Font(bold=True, size=10)

                elif first.startswith("ENSEMBLE"):
                    for cell in row:
                        cell.fill = BLUE_FILL
                        cell.font = Font(bold=True, size=10)

                elif first.startswith("EDGE"):
                    for cell in row:
                        cell.font = Font(bold=True, size=10)
                        val = str(cell.value or "")
                        if val.startswith("+") and not val.startswith("+0."):
                            cell.fill = GREEN_FILL
                        elif val.startswith("-") and not val.startswith("-0."):
                            cell.fill = RED_FILL

                elif first == "SEGNALE":
                    for cell in row:
                        cell.font = Font(bold=True, size=11)
                        val = str(cell.value or "")
                        if val == "BUY":
                            cell.fill = GREEN_FILL
                            cell.font = Font(bold=True, size=11, color="006100")
                        elif val == "SELL":
                            cell.fill = RED_FILL
                            cell.font = Font(bold=True, size=11, color="9C0006")

                elif first.startswith("--- "):
                    for cell in row:
                        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                        cell.font = Font(bold=True, size=9)

                elif first.startswith("DATA:") or first.startswith("Membri"):
                    row[0].font = Font(bold=True, size=10)

            # Larghezze colonne
            ws.column_dimensions["A"].width = 32
            for col_idx in range(2, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 10

            ws.freeze_panes = "B2"

        # ── Foglio riepilogo finale ──
        summary_rows = []
        for date_str, res in results.items():
            for b in res["buckets"]:
                ens_p = res["combined_weighted"].get(b, 0)
                mkt_p = res["polymarket"].get(b, 0)
                edge = ens_p - mkt_p
                summary_rows.append({
                    "Data": date_str,
                    "Bucket": f"{b}C",
                    "Polymarket": f"{mkt_p*100:.1f}%",
                    "Ensemble": f"{ens_p*100:.1f}%",
                    "Edge (pp)": f"{edge*100:+.1f}",
                    "Segnale": "BUY" if edge > 0.05 else ("SELL" if edge < -0.05 else "-"),
                })

        df_sum = pd.DataFrame(summary_rows)
        df_sum.to_excel(writer, sheet_name="Riepilogo Segnali", index=False)

        ws_sum = writer.sheets["Riepilogo Segnali"]
        for cell in ws_sum[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER

        for row in ws_sum.iter_rows(min_row=2, max_row=ws_sum.max_row, max_col=6):
            for cell in row:
                cell.border = THIN_BORDER
                cell.alignment = CENTER
                cell.font = Font(size=10)
            signal = row[5].value
            if signal == "BUY":
                for cell in row:
                    cell.fill = GREEN_FILL
            elif signal == "SELL":
                for cell in row:
                    cell.fill = RED_FILL

        ws_sum.column_dimensions["A"].width = 14
        ws_sum.column_dimensions["B"].width = 10
        ws_sum.column_dimensions["C"].width = 14
        ws_sum.column_dimensions["D"].width = 14
        ws_sum.column_dimensions["E"].width = 12
        ws_sum.column_dimensions["F"].width = 10

        # Sposta riepilogo come primo foglio
        wb = writer.book
        wb.move_sheet("Riepilogo Segnali", offset=-len(wb.sheetnames) + 1)

    print(f"\nFile salvato: {OUT_PATH}")

    # ── Stampa riepilogo a console ──
    print("\n" + "=" * 70)
    print("RIEPILOGO SEGNALI")
    print("=" * 70)

    for date_str, res in results.items():
        print(f"\n  {date_str}  ({res['n_total']} membri)")
        print(f"  {'Bucket':>6}  {'Polymarket':>11}  {'Ensemble':>11}  {'Edge':>8}  {'Segnale':>8}")
        print(f"  {'-'*6}  {'-'*11}  {'-'*11}  {'-'*8}  {'-'*8}")

        for b in res["buckets"]:
            ens_p = res["combined_weighted"].get(b, 0)
            mkt_p = res["polymarket"].get(b, 0)
            edge = (ens_p - mkt_p) * 100
            signal = "BUY" if edge > 5 else ("SELL" if edge < -5 else "-")
            marker = " <<<" if abs(edge) > 5 else ""
            print(f"  {b:>4}C  {mkt_p*100:>10.1f}%  {ens_p*100:>10.1f}%  {edge:>+7.1f}pp  {signal:>8}{marker}")


if __name__ == "__main__":
    main()
