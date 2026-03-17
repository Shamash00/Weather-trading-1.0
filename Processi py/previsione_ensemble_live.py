"""
Previsione Ensemble Live + Raccolta Storica

Scarica previsioni ensemble probabilistiche da Open-Meteo per tutte le citta,
calcola distribuzione di probabilita per i bucket di temperatura, e salva
lo storico giornaliero per analisi future.

Modelli ensemble (179 membri totali per citta):
- ECMWF IFS 0.25  (51 membri)
- GFS GEFS 0.25   (31 membri)
- ICON Seamless    (40 membri)
- GEM Global       (21 membri)
- BOM ACCESS       (18 membri)
- UKMO Global      (18 membri)
"""

import pandas as pd
import numpy as np
import requests
import os
import time
import json
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

CITIES = {
    "Ankara":      {"lat": 40.128082, "lon": 32.995083},
    "Atlanta":     {"lat": 33.62972,  "lon": -84.44224},
    "BuenosAires": {"lat": -34.822222,"lon": -58.535833},
    "Chicago":     {"lat": 41.96019,  "lon": -87.93162},
    "Dallas":      {"lat": 32.8519,   "lon": -96.8555},
    "Londra":      {"lat": 51.505278, "lon": 0.055278},
    "Lucknow":     {"lat": 26.7606,   "lon": 80.8893},
    "Miami":       {"lat": 25.78805,  "lon": -80.31694},
    "Monaco":      {"lat": 48.353783, "lon": 11.786086},
    "New York":    {"lat": 40.77945,  "lon": -73.88027},
    "Parigi":      {"lat": 49.012779, "lon": 2.55},
    "SaoPaulo":    {"lat": -23.432075,"lon": -46.469511},
    "Seattle":     {"lat": 47.4444,   "lon": -122.3138},
    "Seoul":       {"lat": 37.469075, "lon": 126.450517},
    "Shanghai":    {"lat": 31.143378, "lon": 121.805214},
    "Singapore":   {"lat": 1.350189,  "lon": 103.994433},
    "TelAviv":     {"lat": 32.011389, "lon": 34.886667},
    "Tokyo":       {"lat": 35.552258, "lon": 139.779694},
    "Toronto":     {"lat": 43.677223, "lon": -79.630556},
    "Wellington":  {"lat": -41.3333333,"lon": 174.8},
}

ENSEMBLE_URL = "https://ensemble-api.open-meteo.com/v1/ensemble"
ENSEMBLE_MODELS_PARAM = ("ecmwf_ifs025,gfs025,icon_seamless,gem_global,"
                         "bom_access_global_ensemble,ukmo_global_ensemble_20km")

# Suffissi colonne -> nome leggibile
MODEL_DISPLAY = {
    "ecmwf_ifs025_ensemble": "ECMWF IFS",
    "ncep_gefs025":          "GFS GEFS",
    "icon_seamless_eps":     "ICON",
    "gem_global_ensemble":   "GEM",
    "bom_access_global_ensemble": "BOM ACCESS",
    "ukmo_global_ensemble_20km":  "UKMO",
}

HISTORY_DIR = os.path.join(BASE_DIR, "_storico_ensemble")
os.makedirs(HISTORY_DIR, exist_ok=True)
out_path = os.path.join(BASE_DIR, "Previsione Ensemble Live.xlsx")

green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
gold_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
blue_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
light_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
bold = Font(bold=True)
bold_white = Font(bold=True, color="FFFFFF")


def fetch_ensemble_city(lat, lon):
    """Fetch all ensemble models for a city in a single API call."""
    r = requests.get(ENSEMBLE_URL, params={
        "latitude": lat, "longitude": lon,
        "hourly": "temperature_2m",
        "models": ENSEMBLE_MODELS_PARAM,
        "forecast_days": 3,
        "timezone": "auto",
        "temperature_unit": "fahrenheit",
    }, timeout=120)

    if r.status_code == 429:
        return None
    r.raise_for_status()
    return r.json()


def parse_members_by_model(data):
    """Parse API response: daily max per member, organized by model."""
    df = pd.DataFrame(data["hourly"])
    df["time"] = pd.to_datetime(df["time"])
    df["date"] = df["time"].dt.date
    dates = sorted(df["date"].unique())

    result = {}  # model_suffix -> {date -> array of member daily maxes}

    for suffix, display_name in MODEL_DISPLAY.items():
        member_cols = [c for c in df.columns if suffix in c and "temperature_2m" in c]
        if not member_cols:
            continue

        daily_max = df.groupby("date")[member_cols].max()
        model_dates = {}
        for d in dates:
            if d in daily_max.index:
                vals = daily_max.loc[d].dropna().values
                if len(vals) > 0:
                    model_dates[str(d)] = vals.tolist()

        result[suffix] = {
            "name": display_name,
            "n_members": len(member_cols),
            "dates": model_dates,
        }

    return result, [str(d) for d in dates]


def analyze_distribution(values_f):
    """Analyze temperature distribution from ensemble members (in F)."""
    vals_f = np.array(values_f)
    vals_c = (vals_f - 32) * 5 / 9
    vals_c_round = np.round(vals_c).astype(int)

    unique, counts = np.unique(vals_c_round, return_counts=True)
    probs = counts / len(vals_c_round) * 100
    dist = dict(zip(unique.tolist(), probs.tolist()))

    # Mode (most likely value)
    mode_idx = np.argmax(counts)
    mode_c = int(unique[mode_idx])
    mode_prob = float(probs[mode_idx])

    # P(+-1C) around mode = equivalent to verde%
    pm1_prob = sum(p for c, p in dist.items() if abs(c - mode_c) <= 1)

    # Best 2 adjacent integer C pair
    sorted_c = sorted(dist.keys())
    best_pair_prob = mode_prob
    best_pair = (mode_c, mode_c)
    for i in range(len(sorted_c) - 1):
        if sorted_c[i + 1] - sorted_c[i] == 1:
            pair_prob = dist[sorted_c[i]] + dist[sorted_c[i + 1]]
            if pair_prob > best_pair_prob:
                best_pair_prob = pair_prob
                best_pair = (sorted_c[i], sorted_c[i + 1])

    return {
        "mean_f": round(float(np.mean(vals_f)), 1),
        "std_f": round(float(np.std(vals_f)), 1),
        "mean_c": round(float(np.mean(vals_c)), 1),
        "std_c": round(float(np.std(vals_c)), 1),
        "mode_c": mode_c,
        "mode_f": round(mode_c * 9 / 5 + 32),
        "mode_prob": round(mode_prob, 1),
        "pm1_prob": round(float(pm1_prob), 1),
        "best_pair": list(best_pair),
        "best_pair_prob": round(float(best_pair_prob), 1),
        "distribution_c": {str(k): round(v, 1) for k, v in dist.items()},
        "n_members": len(vals_f),
    }


# ── Main ──
today = datetime.now().date()
print(f"Data esecuzione: {today}")
print(f"Modelli: 6 ensemble ({sum(m['n_members'] for m in [{'n_members': 51}, {'n_members': 31}, {'n_members': 40}, {'n_members': 21}, {'n_members': 18}, {'n_members': 18}])} membri)")
print(f"Citta: {len(CITIES)}")

# Check cache
history_file = os.path.join(HISTORY_DIR, f"{today.isoformat()}.json")
use_cache = False
cached_data = None
if os.path.exists(history_file):
    print(f"\nDati gia scaricati oggi, uso cache.")
    with open(history_file, encoding="utf-8") as f:
        cached_data = json.load(f)
    use_cache = True

all_results = {}
api_calls = 0

for city_name, coords in CITIES.items():
    lat, lon = coords["lat"], coords["lon"]

    print(f"\n{'=' * 70}")
    print(f"  {city_name}")
    print(f"{'=' * 70}")

    if use_cache and city_name in cached_data.get("cities", {}):
        city_data = cached_data["cities"][city_name]
        for h_key in ["1gg", "2gg"]:
            if h_key in city_data:
                d = city_data[h_key]
                print(f"  {h_key.upper()} (cache) data={d.get('target_date','')}  "
                      f"Moda={d['mode_c']}C  P(+-1C)={d['pm1_prob']}%")
        all_results[city_name] = city_data
        continue

    # Fetch
    try:
        data = fetch_ensemble_city(lat, lon)
        if data is None:
            print(f"  Rate limited! Attendo 60s...")
            time.sleep(60)
            data = fetch_ensemble_city(lat, lon)
        if data is None:
            print(f"  Ancora rate limited, skip.")
            continue

        api_calls += 1
        model_results, all_dates = parse_members_by_model(data)

        # Print per-model summary
        for suffix, mdata in model_results.items():
            if len(all_dates) >= 2 and all_dates[1] in mdata["dates"]:
                vals = mdata["dates"][all_dates[1]]
                mean_c = (np.mean(vals) - 32) * 5 / 9
                print(f"  {mdata['name']:<14} {mdata['n_members']:3d} membri  "
                      f"1GG media={np.mean(vals):.1f}F ({mean_c:.1f}C)")

    except Exception as e:
        print(f"  ERRORE: {e}")
        continue

    # Combine all members per horizon
    city_result = {}
    for h_idx, h_label in [(1, "1gg"), (2, "2gg")]:
        if h_idx >= len(all_dates):
            continue

        target_date = all_dates[h_idx]
        all_vals = []
        per_model = {}

        for suffix, mdata in model_results.items():
            if target_date in mdata["dates"]:
                vals = mdata["dates"][target_date]
                all_vals.extend(vals)
                per_model[mdata["name"]] = round((np.mean(vals) - 32) * 5 / 9, 1)

        if not all_vals:
            continue

        analysis = analyze_distribution(all_vals)
        analysis["target_date"] = target_date
        analysis["per_model_mean_c"] = per_model

        city_result[h_label] = analysis

        # Print distribution
        print(f"\n  {h_label.upper()} ({target_date}) - {analysis['n_members']} membri:")
        print(f"  Media: {analysis['mean_f']}F ({analysis['mean_c']}C)  "
              f"Std: {analysis['std_f']}F ({analysis['std_c']}C)")

        for c_val in sorted(analysis["distribution_c"].keys(), key=int):
            prob = analysis["distribution_c"][c_val]
            f_val = int(c_val) * 9 / 5 + 32
            bar = "#" * int(prob / 2)
            marker = " << MODA" if int(c_val) == analysis["mode_c"] else ""
            print(f"    {c_val:>3}C ({f_val:3.0f}F): {prob:5.1f}% {bar}{marker}")

        # Key metrics
        print(f"\n  Prob singolo bucket ({analysis['mode_c']}C): {analysis['mode_prob']}%")
        pm1_str = f"  Prob +-1C ({analysis['mode_c'] - 1} a {analysis['mode_c'] + 1}C): {analysis['pm1_prob']}%"
        if analysis["pm1_prob"] >= 80:
            pm1_str += "  ** SCOMMESSA OK **"
        print(pm1_str)

        bp = analysis["best_pair"]
        if bp[0] != bp[1]:
            print(f"  Miglior coppia ({bp[0]}-{bp[1]}C): {analysis['best_pair_prob']}%")

        # Model agreement
        if per_model:
            model_vals = list(per_model.values())
            spread = max(model_vals) - min(model_vals)
            agreement = "ALTA" if spread <= 1 else ("MEDIA" if spread <= 2 else "BASSA")
            print(f"  Accordo modelli: {agreement} (spread {spread:.1f}C)")

    all_results[city_name] = city_result
    time.sleep(1)  # rate limiting

# ── Save history ──
print(f"\n\nSalvataggio storico...")
save_data = {
    "date": today.isoformat(),
    "models": list(MODEL_DISPLAY.values()),
    "api_calls": api_calls,
    "cities": all_results,
}
with open(history_file, "w", encoding="utf-8") as f:
    json.dump(save_data, f, indent=2, ensure_ascii=False)
print(f"Salvato: {history_file}")


# ── Excel output ──
print(f"\nScrittura Excel: {out_path}")

summary_rows = {h: [] for h in ["1gg", "2gg"]}

for city_name in CITIES:
    if city_name not in all_results:
        continue
    cr = all_results[city_name]

    for h_key in ["1gg", "2gg"]:
        if h_key not in cr:
            continue
        d = cr[h_key]
        row = {
            "Citta": city_name,
            "Data": d.get("target_date", ""),
            "Media C": d["mean_c"],
            "Media F": d["mean_f"],
            "Std C": d["std_c"],
            "Moda C": d["mode_c"],
            "Moda F": d["mode_f"],
            "P(Moda)%": d["mode_prob"],
            "P(+-1C)%": d["pm1_prob"],
            "Coppia": f"{d['best_pair'][0]}-{d['best_pair'][1]}C"
                      if d["best_pair"][0] != d["best_pair"][1]
                      else f"{d['best_pair'][0]}C",
            "P(Coppia)%": d["best_pair_prob"],
            "N Membri": d["n_members"],
        }

        # Add per-model means
        for model_name, mean_c in d.get("per_model_mean_c", {}).items():
            row[model_name] = mean_c

        summary_rows[h_key].append(row)


with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
    for h_key, sheet_name in [("1gg", "Previsioni 1GG"), ("2gg", "Previsioni 2GG")]:
        if not summary_rows[h_key]:
            continue

        df = pd.DataFrame(summary_rows[h_key])
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        col_names = list(df.columns)

        for col_idx in range(1, len(col_names) + 1):
            ws.cell(row=1, column=col_idx).font = bold_white
            ws.cell(row=1, column=col_idx).fill = blue_fill
            ws.cell(row=1, column=col_idx).alignment = Alignment(horizontal="center")

        for row_idx in range(2, len(df) + 2):
            for col_idx, col_name in enumerate(col_names, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if col_name == "P(+-1C)%" and cell.value is not None:
                    if cell.value >= 80:
                        cell.fill = gold_fill
                    elif cell.value >= 70:
                        cell.fill = green_fill
                    else:
                        cell.fill = light_red
                    cell.font = bold
                elif col_name == "P(Moda)%" and cell.value is not None:
                    cell.font = bold
                elif col_name == "P(Coppia)%" and cell.value is not None:
                    if cell.value >= 80:
                        cell.fill = gold_fill
                    elif cell.value >= 70:
                        cell.fill = green_fill
                    cell.font = bold

        for col_idx in range(1, len(col_names) + 1):
            max_len = max(len(str(ws.cell(row=r, column=col_idx).value or ""))
                          for r in range(1, len(df) + 2))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 20)

    # Sheet 3: Distribuzioni dettagliate 1GG
    dist_rows = []
    for city_name in CITIES:
        if city_name not in all_results or "1gg" not in all_results[city_name]:
            continue
        d = all_results[city_name]["1gg"]
        for c_str, prob in sorted(d["distribution_c"].items(), key=lambda x: int(x[0])):
            c_val = int(c_str)
            dist_rows.append({
                "Citta": city_name,
                "C": c_val,
                "F": round(c_val * 9 / 5 + 32),
                "Probabilita%": prob,
                "Moda": "SI" if c_val == d["mode_c"] else "",
            })

    if dist_rows:
        df_dist = pd.DataFrame(dist_rows)
        df_dist.to_excel(writer, sheet_name="Distribuzioni 1GG", index=False)
        ws_d = writer.sheets["Distribuzioni 1GG"]
        for col_idx in range(1, len(df_dist.columns) + 1):
            ws_d.cell(row=1, column=col_idx).font = bold_white
            ws_d.cell(row=1, column=col_idx).fill = blue_fill
        for row_idx in range(2, len(df_dist) + 2):
            prob_cell = ws_d.cell(row=row_idx, column=4)
            moda_cell = ws_d.cell(row=row_idx, column=5)
            if prob_cell.value is not None and prob_cell.value >= 20:
                prob_cell.fill = gold_fill
                prob_cell.font = bold
            elif prob_cell.value is not None and prob_cell.value >= 10:
                prob_cell.fill = green_fill
            if moda_cell.value == "SI":
                for c in range(1, 6):
                    ws_d.cell(row=row_idx, column=c).font = bold


print(f"File salvato: {out_path}")

# ── Riepilogo finale ──
print(f"\n{'=' * 70}")
print(f"OPPORTUNITA LIVE - P(+-1C) >= 80%")
print(f"{'=' * 70}")
found = False
for city_name in CITIES:
    if city_name not in all_results:
        continue
    for h_key in ["1gg", "2gg"]:
        if h_key not in all_results[city_name]:
            continue
        d = all_results[city_name][h_key]
        if d["pm1_prob"] >= 80:
            found = True
            print(f"  {city_name:<15} {h_key.upper()}  data={d.get('target_date','')}  "
                  f"Moda={d['mode_c']}C  P(+-1C)={d['pm1_prob']}%  "
                  f"Std={d['std_c']}C")
if not found:
    print("  Nessuna citta con P(+-1C) >= 80% oggi.")

print(f"\n{'=' * 70}")
print(f"TUTTE LE CITTA - Riepilogo 1GG")
print(f"{'=' * 70}")
for city_name in CITIES:
    if city_name not in all_results or "1gg" not in all_results[city_name]:
        continue
    d = all_results[city_name]["1gg"]
    pm1 = d["pm1_prob"]
    marker = " **" if pm1 >= 80 else (" *" if pm1 >= 70 else "")
    print(f"  {city_name:<15} Moda={d['mode_c']:3d}C ({d['mode_f']:3d}F)  "
          f"P(+-1C)={pm1:5.1f}%  Std={d['std_c']:.1f}C{marker}")

print(f"\nAPI calls: {api_calls}")
n_history = len([f for f in os.listdir(HISTORY_DIR) if f.endswith(".json")])
print(f"Giorni nello storico ensemble: {n_history}")
print(f"\nPer raccogliere dati giornalmente, esegui questo script ogni giorno.")
print(f"Lo storico si accumula in: {HISTORY_DIR}")
