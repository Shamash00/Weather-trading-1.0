"""
Scarica previsioni ensemble da Open-Meteo per Londra (tutti i modelli)
e crea un foglio Excel con i risultati organizzati per modello.
"""

import requests
import pandas as pd
import numpy as np
import os
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUT_PATH = os.path.join(BASE_DIR, "Ensemble Tutti Modelli Londra.xlsx")

MODELS = [
    "icon_seamless_eps",
    "icon_global_eps",
    "icon_eu_eps",
    "icon_d2_eps",
    "meteoswiss_icon_ch1_ensemble",
    "meteoswiss_icon_ch2_ensemble",
    "ncep_gefs_seamless",
    "ncep_gefs025",
    "ncep_gefs05",
    "ncep_aigefs025",
    "ecmwf_ifs025_ensemble",
    "ecmwf_aifs025_ensemble",
    "gem_global_ensemble",
    "bom_access_global_ensemble",
    "ukmo_global_ensemble_20km",
    "ukmo_uk_ensemble_2km",
]

API_URL = "https://ensemble-api.open-meteo.com/v1/ensemble"
LAT = 51.505278
LON = 0.055278
START = "2026-03-07"
END = "2026-03-17"

# Colori
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=9)
STATS_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
STATS_FONT = Font(bold=True, size=9)
DATA_FONT = Font(size=9)
CENTER = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def fetch_model(model_name, variable_type="daily"):
    """Scarica dati ensemble per un singolo modello."""
    params = {
        "latitude": LAT,
        "longitude": LON,
        "models": model_name,
        "timezone": "auto",
        "start_date": START,
        "end_date": END,
    }

    if variable_type == "daily":
        params["daily"] = "temperature_2m_max,temperature_2m_min"
    else:
        params["hourly"] = "temperature_2m"

    try:
        r = requests.get(API_URL, params=params, timeout=60)
        if r.status_code != 200:
            print(f"  {model_name}: HTTP {r.status_code}")
            return None
        return r.json()
    except Exception as e:
        print(f"  {model_name}: {e}")
        return None


def extract_members(data, variable_type="daily"):
    """
    Estrae i dati dei membri dall'API response.
    Ritorna un dict: { variable: DataFrame con colonne [time, member_00, member_01, ...] }
    """
    section = data.get(variable_type, {})
    if not section:
        return {}

    times = section.get("time", [])
    if not times:
        return {}

    # Raggruppa le colonne per variabile base
    variables = {}
    for key in section:
        if key == "time":
            continue

        # Determina variabile base e member_id
        if "_member" in key:
            # es. "temperature_2m_max_member03" -> base="temperature_2m_max", member=3
            parts = key.rsplit("_member", 1)
            base_var = parts[0]
            member_id = int(parts[1])
        else:
            # Control run (member 00)
            base_var = key
            member_id = 0

        if base_var not in variables:
            variables[base_var] = {}
        variables[base_var][member_id] = section[key]

    result = {}
    for var_name, members in variables.items():
        df = pd.DataFrame({"time": times})
        for m_id in sorted(members.keys()):
            df[f"member_{m_id:02d}"] = members[m_id]
        result[var_name] = df

    return result


def write_model_sheet(writer, sheet_name, model_name, daily_data, n_members_info):
    """Scrive un foglio Excel per un modello con T max e T min."""

    var_max = None
    var_min = None
    for var_name, df in daily_data.items():
        if "max" in var_name:
            var_max = df
        elif "min" in var_name:
            var_min = df

    if var_max is None and var_min is None:
        return

    # Prepara il DataFrame principale
    all_sections = []

    for label, var_df in [("T MAX (C)", var_max), ("T MIN (C)", var_min)]:
        if var_df is None:
            continue

        member_cols = [c for c in var_df.columns if c.startswith("member_")]
        n_members = len(member_cols)
        n_members_info[sheet_name] = n_members

        # Trasponi: righe = membri, colonne = date
        dates = var_df["time"].tolist()
        member_data = []

        for col in member_cols:
            row = {"Membro": col}
            for i, d in enumerate(dates):
                row[d] = var_df[col].iloc[i]
            member_data.append(row)

        df_out = pd.DataFrame(member_data)

        # Aggiungi statistiche
        stats_rows = []
        date_cols = [c for c in df_out.columns if c != "Membro"]

        for stat_name, func in [
            ("Media", np.nanmean),
            ("Mediana", np.nanmedian),
            ("Std Dev", lambda x: np.nanstd(x, ddof=1)),
            ("Min", np.nanmin),
            ("Max", np.nanmax),
            ("P10", lambda x: np.nanpercentile(x, 10)),
            ("P25", lambda x: np.nanpercentile(x, 25)),
            ("P75", lambda x: np.nanpercentile(x, 75)),
            ("P90", lambda x: np.nanpercentile(x, 90)),
        ]:
            stat_row = {"Membro": stat_name}
            for d in date_cols:
                vals = df_out[d].dropna().values
                if len(vals) > 0:
                    stat_row[d] = round(func(vals), 2)
                else:
                    stat_row[d] = None
            stats_rows.append(stat_row)

        # Probabilità per soglie
        for threshold in [8, 9, 10, 11, 12, 13, 14, 15]:
            prob_row = {"Membro": f"P(T>={threshold})"}
            for d in date_cols:
                vals = df_out[d].dropna().values
                if len(vals) > 0:
                    prob_row[d] = f"{np.sum(vals >= threshold) / len(vals) * 100:.0f}%"
                else:
                    prob_row[d] = None
            stats_rows.append(prob_row)

        df_stats = pd.DataFrame(stats_rows)

        # Riga separatore con label
        sep = pd.DataFrame([{c: "" for c in df_out.columns}])
        sep.iloc[0, 0] = label

        all_sections.append(sep)
        all_sections.append(df_out)
        all_sections.append(pd.DataFrame([{c: "" for c in df_out.columns}]))
        all_sections.append(df_stats)
        all_sections.append(pd.DataFrame([{c: "" for c in df_out.columns}]))

    if not all_sections:
        return

    combined = pd.concat(all_sections, ignore_index=True)
    combined.to_excel(writer, sheet_name=sheet_name, index=False)


def format_sheet(ws):
    """Applica formattazione al foglio."""
    # Header row
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # Tutte le celle
    stats_labels = {"Media", "Mediana", "Std Dev", "Min", "Max",
                    "P10", "P25", "P75", "P90"}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        first_val = row[0].value
        is_section_header = first_val in ("T MAX (C)", "T MIN (C)")
        is_stats = first_val in stats_labels
        is_prob = isinstance(first_val, str) and first_val.startswith("P(T>=")

        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = CENTER

            if is_section_header:
                cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF", size=10)
            elif is_stats:
                cell.fill = STATS_FILL
                cell.font = STATS_FONT
                if isinstance(cell.value, float):
                    cell.number_format = "0.00"
            elif is_prob:
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                cell.font = Font(bold=True, size=9)
            else:
                cell.font = DATA_FONT
                if isinstance(cell.value, float):
                    cell.number_format = "0.0"

    # Larghezza colonne
    ws.column_dimensions["A"].width = 14
    for col_idx in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12

    # Freeze pane
    ws.freeze_panes = "B2"


def main():
    import time

    print("=== Ensemble Tutti Modelli - Londra ===\n")
    print(f"Periodo: {START} -> {END}")
    print(f"Modelli: {len(MODELS)}\n")

    model_data = {}
    n_members_info = {}

    for i, model in enumerate(MODELS):
        print(f"[{i+1}/{len(MODELS)}] {model}...", end=" ", flush=True)

        data = fetch_model(model, "daily")
        if data is None:
            print("SKIP")
            continue

        members = extract_members(data, "daily")
        if not members:
            print("nessun dato")
            continue

        # Conta membri
        first_var = next(iter(members.values()))
        n = len([c for c in first_var.columns if c.startswith("member_")])
        print(f"{n} membri")

        model_data[model] = members
        time.sleep(1)

    if not model_data:
        print("\nNessun dato scaricato!")
        return

    # ── Scrivi Excel ──
    print(f"\nScrittura Excel ({len(model_data)} modelli)...")

    with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
        # Foglio riepilogo
        summary_rows = []

        for model_name, daily in model_data.items():
            # Nome foglio (max 31 char per Excel)
            sheet_name = model_name.replace("_ensemble", "").replace("_seamless", "")
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]

            write_model_sheet(writer, sheet_name, model_name, daily, n_members_info)
            format_sheet(writer.sheets[sheet_name])

            # Info per riepilogo
            n_mem = n_members_info.get(sheet_name, "?")
            summary_rows.append({
                "Modello": model_name,
                "Foglio": sheet_name,
                "N. Membri": n_mem,
            })

        # Foglio riepilogo come primo foglio
        df_summary = pd.DataFrame(summary_rows)
        df_summary.to_excel(writer, sheet_name="Riepilogo", index=False)
        ws_sum = writer.sheets["Riepilogo"]
        for cell in ws_sum[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER
        ws_sum.column_dimensions["A"].width = 35
        ws_sum.column_dimensions["B"].width = 32
        ws_sum.column_dimensions["C"].width = 12

        # Sposta Riepilogo come primo foglio
        wb = writer.book
        wb.move_sheet("Riepilogo", offset=-len(wb.sheetnames) + 1)

    print(f"\nFile salvato: {OUT_PATH}")


if __name__ == "__main__":
    main()
