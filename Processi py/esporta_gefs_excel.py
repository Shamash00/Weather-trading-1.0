"""Esporta i dati GEFS storico scaricati in un Excel formattato."""

import pickle, os, glob
import pandas as pd
import numpy as np
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CACHE_DIR = os.path.join(BASE_DIR, "_gefs_storico")
out_path = os.path.join(BASE_DIR, "GEFS Storico Ensemble.xlsx")

files = sorted(glob.glob(os.path.join(CACHE_DIR, "*.pkl")))
print(f"File trovati: {len(files)}")

# --- Foglio 1: Dettaglio completo ---
rows = []
for f in files:
    with open(f, "rb") as fp:
        data = pickle.load(fp)
    for city, info in data["cities"].items():
        mean_c = round((info["mean_f"] - 32) * 5 / 9, 1)
        std_c = round(info["std_f"] * 5 / 9, 1)
        rows.append({
            "Data Run": data["date"],
            "Data Previsione": data["forecast_date"],
            "Citta": city,
            "Media F": info["mean_f"],
            "Media C": mean_c,
            "Std F": info["std_f"],
            "Std C": std_c,
            "Membri": info["n_members"],
            "Min F": round(min(info["members_f"]), 1),
            "Max F": round(max(info["members_f"]), 1),
            "Range F": round(max(info["members_f"]) - min(info["members_f"]), 1),
        })

df_detail = pd.DataFrame(rows)

# --- Foglio 2: Matrice Date x Citta (Media C) ---
pivot_mean = df_detail.pivot_table(index="Data Previsione", columns="Citta", values="Media C")
pivot_mean = pivot_mean.reset_index()

# --- Foglio 3: Matrice Date x Citta (Std F) ---
pivot_std = df_detail.pivot_table(index="Data Previsione", columns="Citta", values="Std F")
pivot_std = pivot_std.reset_index()

# --- Foglio 4: Distribuzione membri per ogni data/citta ---
rows_members = []
for f in files:
    with open(f, "rb") as fp:
        data = pickle.load(fp)
    for city, info in data["cities"].items():
        row = {
            "Data Previsione": data["forecast_date"],
            "Citta": city,
        }
        for i, val in enumerate(info["members_f"]):
            val_c = round((val - 32) * 5 / 9, 1)
            row[f"M{i:02d} (C)"] = val_c
        rows_members.append(row)

df_members = pd.DataFrame(rows_members)

# --- Stili ---
blue_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
yellow_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
bold_white = Font(bold=True, color="FFFFFF")
bold = Font(bold=True)
center = Alignment(horizontal="center")

# --- Scrivi Excel ---
print(f"Scrittura Excel...")

with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
    # Foglio 1: Dettaglio
    df_detail.to_excel(writer, sheet_name="Dettaglio", index=False)
    ws = writer.sheets["Dettaglio"]
    for col_idx in range(1, len(df_detail.columns) + 1):
        ws.cell(row=1, column=col_idx).font = bold_white
        ws.cell(row=1, column=col_idx).fill = blue_fill
        ws.cell(row=1, column=col_idx).alignment = center

    std_col = list(df_detail.columns).index("Std F") + 1
    for row_idx in range(2, len(df_detail) + 2):
        cell = ws.cell(row=row_idx, column=std_col)
        if cell.value is not None:
            if cell.value <= 1.0:
                cell.fill = green_fill
            elif cell.value >= 3.0:
                cell.fill = red_fill
            elif cell.value >= 2.0:
                cell.fill = yellow_fill

    for col_idx in range(1, len(df_detail.columns) + 1):
        max_len = max(len(str(ws.cell(row=r, column=col_idx).value or ""))
                      for r in range(1, len(df_detail) + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 20)

    # Foglio 2: Matrice Temperature
    pivot_mean.to_excel(writer, sheet_name="Matrice Temperature (C)", index=False)
    ws2 = writer.sheets["Matrice Temperature (C)"]
    cols2 = list(pivot_mean.columns)
    for col_idx in range(1, len(cols2) + 1):
        ws2.cell(row=1, column=col_idx).font = bold_white
        ws2.cell(row=1, column=col_idx).fill = blue_fill
        ws2.cell(row=1, column=col_idx).alignment = center

    for row_idx in range(2, len(pivot_mean) + 2):
        for col_idx in range(2, len(cols2) + 1):
            cell = ws2.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                cell.number_format = "0.0"
                cell.alignment = center
                v = cell.value
                if v >= 35:
                    cell.fill = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                elif v >= 25:
                    cell.fill = PatternFill(start_color="FF9933", end_color="FF9933", fill_type="solid")
                elif v >= 15:
                    cell.fill = yellow_fill
                elif v >= 5:
                    cell.fill = PatternFill(start_color="B4D7A8", end_color="B4D7A8", fill_type="solid")
                elif v >= -5:
                    cell.fill = PatternFill(start_color="9DC3E6", end_color="9DC3E6", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="6C8EBF", end_color="6C8EBF", fill_type="solid")
                    cell.font = Font(color="FFFFFF")

    for col_idx in range(1, len(cols2) + 1):
        max_len = max(len(str(ws2.cell(row=r, column=col_idx).value or ""))
                      for r in range(1, len(pivot_mean) + 2))
        ws2.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 18)

    # Foglio 3: Matrice Incertezza
    pivot_std.to_excel(writer, sheet_name="Matrice Incertezza (Std F)", index=False)
    ws3 = writer.sheets["Matrice Incertezza (Std F)"]
    cols3 = list(pivot_std.columns)
    for col_idx in range(1, len(cols3) + 1):
        ws3.cell(row=1, column=col_idx).font = bold_white
        ws3.cell(row=1, column=col_idx).fill = blue_fill
        ws3.cell(row=1, column=col_idx).alignment = center

    for row_idx in range(2, len(pivot_std) + 2):
        for col_idx in range(2, len(cols3) + 1):
            cell = ws3.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                cell.number_format = "0.0"
                cell.alignment = center
                v = cell.value
                if v <= 1.0:
                    cell.fill = green_fill
                    cell.font = bold
                elif v >= 3.0:
                    cell.fill = red_fill
                    cell.font = Font(color="FFFFFF", bold=True)
                elif v >= 2.0:
                    cell.fill = yellow_fill

    for col_idx in range(1, len(cols3) + 1):
        max_len = max(len(str(ws3.cell(row=r, column=col_idx).value or ""))
                      for r in range(1, len(pivot_std) + 2))
        ws3.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 18)

    # Foglio 4: Membri singoli
    df_members.to_excel(writer, sheet_name="Membri Ensemble", index=False)
    ws4 = writer.sheets["Membri Ensemble"]
    for col_idx in range(1, len(df_members.columns) + 1):
        ws4.cell(row=1, column=col_idx).font = bold_white
        ws4.cell(row=1, column=col_idx).fill = blue_fill
        ws4.cell(row=1, column=col_idx).alignment = center
    ws4.column_dimensions["A"].width = 16
    ws4.column_dimensions["B"].width = 15
    for col_idx in range(3, len(df_members.columns) + 1):
        ws4.column_dimensions[get_column_letter(col_idx)].width = 8

print(f"Excel salvato: {out_path}")
print(f"Righe dettaglio: {len(df_detail)}")
print(f"Date: {len(files)}")
print(f"Fogli: Dettaglio, Matrice Temperature (C), Matrice Incertezza (Std F), Membri Ensemble")
