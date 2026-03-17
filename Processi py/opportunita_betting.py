import pandas as pd
import os
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
input_path = os.path.join(BASE_DIR, "Ottimizzazione Ensemble Stagioni.xlsx")
out_path = os.path.join(BASE_DIR, "Opportunita Betting.xlsx")

SOGLIA = 80.0  # Verde% minimo per considerare l'opportunita

green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
gold_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
blue_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
light_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
bold = Font(bold=True)
bold_white = Font(bold=True, color="FFFFFF")
bold_big = Font(bold=True, size=13)


def format_sheet(ws, df, col_names):
    """Format header + color verde% cells + auto-width."""
    for col_idx in range(1, len(col_names) + 1):
        ws.cell(row=1, column=col_idx).font = bold_white
        ws.cell(row=1, column=col_idx).fill = blue_fill
        ws.cell(row=1, column=col_idx).alignment = Alignment(horizontal="center")

    for row_idx in range(2, len(df) + 2):
        for col_idx, col_name in enumerate(col_names, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if col_name == "Verde%" and cell.value is not None:
                if cell.value >= 85:
                    cell.fill = gold_fill
                else:
                    cell.fill = green_fill
                cell.font = bold

    for col_idx in range(1, len(col_names) + 1):
        max_len = max(len(str(ws.cell(row=r, column=col_idx).value or ""))
                      for r in range(1, len(df) + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)


print(f"Lettura dati da: {input_path}")
df_1gg = pd.read_excel(input_path, sheet_name="Operativa 1GG")
df_2gg = pd.read_excel(input_path, sheet_name="Operativa 2GG")

# Filtra sopra soglia
hot_1gg = df_1gg[df_1gg["Verde%"] >= SOGLIA].copy()
hot_2gg = df_2gg[df_2gg["Verde%"] >= SOGLIA].copy()

# Ordina per Verde% decrescente
hot_1gg = hot_1gg.sort_values("Verde%", ascending=False).reset_index(drop=True)
hot_2gg = hot_2gg.sort_values("Verde%", ascending=False).reset_index(drop=True)

# Seleziona colonne chiave per la tabella operativa
cols_out = ["Citta", "Stagione", "Metodo", "N Modelli", "Verde%", "MAE",
            "Modelli", "Correzioni BC (C)"]

hot_1gg_out = hot_1gg[cols_out].copy()
hot_2gg_out = hot_2gg[cols_out].copy()

print(f"\n1GG: {len(hot_1gg_out)} combinazioni citta x stagione con Verde% >= {SOGLIA}%")
print(f"2GG: {len(hot_2gg_out)} combinazioni citta x stagione con Verde% >= {SOGLIA}%")

# Console preview
print(f"\n{'='*80}")
print(f"OPPORTUNITA 1GG (Verde% >= {SOGLIA}%)")
print(f"{'='*80}")
for _, r in hot_1gg_out.iterrows():
    print(f"  {r['Citta']:<15} {r['Stagione']:<12} Verde%={r['Verde%']:5.1f}%  "
          f"Metodo={r['Metodo']}  Modelli: {r['Modelli'][:60]}")

if not hot_2gg_out.empty:
    print(f"\n{'='*80}")
    print(f"OPPORTUNITA 2GG (Verde% >= {SOGLIA}%)")
    print(f"{'='*80}")
    for _, r in hot_2gg_out.iterrows():
        print(f"  {r['Citta']:<15} {r['Stagione']:<12} Verde%={r['Verde%']:5.1f}%  "
              f"Metodo={r['Metodo']}  Modelli: {r['Modelli'][:60]}")

# Scrivi Excel
print(f"\nScrittura Excel: {out_path}")

with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
    # Foglio 1: Opportunita 1GG
    hot_1gg_out.to_excel(writer, sheet_name="Opportunita 1GG", index=False)
    format_sheet(writer.sheets["Opportunita 1GG"], hot_1gg_out, cols_out)

    # Foglio 2: Opportunita 2GG
    if not hot_2gg_out.empty:
        hot_2gg_out.to_excel(writer, sheet_name="Opportunita 2GG", index=False)
        format_sheet(writer.sheets["Opportunita 2GG"], hot_2gg_out, cols_out)

    # Foglio 3: Matrice citta x stagione (1GG) - vista rapida
    season_order = ["Inverno", "Primavera", "Estate", "Autunno"]
    cities = df_1gg["Citta"].unique()

    matrix_rows = []
    for city in cities:
        row = {"Citta": city}
        city_data = df_1gg[df_1gg["Citta"] == city]
        for season in season_order:
            s_data = city_data[city_data["Stagione"] == season]
            if not s_data.empty:
                v = s_data["Verde%"].values[0]
                row[season] = v
        matrix_rows.append(row)

    df_matrix = pd.DataFrame(matrix_rows)
    df_matrix.to_excel(writer, sheet_name="Matrice 1GG", index=False)
    ws_m = writer.sheets["Matrice 1GG"]

    mat_cols = list(df_matrix.columns)
    for col_idx in range(1, len(mat_cols) + 1):
        ws_m.cell(row=1, column=col_idx).font = bold_white
        ws_m.cell(row=1, column=col_idx).fill = blue_fill
        ws_m.cell(row=1, column=col_idx).alignment = Alignment(horizontal="center")

    for row_idx in range(2, len(df_matrix) + 2):
        for col_idx, col_name in enumerate(mat_cols, 1):
            cell = ws_m.cell(row=row_idx, column=col_idx)
            if col_name in season_order and cell.value is not None:
                v = cell.value
                if v >= 85:
                    cell.fill = gold_fill
                    cell.font = bold
                elif v >= SOGLIA:
                    cell.fill = green_fill
                    cell.font = bold
                elif v >= 75:
                    cell.fill = light_green
                cell.number_format = "0.0"
                cell.alignment = Alignment(horizontal="center")

    for col_idx in range(1, len(mat_cols) + 1):
        max_len = max(len(str(ws_m.cell(row=r, column=col_idx).value or ""))
                      for r in range(1, len(df_matrix) + 2))
        ws_m.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 20)

print(f"\nFile salvato: {out_path}")
print(f"\nLegenda colori: ORO = Verde% >= 85%, VERDE = Verde% >= 80%, VERDE CHIARO = >= 75%")
