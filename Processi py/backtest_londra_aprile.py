"""
Backtest GEFS Ensemble vs Polymarket – Londra Aprile 2025

Per ogni giorno di aprile 2025:
1. Carica le previsioni GEFS ensemble (31 membri) dal giorno PRIMA
2. Calcola la probabilita' di ogni bucket Polymarket (in °F)
3. Confronta con le odds REALI Polymarket del giorno prima (fidelity=1440)
4. Mostra edge (GEFS vs mercato) e il bucket vincitore reale
5. Calcola P&L simulato con strategia Kelly semplificata

Output: Excel "Backtest Londra Aprile 2025.xlsx"
"""

import pickle, os, re
import numpy as np
import pandas as pd
from collections import Counter
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

PROJECT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
GEFS_DIR = os.path.join(PROJECT_DIR, "_gefs_storico")
POLY_PKL = os.path.join(PROJECT_DIR, "_polymarket_april_2025_london_prices.pkl")
OUT_PATH = os.path.join(PROJECT_DIR, "Backtest Londra Aprile 2025.xlsx")

# ── Carica dati Polymarket con prezzi reali ──────────────────────────────

with open(POLY_PKL, "rb") as f:
    poly_data = pickle.load(f)


def member_to_bucket(val_f, buckets):
    """Assegna un valore in °F al bucket Polymarket corrispondente."""
    val_round = round(val_f)
    for b in buckets:
        lo = b["lo_f"]
        hi = b["hi_f"]
        if lo is None and hi is not None:
            if val_round <= hi:
                return b["label"]
        elif hi is None and lo is not None:
            if val_round >= lo:
                return b["label"]
        elif lo is not None and hi is not None:
            if lo <= val_round <= hi:
                return b["label"]
    return None


# ── Processa ogni giorno ──────────────────────────────────────────────────

results = []

for day in range(1, 31):
    if day not in poly_data:
        continue

    buckets = poly_data[day]
    forecast_date = f"2025-04-{day:02d}"
    run_date = f"2025-04-{day - 1:02d}" if day > 1 else "2025-03-31"

    pkl_path = os.path.join(GEFS_DIR, f"{run_date}.pkl")
    if not os.path.exists(pkl_path):
        print(f"  Apr {day}: GEFS mancante per run {run_date}")
        continue

    with open(pkl_path, "rb") as f:
        gefs_data = pickle.load(f)

    london = gefs_data["cities"].get("Londra")
    if not london:
        continue

    members_f = london["members_f"]

    # Probabilita' GEFS per ogni bucket
    bucket_counts = Counter()
    for val in members_f:
        bl = member_to_bucket(val, buckets)
        if bl:
            bucket_counts[bl] += 1

    total = len(members_f)

    # Bucket vincitore
    winner_label = None
    winner_temp_f = None
    for b in buckets:
        if b["winner"]:
            winner_label = b["label"]
            if b["lo_f"] is not None and b["hi_f"] is not None:
                winner_temp_f = (b["lo_f"] + b["hi_f"]) / 2
            elif b["lo_f"] is not None:
                winner_temp_f = b["lo_f"]
            elif b["hi_f"] is not None:
                winner_temp_f = b["hi_f"]
            break

    gefs_mean_f = np.mean(members_f)
    gefs_std_f = np.std(members_f)
    gefs_mean_c = (gefs_mean_f - 32) * 5 / 9
    winner_c = (winner_temp_f - 32) * 5 / 9 if winner_temp_f else None

    gefs_prob_winner = bucket_counts.get(winner_label, 0) / total * 100

    if bucket_counts:
        gefs_top_bucket = max(bucket_counts, key=bucket_counts.get)
        gefs_top_prob = bucket_counts[gefs_top_bucket] / total * 100
    else:
        gefs_top_bucket = "N/A"
        gefs_top_prob = 0

    hit = gefs_top_bucket == winner_label
    sorted_buckets_list = sorted(bucket_counts.items(), key=lambda x: -x[1])
    top2_labels = [x[0] for x in sorted_buckets_list[:2]]
    near_hit = winner_label in top2_labels
    error_f = abs(gefs_mean_f - winner_temp_f) if winner_temp_f else None

    # Distribuzione completa + prezzi Polymarket
    bucket_details = []
    for b in buckets:
        prob_gefs = bucket_counts.get(b["label"], 0) / total * 100
        poly_price = b.get("poly_price")  # Prezzo reale Polymarket giorno prima
        poly_pct = poly_price * 100 if poly_price is not None else None
        edge = (prob_gefs - poly_pct) if poly_pct is not None else None
        bucket_details.append({
            "label": b["label"],
            "lo_f": b["lo_f"],
            "hi_f": b["hi_f"],
            "winner": b["winner"],
            "gefs_pct": round(prob_gefs, 1),
            "poly_pct": round(poly_pct, 1) if poly_pct is not None else None,
            "edge": round(edge, 1) if edge is not None else None,
            "poly_price": poly_price,
        })

    results.append({
        "day": day,
        "forecast_date": forecast_date,
        "run_date": run_date,
        "gefs_mean_f": round(gefs_mean_f, 1),
        "gefs_mean_c": round(gefs_mean_c, 1),
        "gefs_std_f": round(gefs_std_f, 1),
        "winner_label": winner_label,
        "winner_temp_f": winner_temp_f,
        "winner_temp_c": round(winner_c, 1) if winner_c else None,
        "gefs_prob_winner": round(gefs_prob_winner, 1),
        "gefs_top_bucket": gefs_top_bucket,
        "gefs_top_prob": round(gefs_top_prob, 1),
        "hit": hit,
        "near_hit": near_hit,
        "error_f": round(error_f, 1) if error_f else None,
        "bucket_details": bucket_details,
        "members_f": members_f,
    })

print(f"Giorni processati: {len(results)} / {len(poly_data)}")

# ── Foglio 1: Riepilogo giornaliero ──────────────────────────────────────

rows_summary = []
for r in results:
    # Trova edge sul bucket vincitore
    winner_detail = [d for d in r["bucket_details"] if d["winner"]]
    edge_winner = winner_detail[0]["edge"] if winner_detail and winner_detail[0]["edge"] is not None else None
    poly_winner = winner_detail[0]["poly_pct"] if winner_detail and winner_detail[0]["poly_pct"] is not None else None

    # Trova il bucket con edge massimo
    best_edge_bucket = max(r["bucket_details"], key=lambda x: x["edge"] if x["edge"] is not None else -999)

    rows_summary.append({
        "Data": r["forecast_date"],
        "Run GEFS": r["run_date"],
        "Media GEFS (F)": r["gefs_mean_f"],
        "Media GEFS (C)": r["gefs_mean_c"],
        "Std (F)": r["gefs_std_f"],
        "Bucket Vincitore": r["winner_label"],
        "Temp Reale (F)": r["winner_temp_f"],
        "Temp Reale (C)": r["winner_temp_c"],
        "P(GEFS) Vinc.": r["gefs_prob_winner"],
        "P(Poly) Vinc.": poly_winner,
        "Edge Vinc.": edge_winner,
        "Top GEFS": r["gefs_top_bucket"],
        "P(GEFS) Top": r["gefs_top_prob"],
        "Centrato?": "SI" if r["hit"] else "NO",
        "Top 2?": "SI" if r["near_hit"] else "NO",
        "Errore (F)": r["error_f"],
        "Best Edge Bucket": best_edge_bucket["label"],
        "Best Edge %": best_edge_bucket["edge"],
    })

df_summary = pd.DataFrame(rows_summary)

# ── Foglio 2: Confronto GEFS vs Polymarket per ogni bucket ───────────────

rows_confronto = []
for r in results:
    for d in r["bucket_details"]:
        rows_confronto.append({
            "Data": r["forecast_date"],
            "Bucket": d["label"],
            "P(GEFS) %": d["gefs_pct"],
            "P(Polymarket) %": d["poly_pct"],
            "Edge (GEFS-Poly) %": d["edge"],
            "Vincitore": "SI" if d["winner"] else "",
            "N. Membri /31": round(d["gefs_pct"] / 100 * 31),
        })

df_confronto = pd.DataFrame(rows_confronto)

# ── Foglio 3: Simulazione P&L con prezzi REALI ──────────────────────────
# Strategia: scommetti $10 sul bucket dove GEFS ha piu' edge vs Polymarket
# Condizione: edge >= 10% (GEFS crede molto di piu' del mercato)

BET_AMOUNT = 10
MIN_EDGE = 10  # Scommetti solo se edge >= 10%

rows_pnl = []
total_pnl = 0
total_invested = 0

for r in results:
    # Trova il bucket con edge massimo
    valid_bets = [d for d in r["bucket_details"]
                  if d["edge"] is not None and d["edge"] >= MIN_EDGE
                  and d["poly_price"] is not None and d["poly_price"] > 0.01]

    if not valid_bets:
        rows_pnl.append({
            "Data": r["forecast_date"],
            "Azione": "SKIP (edge < 10%)",
            "Bucket Scelto": "-",
            "P(GEFS) %": "-",
            "P(Poly) %": "-",
            "Edge %": "-",
            "Prezzo Acquisto": "-",
            "Risultato": "-",
            "P&L ($)": 0,
            "P&L Cumulato ($)": round(total_pnl, 2),
        })
        continue

    # Scegli il bucket con edge piu' alto
    best = max(valid_bets, key=lambda x: x["edge"])
    buy_price = best["poly_price"]
    shares = BET_AMOUNT / buy_price
    won = best["winner"]

    if won:
        pnl = shares * 1.0 - BET_AMOUNT  # Payout $1 per share
    else:
        pnl = -BET_AMOUNT

    total_pnl += pnl
    total_invested += BET_AMOUNT

    rows_pnl.append({
        "Data": r["forecast_date"],
        "Azione": "BET",
        "Bucket Scelto": best["label"],
        "P(GEFS) %": best["gefs_pct"],
        "P(Poly) %": best["poly_pct"],
        "Edge %": best["edge"],
        "Prezzo Acquisto": round(buy_price, 3),
        "Risultato": "VINTO" if won else "PERSO",
        "P&L ($)": round(pnl, 2),
        "P&L Cumulato ($)": round(total_pnl, 2),
    })

df_pnl = pd.DataFrame(rows_pnl)

# ── Foglio 4: Statistiche ────────────────────────────────────────────────

n_days = len(results)
n_hits = sum(1 for r in results if r["hit"])
n_near = sum(1 for r in results if r["near_hit"])
avg_prob_winner = np.mean([r["gefs_prob_winner"] for r in results])
avg_error = np.mean([r["error_f"] for r in results if r["error_f"] is not None])
avg_top_prob = np.mean([r["gefs_top_prob"] for r in results])

# Edge medio GEFS vs Poly sul bucket vincitore
edges_winner = []
for r in results:
    for d in r["bucket_details"]:
        if d["winner"] and d["edge"] is not None:
            edges_winner.append(d["edge"])
avg_edge_winner = np.mean(edges_winner) if edges_winner else 0

n_bets = sum(1 for _, row in df_pnl.iterrows() if row["Azione"] == "BET")
n_wins = sum(1 for _, row in df_pnl.iterrows() if row["Risultato"] == "VINTO")

stats = [
    ("Giorni analizzati", n_days),
    ("", ""),
    ("=== ACCURATEZZA GEFS ===", ""),
    ("Bucket centrato (top 1)", f"{n_hits}/{n_days} ({n_hits/n_days*100:.0f}%)"),
    ("Bucket centrato (top 2)", f"{n_near}/{n_days} ({n_near/n_days*100:.0f}%)"),
    ("P(GEFS) media del vincitore", f"{avg_prob_winner:.1f}%"),
    ("P(GEFS) media del top bucket", f"{avg_top_prob:.1f}%"),
    ("Errore medio (F)", f"{avg_error:.1f}°F"),
    ("Errore medio (C)", f"{avg_error * 5/9:.1f}°C"),
    ("", ""),
    ("=== GEFS vs POLYMARKET ===", ""),
    ("Edge medio GEFS sul vincitore", f"{avg_edge_winner:+.1f}%"),
    ("Interpretazione", "Positivo = GEFS dava piu probabilita del mercato al vincitore"),
    ("", ""),
    ("=== SIMULAZIONE P&L (prezzi REALI Polymarket) ===", ""),
    ("Soglia edge minima", f"{MIN_EDGE}%"),
    ("Importo per scommessa", f"${BET_AMOUNT}"),
    ("Scommesse piazzate", f"{n_bets}/{n_days}"),
    ("Scommesse vinte", f"{n_wins}/{n_bets} ({n_wins/n_bets*100:.0f}%)" if n_bets else "0"),
    ("P&L totale", f"${total_pnl:.2f}"),
    ("Totale investito", f"${total_invested:.2f}"),
    ("P&L medio per scommessa", f"${total_pnl/n_bets:.2f}" if n_bets else "$0"),
    ("ROI", f"{total_pnl/total_invested*100:.1f}%" if total_invested else "0%"),
    ("", ""),
    ("=== CONFIGURAZIONE ===", ""),
    ("Modello", "GEFS 31 membri, run 00Z giorno prima, TMAX forecast 1GG"),
    ("Prezzi Polymarket", "Reali (fidelity=1440, ultimo prezzo del giorno prima)"),
    ("Citta", "Londra (London City Airport - EGLC)"),
    ("Periodo", "Aprile 2025 (1-30)"),
]

df_stats = pd.DataFrame(stats, columns=["Metrica", "Valore"])

# ── Scrivi Excel ──────────────────────────────────────────────────────────

print("Scrittura Excel...")

blue = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
green = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
dark_green = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
red = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
gold = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
light_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
light_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
orange = PatternFill(start_color="FF9933", end_color="FF9933", fill_type="solid")
bold_white = Font(bold=True, color="FFFFFF")
bold = Font(bold=True)
center = Alignment(horizontal="center")

with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:

    # --- Foglio 1: Riepilogo ---
    df_summary.to_excel(writer, sheet_name="Riepilogo", index=False)
    ws1 = writer.sheets["Riepilogo"]
    for c in range(1, len(df_summary.columns) + 1):
        ws1.cell(row=1, column=c).font = bold_white
        ws1.cell(row=1, column=c).fill = blue
        ws1.cell(row=1, column=c).alignment = center

    col_hit = list(df_summary.columns).index("Centrato?") + 1
    col_near = list(df_summary.columns).index("Top 2?") + 1
    col_prob_g = list(df_summary.columns).index("P(GEFS) Vinc.") + 1
    col_prob_p = list(df_summary.columns).index("P(Poly) Vinc.") + 1
    col_edge = list(df_summary.columns).index("Edge Vinc.") + 1
    col_best_edge = list(df_summary.columns).index("Best Edge %") + 1

    for row in range(2, len(df_summary) + 2):
        if ws1.cell(row=row, column=col_hit).value == "SI":
            ws1.cell(row=row, column=col_hit).fill = green
            ws1.cell(row=row, column=col_hit).font = bold
        else:
            ws1.cell(row=row, column=col_hit).fill = light_red
        if ws1.cell(row=row, column=col_near).value == "SI":
            ws1.cell(row=row, column=col_near).fill = light_green

        # Edge coloring
        for ec in [col_edge, col_best_edge]:
            e = ws1.cell(row=row, column=ec).value
            if e is not None and isinstance(e, (int, float)):
                if e >= 15:
                    ws1.cell(row=row, column=ec).fill = dark_green
                    ws1.cell(row=row, column=ec).font = Font(color="FFFFFF", bold=True)
                elif e >= 5:
                    ws1.cell(row=row, column=ec).fill = green
                elif e <= -10:
                    ws1.cell(row=row, column=ec).fill = red
                    ws1.cell(row=row, column=ec).font = Font(color="FFFFFF")

    for c in range(1, len(df_summary.columns) + 1):
        max_len = max(len(str(ws1.cell(row=r, column=c).value or ""))
                      for r in range(1, len(df_summary) + 2))
        ws1.column_dimensions[get_column_letter(c)].width = min(max_len + 3, 24)

    # --- Foglio 2: Confronto GEFS vs Polymarket ---
    df_confronto.to_excel(writer, sheet_name="GEFS vs Polymarket", index=False)
    ws2 = writer.sheets["GEFS vs Polymarket"]
    for c in range(1, len(df_confronto.columns) + 1):
        ws2.cell(row=1, column=c).font = bold_white
        ws2.cell(row=1, column=c).fill = blue
        ws2.cell(row=1, column=c).alignment = center

    col_win = list(df_confronto.columns).index("Vincitore") + 1
    col_edge2 = list(df_confronto.columns).index("Edge (GEFS-Poly) %") + 1
    for row in range(2, len(df_confronto) + 2):
        if ws2.cell(row=row, column=col_win).value == "SI":
            for c in range(1, len(df_confronto.columns) + 1):
                ws2.cell(row=row, column=c).fill = gold
                ws2.cell(row=row, column=c).font = bold
        else:
            e = ws2.cell(row=row, column=col_edge2).value
            if e is not None and isinstance(e, (int, float)):
                if e >= 15:
                    ws2.cell(row=row, column=col_edge2).fill = dark_green
                    ws2.cell(row=row, column=col_edge2).font = Font(color="FFFFFF", bold=True)
                elif e >= 5:
                    ws2.cell(row=row, column=col_edge2).fill = green
                elif e <= -10:
                    ws2.cell(row=row, column=col_edge2).fill = light_red

    for c in range(1, len(df_confronto.columns) + 1):
        max_len = max(len(str(ws2.cell(row=r, column=c).value or ""))
                      for r in range(1, len(df_confronto) + 2))
        ws2.column_dimensions[get_column_letter(c)].width = min(max_len + 3, 25)

    # --- Foglio 3: Simulazione P&L ---
    df_pnl.to_excel(writer, sheet_name="Simulazione P&L", index=False)
    ws3 = writer.sheets["Simulazione P&L"]
    for c in range(1, len(df_pnl.columns) + 1):
        ws3.cell(row=1, column=c).font = bold_white
        ws3.cell(row=1, column=c).fill = blue
        ws3.cell(row=1, column=c).alignment = center

    col_res = list(df_pnl.columns).index("Risultato") + 1
    col_pnl_val = list(df_pnl.columns).index("P&L ($)") + 1
    col_cum = list(df_pnl.columns).index("P&L Cumulato ($)") + 1
    for row in range(2, len(df_pnl) + 2):
        if ws3.cell(row=row, column=col_res).value == "VINTO":
            ws3.cell(row=row, column=col_res).fill = green
            ws3.cell(row=row, column=col_res).font = bold
        elif ws3.cell(row=row, column=col_res).value == "PERSO":
            ws3.cell(row=row, column=col_res).fill = red
            ws3.cell(row=row, column=col_res).font = Font(color="FFFFFF", bold=True)
        pnl_val = ws3.cell(row=row, column=col_pnl_val).value
        if pnl_val is not None and isinstance(pnl_val, (int, float)):
            if pnl_val > 0:
                ws3.cell(row=row, column=col_pnl_val).fill = light_green
            elif pnl_val < 0:
                ws3.cell(row=row, column=col_pnl_val).fill = light_red
        cum_val = ws3.cell(row=row, column=col_cum).value
        if cum_val is not None and isinstance(cum_val, (int, float)):
            if cum_val > 0:
                ws3.cell(row=row, column=col_cum).fill = light_green
            elif cum_val < 0:
                ws3.cell(row=row, column=col_cum).fill = light_red

    for c in range(1, len(df_pnl.columns) + 1):
        max_len = max(len(str(ws3.cell(row=r, column=c).value or ""))
                      for r in range(1, len(df_pnl) + 2))
        ws3.column_dimensions[get_column_letter(c)].width = min(max_len + 3, 22)

    # --- Foglio 4: Statistiche ---
    df_stats.to_excel(writer, sheet_name="Statistiche", index=False)
    ws4 = writer.sheets["Statistiche"]
    ws4.cell(row=1, column=1).font = bold_white
    ws4.cell(row=1, column=1).fill = blue
    ws4.cell(row=1, column=2).font = bold_white
    ws4.cell(row=1, column=2).fill = blue
    ws4.column_dimensions["A"].width = 40
    ws4.column_dimensions["B"].width = 65

    for row in range(2, len(df_stats) + 2):
        val = str(ws4.cell(row=row, column=1).value or "")
        if val.startswith("==="):
            ws4.cell(row=row, column=1).font = Font(bold=True, size=12)
            ws4.cell(row=row, column=2).font = Font(bold=True, size=12)

print(f"\nExcel salvato: {OUT_PATH}")
print(f"\n{'='*60}")
print(f"RISULTATI BACKTEST LONDRA - APRILE 2025")
print(f"{'='*60}")
print(f"Giorni: {n_days}")
print(f"Centrato (top 1): {n_hits}/{n_days} ({n_hits/n_days*100:.0f}%)")
print(f"Centrato (top 2): {n_near}/{n_days} ({n_near/n_days*100:.0f}%)")
print(f"P(GEFS) media vincitore: {avg_prob_winner:.1f}%")
print(f"Edge medio GEFS sul vincitore: {avg_edge_winner:+.1f}%")
print(f"Errore medio: {avg_error:.1f}F ({avg_error*5/9:.1f}C)")
print(f"\nSimulazione P&L (edge >= {MIN_EDGE}%):")
print(f"  Scommesse: {n_bets}, Vinte: {n_wins}")
print(f"  Investito: ${total_invested:.0f}")
print(f"  P&L totale: ${total_pnl:.2f}")
print(f"  ROI: {total_pnl/total_invested*100:.1f}%" if total_invested else "  ROI: N/A")
