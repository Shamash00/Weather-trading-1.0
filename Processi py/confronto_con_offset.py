import pandas as pd
import os
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill, Font

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
forecast_path = os.path.join(BASE_DIR, "Previsioni FM-15 Tutte le Citta 2021-2025 Daily Max.xlsx")
storico_path = os.path.join(BASE_DIR, "Temperature Storiche FM-15 Tutte le Citta 2021-2025 Daily Max.xlsx")
out_path = os.path.join(BASE_DIR, "Confronto con Offset Corretto.xlsx")

green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

print("Caricamento file...")
xl_forecast = pd.ExcelFile(forecast_path)
xl_storico = pd.ExcelFile(storico_path)

# Step 1: identify cities with bias/MAE >= 70% and compute their bias
print("\n--- Analisi bias ---")
city_bias = {}
for sheet in xl_forecast.sheet_names:
    if sheet not in xl_storico.sheet_names:
        continue
    df_fc = pd.read_excel(xl_forecast, sheet_name=sheet)
    df_st = pd.read_excel(xl_storico, sheet_name=sheet)
    df_fc["Data"] = pd.to_datetime(df_fc["Data"]).dt.date
    df_st["Data"] = pd.to_datetime(df_st["Data"]).dt.date
    merged = pd.merge(df_fc, df_st, on="Data", how="inner")

    # Compute bias on Forecast (ultimo aggiornamento) in F
    delta_f = merged["Max_Forecast_F"] - merged["Max_Temperatura_F"]
    valid = delta_f.dropna()
    bias_f = valid.mean()
    mae_f = valid.abs().mean()
    ratio = abs(bias_f) / mae_f * 100 if mae_f > 0 else 0

    # Same for Celsius
    delta_c = merged["Max_Forecast_C"] - merged["Max_Temperatura_C"]
    valid_c = delta_c.dropna()
    bias_c = valid_c.mean()

    if ratio >= 70:
        city_bias[sheet] = {"bias_f": round(bias_f, 1), "bias_c": round(bias_c, 1),
                            "ratio": round(ratio, 0)}
        print(f"  {sheet}: bias_F={bias_f:+.1f}, bias_C={bias_c:+.1f}, ratio={ratio:.0f}%")

print(f"\n{len(city_bias)} citta da correggere: {', '.join(city_bias.keys())}")

# Step 2: generate corrected file for these cities
print("\n--- Generazione file corretto ---")
with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
    for sheet in city_bias:
        print(f"\n{sheet} (offset F: {city_bias[sheet]['bias_f']:+.1f}, "
              f"C: {city_bias[sheet]['bias_c']:+.1f})...")

        df_fc = pd.read_excel(xl_forecast, sheet_name=sheet)
        df_st = pd.read_excel(xl_storico, sheet_name=sheet)
        df_fc["Data"] = pd.to_datetime(df_fc["Data"]).dt.date
        df_st["Data"] = pd.to_datetime(df_st["Data"]).dt.date
        merged = pd.merge(df_fc, df_st, on="Data", how="inner")

        bias_f = city_bias[sheet]["bias_f"]
        bias_c = city_bias[sheet]["bias_c"]

        # Corrected forecasts = original - bias
        out = pd.DataFrame()
        out["Data"] = merged["Data"]

        # Fahrenheit: corrected values and deltas
        out["Previsione_F"] = merged["Max_Forecast_F"]
        out["Prev_Corretta_F"] = (merged["Max_Forecast_F"] - bias_f).round(0).astype("Int64")
        out["Prev_1GG_F"] = merged["Max_PrevDay1_F"]
        out["Prev1GG_Corretta_F"] = (merged["Max_PrevDay1_F"] - bias_f).round(0).astype("Int64")
        out["Prev_2GG_F"] = merged["Max_PrevDay2_F"]
        out["Prev2GG_Corretta_F"] = (merged["Max_PrevDay2_F"] - bias_f).round(0).astype("Int64")
        out["Registrata_F"] = merged["Max_Temperatura_F"]
        out["Delta_Orig_F"] = merged["Max_Forecast_F"] - merged["Max_Temperatura_F"]
        out["Delta_Corr_F"] = out["Prev_Corretta_F"] - merged["Max_Temperatura_F"]
        out["Delta1GG_Corr_F"] = out["Prev1GG_Corretta_F"] - merged["Max_Temperatura_F"]
        out["Delta2GG_Corr_F"] = out["Prev2GG_Corretta_F"] - merged["Max_Temperatura_F"]

        # Celsius: corrected values and deltas
        out["Previsione_C"] = merged["Max_Forecast_C"]
        out["Prev_Corretta_C"] = (merged["Max_Forecast_C"] - bias_c).round(1)
        out["Prev_1GG_C"] = merged["Max_PrevDay1_C"]
        out["Prev1GG_Corretta_C"] = (merged["Max_PrevDay1_C"] - bias_c).round(1)
        out["Prev_2GG_C"] = merged["Max_PrevDay2_C"]
        out["Prev2GG_Corretta_C"] = (merged["Max_PrevDay2_C"] - bias_c).round(1)
        out["Registrata_C"] = merged["Max_Temperatura_C"]
        out["Delta_Orig_C"] = (merged["Max_Forecast_C"] - merged["Max_Temperatura_C"]).round(1)
        out["Delta_Corr_C"] = (out["Prev_Corretta_C"] - merged["Max_Temperatura_C"]).round(1)
        out["Delta1GG_Corr_C"] = (out["Prev1GG_Corretta_C"] - merged["Max_Temperatura_C"]).round(1)
        out["Delta2GG_Corr_C"] = (out["Prev2GG_Corretta_C"] - merged["Max_Temperatura_C"]).round(1)

        out.to_excel(writer, sheet_name=sheet, index=False)

        ws = writer.sheets[sheet]
        max_row = ws.max_row

        # Conditional formatting on corrected delta columns
        # F corrected deltas: J (Delta_Corr_F), K (Delta1GG_Corr_F), L (Delta2GG_Corr_F)
        # Also original delta: I (Delta_Orig_F)
        # C corrected deltas: V (Delta_Corr_C), W (Delta1GG_Corr_C), X (Delta2GG_Corr_C)
        # Also original delta: U (Delta_Orig_C)

        # F original delta col I, corrected J, K, L
        for col in ["I", "J", "K", "L"]:
            rng = f"{col}2:{col}{max_row}"
            ws.conditional_formatting.add(rng, FormulaRule(
                formula=[f'AND({col}2<>"", ABS({col}2)<=1)'],
                fill=green_fill, stopIfTrue=True))
            ws.conditional_formatting.add(rng, FormulaRule(
                formula=[f'AND({col}2<>"", ABS({col}2)>1, ABS({col}2)<3)'],
                fill=yellow_fill, stopIfTrue=True))
            ws.conditional_formatting.add(rng, FormulaRule(
                formula=[f'AND({col}2<>"", ABS({col}2)>=3)'],
                fill=red_fill, stopIfTrue=True))

        # C original delta col U, corrected V, W, X
        for col in ["U", "V", "W", "X"]:
            rng = f"{col}2:{col}{max_row}"
            ws.conditional_formatting.add(rng, FormulaRule(
                formula=[f'AND({col}2<>"", ABS({col}2)<=1)'],
                fill=green_fill, stopIfTrue=True))
            ws.conditional_formatting.add(rng, FormulaRule(
                formula=[f'AND({col}2<>"", ABS({col}2)>1, ABS({col}2)<2)'],
                fill=yellow_fill, stopIfTrue=True))
            ws.conditional_formatting.add(rng, FormulaRule(
                formula=[f'AND({col}2<>"", ABS({col}2)>=2)'],
                fill=red_fill, stopIfTrue=True))

        # --- Summary ---
        delta_f_orig = ["Delta_Orig_F"]
        delta_f_corr = ["Delta_Corr_F", "Delta1GG_Corr_F", "Delta2GG_Corr_F"]
        delta_c_orig = ["Delta_Orig_C"]
        delta_c_corr = ["Delta_Corr_C", "Delta1GG_Corr_C", "Delta2GG_Corr_C"]
        all_cols = delta_f_orig + delta_f_corr + delta_c_orig + delta_c_corr

        valid_mask = out[all_cols].notna().all(axis=1)
        valid_data = out[valid_mask]
        n_valid = len(valid_data)

        # Column name -> Excel column index (1-based)
        col_names = list(out.columns)
        col_idx = {name: col_names.index(name) + 1 for name in all_cols}

        summary_row = max_row + 2
        bold = Font(bold=True)

        for i, label in enumerate(["Verde %", "Giallo %", "Rosso %",
                                    "Delta Medio", "Errore Medio Abs"]):
            cell = ws.cell(row=summary_row + i, column=1, value=label)
            cell.font = bold

        # Write offset info
        ws.cell(row=summary_row + 6, column=1, value=f"Offset applicato: F {bias_f:+.1f}, C {bias_c:+.1f}").font = bold

        for col_name in delta_f_orig + delta_f_corr:
            ci = col_idx[col_name]
            abs_vals = valid_data[col_name].abs()
            g = (abs_vals <= 1).sum() / n_valid * 100
            y = ((abs_vals > 1) & (abs_vals < 3)).sum() / n_valid * 100
            r = (abs_vals >= 3).sum() / n_valid * 100
            mean_delta = round(float(valid_data[col_name].mean()), 1)
            mae = round(float(abs_vals.mean()), 1)

            cg = ws.cell(row=summary_row, column=ci, value=round(g, 1))
            cg.fill = green_fill; cg.font = bold; cg.number_format = '0.0"%"'
            cy = ws.cell(row=summary_row + 1, column=ci, value=round(y, 1))
            cy.fill = yellow_fill; cy.font = bold; cy.number_format = '0.0"%"'
            cr = ws.cell(row=summary_row + 2, column=ci, value=round(r, 1))
            cr.fill = red_fill; cr.font = bold; cr.number_format = '0.0"%"'
            cm = ws.cell(row=summary_row + 3, column=ci, value=mean_delta)
            cm.font = bold; cm.number_format = '0.0'
            ca = ws.cell(row=summary_row + 4, column=ci, value=mae)
            ca.font = bold; ca.number_format = '0.0'

        for col_name in delta_c_orig + delta_c_corr:
            ci = col_idx[col_name]
            abs_vals = valid_data[col_name].abs()
            g = (abs_vals <= 1).sum() / n_valid * 100
            y = ((abs_vals > 1) & (abs_vals < 2)).sum() / n_valid * 100
            r = (abs_vals >= 2).sum() / n_valid * 100
            mean_delta = round(float(valid_data[col_name].mean()), 1)
            mae = round(float(abs_vals.mean()), 1)

            cg = ws.cell(row=summary_row, column=ci, value=round(g, 1))
            cg.fill = green_fill; cg.font = bold; cg.number_format = '0.0"%"'
            cy = ws.cell(row=summary_row + 1, column=ci, value=round(y, 1))
            cy.fill = yellow_fill; cy.font = bold; cy.number_format = '0.0"%"'
            cr = ws.cell(row=summary_row + 2, column=ci, value=round(r, 1))
            cr.fill = red_fill; cr.font = bold; cr.number_format = '0.0"%"'
            cm = ws.cell(row=summary_row + 3, column=ci, value=mean_delta)
            cm.font = bold; cm.number_format = '0.0'
            ca = ws.cell(row=summary_row + 4, column=ci, value=mae)
            ca.font = bold; ca.number_format = '0.0'

        # Print comparison
        orig_green = (valid_data["Delta_Orig_F"].abs() <= 1).sum() / n_valid * 100
        corr_green = (valid_data["Delta_Corr_F"].abs() <= 1).sum() / n_valid * 100
        print(f"  Verde% originale: {orig_green:.1f}% -> corretto: {corr_green:.1f}% "
              f"({corr_green - orig_green:+.1f}pp)")

print(f"\nFile salvato: {out_path}")
