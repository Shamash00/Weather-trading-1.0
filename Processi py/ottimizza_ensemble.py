import pandas as pd
import numpy as np
import os
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
storico_path = os.path.join(BASE_DIR, "Temperature Storiche",
                            "Temperature Storiche FM-15 Tutte le Citta 2021-2025 Daily Max.xlsx")
CITIES = ['Ankara','Atlanta','BuenosAires','Chicago','Dallas','Londra','Lucknow',
          'Miami','Monaco','New York','Parigi','SaoPaulo','Seattle','Seoul',
          'Shanghai','Singapore','TelAviv','Tokyo','Toronto','Wellington']
CACHE_DIR = os.path.join(BASE_DIR, "_cache_modelli")
out_path = os.path.join(BASE_DIR, "Ottimizzazione Ensemble per Citta.xlsx")

green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
blue_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
bold = Font(bold=True)
bold_white = Font(bold=True, color="FFFFFF")


def compute_daily_max(hourly_df):
    hourly_df["Data"] = hourly_df["time"].dt.date
    prev1 = [c for c in hourly_df.columns if c.startswith("temperature_2m_previous_day1_")]
    prev2 = [c for c in hourly_df.columns if c.startswith("temperature_2m_previous_day2_")]
    agg = {c: "max" for c in prev1 + prev2}
    daily = hourly_df.groupby("Data").agg(agg).reset_index()
    rename = {}
    for c in prev1:
        rename[c] = f"1GG_{c.replace('temperature_2m_previous_day1_', '')}"
    for c in prev2:
        rename[c] = f"2GG_{c.replace('temperature_2m_previous_day2_', '')}"
    daily.rename(columns=rename, inplace=True)
    return daily


SEAMLESS_FAMILIES = {
    "icon_seamless": ["icon_d2", "icon_eu", "icon_global"],
    "gfs_seamless": ["gfs_hrrr", "gfs_global"],
    "gem_seamless": ["gem_hrdps_continental", "gem_hrdps_west", "gem_regional", "gem_global"],
    "jma_seamless": ["jma_msm", "jma_gsm"],
    "kma_seamless": ["kma_ldps", "kma_gdps"],
    "meteoswiss_icon_seamless": ["meteoswiss_icon_ch1", "meteoswiss_icon_ch2"],
    "meteofrance_seamless": ["meteofrance_arome_france_hd", "meteofrance_arome_france",
                              "meteofrance_arpege_europe", "meteofrance_arpege_world"],
    "ukmo_seamless": ["ukmo_global_deterministic_10km", "ukmo_uk_deterministic_2km"],
    "knmi_seamless": ["knmi_harmonie_arome_europe", "knmi_harmonie_arome_netherlands"],
    "dmi_seamless": ["dmi_harmonie_arome_europe"],
    "metno_seamless": ["metno_nordic"],
}


def deduplicate_models(merged, model_cols, prefix):
    """Remove seamless models that are duplicates of their components."""
    to_remove = set()
    for seamless, components in SEAMLESS_FAMILIES.items():
        s_col = f"{prefix}_{seamless}"
        if s_col not in model_cols:
            continue
        for comp in components:
            c_col = f"{prefix}_{comp}"
            if c_col not in model_cols:
                continue
            both_valid = merged[s_col].notna() & merged[c_col].notna()
            if both_valid.sum() == 0:
                continue
            if (merged.loc[both_valid, s_col] == merged.loc[both_valid, c_col]).all():
                to_remove.add(s_col)
                break
    return [c for c in model_cols if c not in to_remove]


def optimize_horizon(merged, prefix, reg_c):
    """Run ensemble optimization for a given horizon (1GG or 2GG)."""
    model_cols = sorted([c for c in merged.columns if c.startswith(f"{prefix}_")])
    min_valid = len(merged) * 0.10
    model_cols = [c for c in model_cols if merged[c].notna().sum() >= min_valid]
    model_cols = deduplicate_models(merged, model_cols, prefix)

    if not model_cols:
        return None

    model_info = {}
    for c in model_cols:
        name = c.replace(f"{prefix}_", "")
        stima = ((merged[c].round(0) - 32) * 5 / 9).round(0)
        delta = stima - reg_c
        valid = delta.dropna()
        if len(valid) == 0:
            continue
        verde_pct = (valid.abs() <= 1).sum() / len(valid) * 100
        bias_c = float(valid.mean())
        model_info[name] = {
            "stima": stima,
            "delta": delta,
            "verde": verde_pct,
            "bias_c": bias_c,
            "n": len(valid),
            "mae": float(valid.abs().mean()),
        }

    if not model_info:
        return None

    ranked = sorted(model_info.keys(), key=lambda m: model_info[m]["verde"], reverse=True)
    best_single = ranked[0]
    best_single_v = model_info[best_single]["verde"]

    all_configs = []
    all_configs.append({
        "metodo": "Singolo",
        "n_modelli": 1,
        "modelli": [best_single],
        "verde_c": best_single_v,
        "n_dati": model_info[best_single]["n"],
        "mae": model_info[best_single]["mae"],
        "bc": {},
    })

    max_n = len(ranked)
    for top_n in range(2, max_n + 1):
        top_models = ranked[:top_n]
        stime_raw = pd.DataFrame({m: model_info[m]["stima"] for m in top_models})
        stime_bc = pd.DataFrame()
        for m in top_models:
            bias = model_info[m]["bias_c"]
            stime_bc[m] = model_info[m]["stima"] - round(bias)

        for use_bc, bc_label in [(False, ""), (True, " +BC")]:
            stime = stime_bc if use_bc else stime_raw

            media = stime.mean(axis=1).round(0)
            delta_media = media - reg_c
            valid = delta_media.dropna()
            if len(valid) > 0:
                v = (valid.abs() <= 1).sum() / len(valid) * 100
                bc_dict = {}
                if use_bc:
                    bc_dict = {m: round(model_info[m]["bias_c"]) for m in top_models}
                all_configs.append({
                    "metodo": f"Media{bc_label}",
                    "n_modelli": top_n,
                    "modelli": list(top_models),
                    "verde_c": v,
                    "n_dati": len(valid),
                    "mae": float(valid.abs().mean()),
                    "bc": bc_dict,
                })

            mediana = stime.median(axis=1).round(0)
            delta_med = mediana - reg_c
            valid_m = delta_med.dropna()
            if len(valid_m) > 0:
                v = (valid_m.abs() <= 1).sum() / len(valid_m) * 100
                bc_dict = {}
                if use_bc:
                    bc_dict = {m: round(model_info[m]["bias_c"]) for m in top_models}
                all_configs.append({
                    "metodo": f"Mediana{bc_label}",
                    "n_modelli": top_n,
                    "modelli": list(top_models),
                    "verde_c": v,
                    "n_dati": len(valid_m),
                    "mae": float(valid_m.abs().mean()),
                    "bc": bc_dict,
                })

    all_configs.sort(key=lambda x: x["verde_c"], reverse=True)
    best = all_configs[0]
    top_configs = all_configs[:20]

    return {
        "best": best,
        "best_single": best_single,
        "best_single_v": best_single_v,
        "top_configs": top_configs,
        "model_info": model_info,
        "ranked": ranked,
        "all_configs_count": len(all_configs),
    }


xl_st = pd.ExcelFile(storico_path)

print("Ottimizzazione esaustiva ensemble per citta (1GG e 2GG)...")
print("=" * 100)

city_optimal_1gg = {}
city_optimal_2gg = {}

for city_name in CITIES:
    if city_name not in xl_st.sheet_names:
        continue
    cache_file = os.path.join(CACHE_DIR, f"{city_name}.pkl")
    if not os.path.exists(cache_file):
        continue

    raw_df = pd.read_pickle(cache_file)
    raw_df["time"] = pd.to_datetime(raw_df["time"])
    daily = compute_daily_max(raw_df)
    df_st = pd.read_excel(xl_st, sheet_name=city_name)
    df_st["Data"] = pd.to_datetime(df_st["Data"]).dt.date
    merged = pd.merge(daily, df_st[["Data", "Max_Temperatura_F"]], on="Data", how="inner")

    reg_c = ((merged["Max_Temperatura_F"] - 32) * 5 / 9).round(0)

    print(f"\n{city_name}")
    print("-" * 60)

    # 1GG
    opt_1 = optimize_horizon(merged, "1GG", reg_c)
    if opt_1:
        city_optimal_1gg[city_name] = opt_1
        b1 = opt_1["best"]
        diff1 = b1["verde_c"] - opt_1["best_single_v"]
        print(f"  1GG  Singolo: {opt_1['best_single']:<25} {opt_1['best_single_v']:5.1f}%")
        print(f"       Ottimo:  {b1['metodo']} top{b1['n_modelli']:<5} {'':>14} {b1['verde_c']:5.1f}%  ({diff1:+.1f}%)")

    # 2GG
    opt_2 = optimize_horizon(merged, "2GG", reg_c)
    if opt_2:
        city_optimal_2gg[city_name] = opt_2
        b2 = opt_2["best"]
        diff2 = b2["verde_c"] - opt_2["best_single_v"]
        print(f"  2GG  Singolo: {opt_2['best_single']:<25} {opt_2['best_single_v']:5.1f}%")
        print(f"       Ottimo:  {b2['metodo']} top{b2['n_modelli']:<5} {'':>14} {b2['verde_c']:5.1f}%  ({diff2:+.1f}%)")

    # Confronto
    if opt_1 and opt_2:
        drop = opt_2["best"]["verde_c"] - opt_1["best"]["verde_c"]
        print(f"  >>> Differenza 2GG vs 1GG (ottimo): {drop:+.1f}%")


# ── Scrivi Excel ──
print("\n\nScrittura Excel...")

with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
    # Foglio 1: Confronto 1GG vs 2GG
    def format_bc_info(opt):
        """Build practical BC description for a config."""
        b = opt["best"]
        bc = b.get("bc", {})
        modelli = b["modelli"]
        metodo = b["metodo"]
        if not bc:
            # No BC - list models without corrections
            return ", ".join(modelli), ""
        # BC active - build "model: correction" string
        parts = []
        for m in modelli:
            corr = bc.get(m, 0)
            parts.append(f"{m}: {corr:+d}")
        bc_str = ", ".join(parts)
        # Net correction (average of individual corrections)
        corrections = [bc.get(m, 0) for m in modelli]
        net = sum(corrections) / len(corrections)
        return ", ".join(modelli), bc_str

    summary_data = []
    for city in CITIES:
        row = {"Citta": city}
        if city in city_optimal_1gg:
            o1 = city_optimal_1gg[city]
            b1 = o1["best"]
            row["1GG Metodo"] = b1["metodo"]
            row["1GG N Modelli"] = b1["n_modelli"]
            row["1GG Verde%"] = round(b1["verde_c"], 1)
            row["1GG MAE"] = round(b1["mae"], 2)
            row["1GG Singolo"] = o1["best_single"]
            row["1GG V% Singolo"] = round(o1["best_single_v"], 1)
            row["1GG Miglioram."] = round(b1["verde_c"] - o1["best_single_v"], 1)
            modelli_str, bc_str = format_bc_info(o1)
            row["1GG Modelli"] = modelli_str
            row["1GG Correzioni BC (C)"] = bc_str if bc_str else "Nessuna"
        if city in city_optimal_2gg:
            o2 = city_optimal_2gg[city]
            b2 = o2["best"]
            row["2GG Metodo"] = b2["metodo"]
            row["2GG N Modelli"] = b2["n_modelli"]
            row["2GG Verde%"] = round(b2["verde_c"], 1)
            row["2GG MAE"] = round(b2["mae"], 2)
            row["2GG Singolo"] = o2["best_single"]
            row["2GG V% Singolo"] = round(o2["best_single_v"], 1)
            row["2GG Miglioram."] = round(b2["verde_c"] - o2["best_single_v"], 1)
            modelli_str, bc_str = format_bc_info(o2)
            row["2GG Modelli"] = modelli_str
            row["2GG Correzioni BC (C)"] = bc_str if bc_str else "Nessuna"
        if city in city_optimal_1gg and city in city_optimal_2gg:
            row["Diff 2GG-1GG"] = round(
                city_optimal_2gg[city]["best"]["verde_c"] -
                city_optimal_1gg[city]["best"]["verde_c"], 1)
        if len(row) > 1:
            summary_data.append(row)

    df_summary = pd.DataFrame(summary_data)
    df_summary.to_excel(writer, sheet_name="Confronto 1GG vs 2GG", index=False)

    ws = writer.sheets["Confronto 1GG vs 2GG"]
    for col_idx in range(1, len(df_summary.columns) + 1):
        ws.cell(row=1, column=col_idx).font = bold_white
        ws.cell(row=1, column=col_idx).fill = blue_fill

    # Color verde% and diff columns dynamically
    col_names = list(df_summary.columns)
    for row_idx in range(2, len(df_summary) + 2):
        for col_idx, col_name in enumerate(col_names, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if "Verde%" in col_name and cell.value is not None:
                cell.fill = green_fill
                cell.font = bold
            elif col_name == "Diff 2GG-1GG" and cell.value is not None:
                cell.fill = green_fill if cell.value > 0 else red_fill
                cell.font = bold
            elif "Miglioram" in col_name and cell.value is not None:
                cell.fill = green_fill if cell.value > 0 else red_fill
                cell.font = bold

    for col_idx in range(1, len(df_summary.columns) + 1):
        max_len = max(len(str(ws.cell(row=r, column=col_idx).value or ""))
                      for r in range(1, len(df_summary) + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    # Foglio 2 & 3: Configurazione Ottimale 1GG e 2GG separati
    for horizon, city_opt, sheet_name in [
        ("1GG", city_optimal_1gg, "Ottimale 1GG"),
        ("2GG", city_optimal_2gg, "Ottimale 2GG"),
    ]:
        opt_data = []
        for city in CITIES:
            if city not in city_opt:
                continue
            opt = city_opt[city]
            b = opt["best"]
            opt_data.append({
                "Citta": city,
                "Metodo Ottimale": b["metodo"],
                "N Modelli": b["n_modelli"],
                "Verde% C": round(b["verde_c"], 1),
                "MAE C": round(b["mae"], 2),
                "N Dati": b["n_dati"],
                "Singolo Migliore": opt["best_single"],
                "Verde% Singolo": round(opt["best_single_v"], 1),
                "Miglioramento %": round(b["verde_c"] - opt["best_single_v"], 1),
                "Modelli Usati": ", ".join(b["modelli"]),
            })
        if not opt_data:
            continue
        df_opt = pd.DataFrame(opt_data)
        df_opt.to_excel(writer, sheet_name=sheet_name, index=False)
        ws_o = writer.sheets[sheet_name]
        for col_idx in range(1, len(df_opt.columns) + 1):
            ws_o.cell(row=1, column=col_idx).font = bold_white
            ws_o.cell(row=1, column=col_idx).fill = blue_fill
        for row_idx in range(2, len(df_opt) + 2):
            cell_v = ws_o.cell(row=row_idx, column=4)
            cell_v.fill = green_fill
            cell_v.font = bold
            cell_m = ws_o.cell(row=row_idx, column=9)
            val = cell_m.value
            if val is not None and val > 0:
                cell_m.fill = green_fill
                cell_m.font = bold
            elif val is not None and val <= 0:
                cell_m.fill = red_fill
                cell_m.font = bold
        for col_idx in range(1, len(df_opt.columns) + 1):
            max_len = max(len(str(ws_o.cell(row=r, column=col_idx).value or ""))
                          for r in range(1, len(df_opt) + 2))
            ws_o.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 80)

    # Foglio 4: Bias Modelli (1GG + 2GG)
    print("  Bias per modello...")
    bias_rows = []
    for horizon, city_opt in [("1GG", city_optimal_1gg), ("2GG", city_optimal_2gg)]:
        for city in CITIES:
            if city not in city_opt:
                continue
            opt = city_opt[city]
            for model in opt["ranked"]:
                mi = opt["model_info"][model]
                bias_rows.append({
                    "Orizzonte": horizon,
                    "Citta": city,
                    "Modello": model,
                    "Bias C (media delta)": round(mi["bias_c"], 2),
                    "Correzione (round)": round(round(mi["bias_c"])),
                    "Verde%": round(mi["verde"], 1),
                    "MAE": round(mi["mae"], 2),
                    "N Dati": mi["n"],
                })
    df_bias = pd.DataFrame(bias_rows)
    df_bias.to_excel(writer, sheet_name="Bias Modelli", index=False)
    ws_b = writer.sheets["Bias Modelli"]
    for col_idx in range(1, len(df_bias.columns) + 1):
        ws_b.cell(row=1, column=col_idx).font = bold_white
        ws_b.cell(row=1, column=col_idx).fill = blue_fill

    # Foglio 5 & 6: Top 20 per Citta (1GG e 2GG)
    for horizon, city_opt, sheet_name in [
        ("1GG", city_optimal_1gg, "Top 20 - 1GG"),
        ("2GG", city_optimal_2gg, "Top 20 - 2GG"),
    ]:
        if not city_opt:
            continue
        print(f"  Top configurazioni {horizon}...")
        ws_t = writer.book.create_sheet(sheet_name)
        row_t = 1
        for city in CITIES:
            if city not in city_opt:
                continue
            opt = city_opt[city]

            cell = ws_t.cell(row=row_t, column=1, value=city)
            cell.font = Font(bold=True, size=14)
            ws_t.cell(row=row_t, column=2,
                      value=f"({opt['all_configs_count']} combinazioni testate)")
            row_t += 1

            headers = ["#", "Metodo", "N Modelli", "Verde% C", "MAE C", "N Dati", "Modelli"]
            for j, h in enumerate(headers):
                c = ws_t.cell(row=row_t, column=j + 1, value=h)
                c.font = bold_white
                c.fill = blue_fill
            row_t += 1

            for i, cfg in enumerate(opt["top_configs"]):
                ws_t.cell(row=row_t, column=1, value=i + 1)
                ws_t.cell(row=row_t, column=2, value=cfg["metodo"])
                ws_t.cell(row=row_t, column=3, value=cfg["n_modelli"])
                vc = ws_t.cell(row=row_t, column=4, value=round(cfg["verde_c"], 1))
                vc.number_format = "0.0"
                if i == 0:
                    vc.fill = green_fill
                    vc.font = bold
                mae_c = ws_t.cell(row=row_t, column=5, value=round(cfg["mae"], 2))
                mae_c.number_format = "0.00"
                ws_t.cell(row=row_t, column=6, value=cfg["n_dati"])
                modelli_str = ", ".join(cfg["modelli"])
                ws_t.cell(row=row_t, column=7, value=modelli_str)
                row_t += 1

            row_t += 2

print(f"\nFile salvato: {out_path}")
print()
print("=" * 100)
print("RIEPILOGO CONFRONTO 1GG vs 2GG")
print("=" * 100)
print(f"{'Citta':<15} {'1GG Ottimo':>10} {'2GG Ottimo':>10} {'Diff':>8}")
print("-" * 50)
for city in CITIES:
    v1 = city_optimal_1gg[city]["best"]["verde_c"] if city in city_optimal_1gg else None
    v2 = city_optimal_2gg[city]["best"]["verde_c"] if city in city_optimal_2gg else None
    s1 = f"{v1:5.1f}%" if v1 else "  N/A "
    s2 = f"{v2:5.1f}%" if v2 else "  N/A "
    diff = f"{v2 - v1:+5.1f}%" if v1 and v2 else "  N/A "
    print(f"  {city:<15} {s1:>10} {s2:>10} {diff:>8}")
