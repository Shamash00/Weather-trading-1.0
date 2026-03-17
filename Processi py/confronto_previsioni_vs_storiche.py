import pandas as pd
import os
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill, Font

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
forecast_path = os.path.join(BASE_DIR, "Previsioni FM-15 Tutte le Citta 2021-2025 Daily Max.xlsx")
storico_path = os.path.join(BASE_DIR, "Temperature Storiche FM-15 Tutte le Citta 2021-2025 Daily Max.xlsx")
out_path = os.path.join(BASE_DIR, "Confronto Previsioni vs Storiche Daily Max.xlsx")

# Conditional formatting fills
green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

print("Caricamento file...")
xl_forecast = pd.ExcelFile(forecast_path)
xl_storico = pd.ExcelFile(storico_path)

with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
    for sheet in xl_forecast.sheet_names:
        print(f"\n{sheet}...")
        if sheet not in xl_storico.sheet_names:
            print(f"  ATTENZIONE: {sheet} non presente nei dati storici, saltato.")
            continue

        df_fc = pd.read_excel(xl_forecast, sheet_name=sheet)
        df_st = pd.read_excel(xl_storico, sheet_name=sheet)

        df_fc["Data"] = pd.to_datetime(df_fc["Data"]).dt.date
        df_st["Data"] = pd.to_datetime(df_st["Data"]).dt.date

        # Merge on date (inner join: only days present in both)
        merged = pd.merge(df_fc, df_st, on="Data", how="inner")

        # Build output DataFrame
        out = pd.DataFrame()
        out["Data"] = merged["Data"]

        # Fahrenheit columns
        out["Previsione_F"] = merged["Max_Forecast_F"]
        out["Prev_1GG_F"] = merged["Max_PrevDay1_F"]
        out["Prev_2GG_F"] = merged["Max_PrevDay2_F"]
        out["Registrata_F"] = merged["Max_Temperatura_F"]
        out["Delta_Prev_F"] = merged["Max_Forecast_F"] - merged["Max_Temperatura_F"]
        out["Delta_1GG_F"] = merged["Max_PrevDay1_F"] - merged["Max_Temperatura_F"]
        out["Delta_2GG_F"] = merged["Max_PrevDay2_F"] - merged["Max_Temperatura_F"]

        # Celsius columns
        out["Previsione_C"] = merged["Max_Forecast_C"]
        out["Prev_1GG_C"] = merged["Max_PrevDay1_C"]
        out["Prev_2GG_C"] = merged["Max_PrevDay2_C"]
        out["Registrata_C"] = merged["Max_Temperatura_C"]
        out["Delta_Prev_C"] = (merged["Max_Forecast_C"] - merged["Max_Temperatura_C"]).round(1)
        out["Delta_1GG_C"] = (merged["Max_PrevDay1_C"] - merged["Max_Temperatura_C"]).round(1)
        out["Delta_2GG_C"] = (merged["Max_PrevDay2_C"] - merged["Max_Temperatura_C"]).round(1)

        out.to_excel(writer, sheet_name=sheet, index=False)

        # --- Conditional formatting ---
        ws = writer.sheets[sheet]
        max_row = ws.max_row

        # Fahrenheit delta columns: F, G, H (Delta_Prev_F, Delta_1GG_F, Delta_2GG_F)
        for col in ["F", "G", "H"]:
            rng = f"{col}2:{col}{max_row}"
            # Green: |delta| <= 1
            ws.conditional_formatting.add(rng, FormulaRule(
                formula=[f'AND({col}2<>"", ABS({col}2)<=1)'],
                fill=green_fill, stopIfTrue=True))
            # Yellow: 1 < |delta| < 3
            ws.conditional_formatting.add(rng, FormulaRule(
                formula=[f'AND({col}2<>"", ABS({col}2)>1, ABS({col}2)<3)'],
                fill=yellow_fill, stopIfTrue=True))
            # Red: |delta| >= 3
            ws.conditional_formatting.add(rng, FormulaRule(
                formula=[f'AND({col}2<>"", ABS({col}2)>=3)'],
                fill=red_fill, stopIfTrue=True))

        # Celsius delta columns: M, N, O (Delta_Prev_C, Delta_1GG_C, Delta_2GG_C)
        for col in ["M", "N", "O"]:
            rng = f"{col}2:{col}{max_row}"
            # Green: |delta| <= 1
            ws.conditional_formatting.add(rng, FormulaRule(
                formula=[f'AND({col}2<>"", ABS({col}2)<=1)'],
                fill=green_fill, stopIfTrue=True))
            # Yellow: 1 < |delta| < 2
            ws.conditional_formatting.add(rng, FormulaRule(
                formula=[f'AND({col}2<>"", ABS({col}2)>1, ABS({col}2)<2)'],
                fill=yellow_fill, stopIfTrue=True))
            # Red: |delta| >= 2
            ws.conditional_formatting.add(rng, FormulaRule(
                formula=[f'AND({col}2<>"", ABS({col}2)>=2)'],
                fill=red_fill, stopIfTrue=True))

        # --- Summary counts (only rows where all 6 deltas are present) ---
        delta_f_cols = ["Delta_Prev_F", "Delta_1GG_F", "Delta_2GG_F"]
        delta_c_cols = ["Delta_Prev_C", "Delta_1GG_C", "Delta_2GG_C"]
        all_delta_cols = delta_f_cols + delta_c_cols

        valid_mask = out[all_delta_cols].notna().all(axis=1)
        valid_data = out[valid_mask]
        n_valid = len(valid_data)

        # Excel column indices (1-based)
        col_idx = {
            "Delta_Prev_F": 6, "Delta_1GG_F": 7, "Delta_2GG_F": 8,
            "Delta_Prev_C": 13, "Delta_1GG_C": 14, "Delta_2GG_C": 15,
        }

        summary_row = max_row + 2  # blank row separator
        bold = Font(bold=True)

        # Labels
        for i, label in enumerate(["Verde %", "Giallo %", "Rosso %",
                                    "Delta Medio", "Errore Medio Abs"]):
            cell = ws.cell(row=summary_row + i, column=1, value=label)
            cell.font = bold

        # Fahrenheit percentages
        for col_name in delta_f_cols:
            ci = col_idx[col_name]
            abs_vals = valid_data[col_name].abs()
            g = (abs_vals <= 1).sum() / n_valid * 100
            y = ((abs_vals > 1) & (abs_vals < 3)).sum() / n_valid * 100
            r = (abs_vals >= 3).sum() / n_valid * 100

            cg = ws.cell(row=summary_row, column=ci, value=round(g, 1))
            cg.fill = green_fill
            cg.font = bold
            cg.number_format = '0.0"%"'
            cy = ws.cell(row=summary_row + 1, column=ci, value=round(y, 1))
            cy.fill = yellow_fill
            cy.font = bold
            cy.number_format = '0.0"%"'
            cr = ws.cell(row=summary_row + 2, column=ci, value=round(r, 1))
            cr.fill = red_fill
            cr.font = bold
            cr.number_format = '0.0"%"'

            # Delta medio (bias) e errore medio assoluto (MAE)
            mean_delta = round(float(valid_data[col_name].mean()), 1)
            mae = round(float(valid_data[col_name].abs().mean()), 1)
            cm = ws.cell(row=summary_row + 3, column=ci, value=mean_delta)
            cm.font = bold
            cm.number_format = '0.0'
            ca = ws.cell(row=summary_row + 4, column=ci, value=mae)
            ca.font = bold
            ca.number_format = '0.0'

        # Celsius percentages
        for col_name in delta_c_cols:
            ci = col_idx[col_name]
            abs_vals = valid_data[col_name].abs()
            g = (abs_vals <= 1).sum() / n_valid * 100
            y = ((abs_vals > 1) & (abs_vals < 2)).sum() / n_valid * 100
            r = (abs_vals >= 2).sum() / n_valid * 100

            cg = ws.cell(row=summary_row, column=ci, value=round(g, 1))
            cg.fill = green_fill
            cg.font = bold
            cg.number_format = '0.0"%"'
            cy = ws.cell(row=summary_row + 1, column=ci, value=round(y, 1))
            cy.fill = yellow_fill
            cy.font = bold
            cy.number_format = '0.0"%"'
            cr = ws.cell(row=summary_row + 2, column=ci, value=round(r, 1))
            cr.fill = red_fill
            cr.font = bold
            cr.number_format = '0.0"%"'

            # Delta medio (bias) e errore medio assoluto (MAE)
            mean_delta = round(float(valid_data[col_name].mean()), 1)
            mae = round(float(valid_data[col_name].abs().mean()), 1)
            cm = ws.cell(row=summary_row + 3, column=ci, value=mean_delta)
            cm.font = bold
            cm.number_format = '0.0'
            ca = ws.cell(row=summary_row + 4, column=ci, value=mae)
            ca.font = bold
            ca.number_format = '0.0'

        # Console stats
        print(f"  {len(out)} giorni confrontati, {n_valid} con tutte le previsioni")
        if n_valid > 0:
            abs_f = valid_data["Delta_Prev_F"].abs()
            g_pct = (abs_f <= 1).sum() / n_valid * 100
            print(f"  Previsione ultimo agg. F: {g_pct:.1f}% verde")

print(f"\nFile salvato: {out_path}")
